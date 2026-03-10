import io

from openpyxl import Workbook, load_workbook

from models.fund import Fund, LP
from models.gp_entity import GPEntity
from models.lp_contribution import LPContribution
from routers.funds import (
    MIGRATION_FUND_HEADERS,
    MIGRATION_FUND_HEADER_LABELS,
    MIGRATION_FUND_REQUIRED_HEADERS,
    MIGRATION_LP_HEADERS,
    MIGRATION_LP_HEADER_LABELS,
    MIGRATION_LP_REQUIRED_HEADERS,
    MIGRATION_SHEET_NAME_COLUMN_GUIDE,
    MIGRATION_SHEET_NAME_FUND,
    MIGRATION_SHEET_NAME_GUIDE,
    MIGRATION_SHEET_NAME_LP,
    _migration_header_label,
)
from services.lp_types import LP_TYPE_GP, LP_TYPE_INSTITUTIONAL, LP_TYPE_SPECIAL_GROWTH


def _workbook_bytes(*, fund_rows: list[dict], lp_rows: list[dict], legacy: bool = False) -> bytes:
    workbook = Workbook()

    if legacy:
        fund_sheet = workbook.active
        fund_sheet.title = "Funds"
        fund_sheet.append(["fund_key", "name"])
        fund_sheet.append(["1", "구버전 조합"])
        lp_sheet = workbook.create_sheet("LPs")
        lp_sheet.append(["fund_key", "name"])
        lp_sheet.append(["1", "구버전 LP"])
    else:
        fund_sheet = workbook.active
        fund_sheet.title = MIGRATION_SHEET_NAME_FUND
        fund_sheet.append([
            _migration_header_label(key, MIGRATION_FUND_HEADER_LABELS, MIGRATION_FUND_REQUIRED_HEADERS)
            for key in MIGRATION_FUND_HEADERS
        ])
        for row in fund_rows:
            fund_sheet.append([row.get(key, "") for key in MIGRATION_FUND_HEADERS])

        lp_sheet = workbook.create_sheet(MIGRATION_SHEET_NAME_LP)
        lp_sheet.append([
            _migration_header_label(key, MIGRATION_LP_HEADER_LABELS, MIGRATION_LP_REQUIRED_HEADERS)
            for key in MIGRATION_LP_HEADERS
        ])
        for row in lp_rows:
            lp_sheet.append([row.get(key, "") for key in MIGRATION_LP_HEADERS])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _base_fund_row() -> dict:
    return {
        "fund_key": "1",
        "name": "브이온 벤처성장조합 1호",
        "type": "벤처투자조합",
        "status": "운용 중",
        "formation_date": "2026-01-15",
        "registration_number": "123-45-67890",
        "registration_date": "2026-02-01",
        "fund_manager": "김펀드매니저",
        "co_gp": "",
        "trustee": "OO은행",
        "commitment_total": 10000000000,
        "gp_commitment": 1000000000,
        "contribution_type": "분할",
        "investment_period_end": "2030-01-15",
        "maturity_date": "2034-01-15",
        "dissolution_date": "",
        "mgmt_fee_rate": 2.0,
        "performance_fee_rate": 20.0,
        "hurdle_rate": 6.0,
        "account_number": "110-123-456789",
    }


def _base_lp_rows() -> list[dict]:
    return [
        {
            "fund_key": "1",
            "name": "한국성장금융",
            "type": "기관투자자",
            "commitment": 3000000000,
            "paid_in": 750000000,
            "contact": "02-0000-0000",
            "business_number": "111-22-33333",
            "address": "서울시 강남구",
            "due_date": "2026-02-10",
            "round_no": 1,
            "commitment_ratio_percent": 10,
            "actual_paid_date": "2026-02-10",
            "memo": "1차 납입",
        },
        {
            "fund_key": "1",
            "name": "한국성장금융",
            "type": "",
            "commitment": "",
            "paid_in": "",
            "contact": "",
            "business_number": "",
            "address": "",
            "due_date": "2026-05-10",
            "round_no": "",
            "commitment_ratio_percent": 15,
            "actual_paid_date": "2026-05-12",
            "memo": "2차 납입",
        },
    ]


def _post_validate(client, workbook_bytes: bytes):
    return client.post(
        "/api/funds/migration-validate",
        content=workbook_bytes,
        headers={"content-type": "application/octet-stream"},
    )


def _post_import(client, workbook_bytes: bytes):
    return client.post(
        "/api/funds/migration-import?mode=upsert&sync_address_book=false",
        content=workbook_bytes,
        headers={"content-type": "application/octet-stream"},
    )


def test_download_migration_template_uses_new_korean_sheet_contract(client):
    response = client.get("/api/funds/migration-template")
    assert response.status_code == 200

    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == [
        MIGRATION_SHEET_NAME_FUND,
        MIGRATION_SHEET_NAME_LP,
        MIGRATION_SHEET_NAME_GUIDE,
        MIGRATION_SHEET_NAME_COLUMN_GUIDE,
    ]
    assert "LPContributions" not in workbook.sheetnames

    fund_headers = [cell.value for cell in workbook[MIGRATION_SHEET_NAME_FUND][1]]
    lp_sheet = workbook[MIGRATION_SHEET_NAME_LP]
    assert "조합번호*" in fund_headers
    assert "GP" not in fund_headers
    assert LP_TYPE_INSTITUTIONAL in str(lp_sheet["C2"].value)
    assert LP_TYPE_SPECIAL_GROWTH in str(lp_sheet["C2"].value)
    assert lp_sheet["C3"].value == LP_TYPE_INSTITUTIONAL


def test_validate_migration_requires_single_primary_gp(client):
    workbook_bytes = _workbook_bytes(fund_rows=[_base_fund_row()], lp_rows=_base_lp_rows())
    response = _post_validate(client, workbook_bytes)
    assert response.status_code == 200
    payload = response.json()
    assert payload["success"] is False
    assert any(error["column"] == "대표 GP" for error in payload["errors"])


def test_validate_and_import_migration_with_merged_lp_sheet(client, db_session):
    primary_gp = GPEntity(name="브이온 GP", entity_type="vc", business_number="999-88-77777", representative="대표자", address="서울", is_primary=1)
    db_session.add(primary_gp)
    db_session.commit()

    workbook_bytes = _workbook_bytes(fund_rows=[_base_fund_row()], lp_rows=_base_lp_rows())

    validate_response = _post_validate(client, workbook_bytes)
    assert validate_response.status_code == 200
    validate_payload = validate_response.json()
    assert validate_payload["success"] is True
    assert validate_payload["fund_rows"] == 1
    assert validate_payload["lp_rows"] == 1
    assert validate_payload["contribution_rows"] == 2

    import_response = _post_import(client, workbook_bytes)
    assert import_response.status_code == 200
    import_payload = import_response.json()
    assert import_payload["success"] is True
    assert import_payload["created_funds"] == 1
    assert import_payload["created_lps"] == 1
    assert import_payload["created_contributions"] == 2

    fund = db_session.query(Fund).one()
    assert fund.gp_entity_id == primary_gp.id
    assert fund.gp == primary_gp.name

    lps = db_session.query(LP).order_by(LP.name.asc()).all()
    assert len(lps) == 2
    assert any(lp.type == LP_TYPE_GP and lp.name == primary_gp.name for lp in lps)
    target_lp = next(lp for lp in lps if lp.type != LP_TYPE_GP)
    assert target_lp.name == "한국성장금융"
    assert target_lp.type == LP_TYPE_INSTITUTIONAL
    assert target_lp.paid_in == 750000000

    contributions = (
        db_session.query(LPContribution)
        .filter(LPContribution.lp_id == target_lp.id)
        .order_by(LPContribution.round_no.asc())
        .all()
    )
    assert len(contributions) == 2
    assert [row.round_no for row in contributions] == [1, 2]


def test_validate_rejects_legacy_template_sheets(client, db_session):
    db_session.add(GPEntity(name="브이온 GP", entity_type="vc", is_primary=1))
    db_session.commit()

    response = _post_validate(client, _workbook_bytes(fund_rows=[], lp_rows=[], legacy=True))
    assert response.status_code == 200
    payload = response.json()
    assert payload["success"] is False
    assert any("구버전 템플릿" in error["reason"] for error in payload["errors"])


def test_validate_rejects_conflicting_repeated_lp_base_fields(client, db_session):
    db_session.add(GPEntity(name="브이온 GP", entity_type="vc", is_primary=1))
    db_session.commit()

    lp_rows = _base_lp_rows()
    lp_rows[1]["commitment"] = 3500000000

    response = _post_validate(client, _workbook_bytes(fund_rows=[_base_fund_row()], lp_rows=lp_rows))
    assert response.status_code == 200
    payload = response.json()
    assert payload["success"] is False
    assert any("같은 LP의 기본정보는 모든 행에서 동일" in error["reason"] for error in payload["errors"])


def test_validate_rejects_manual_gp_rows_in_lp_sheet(client, db_session):
    db_session.add(GPEntity(name="브이온 GP", entity_type="vc", is_primary=1))
    db_session.commit()

    lp_rows = _base_lp_rows()
    lp_rows[0]["type"] = "GP"

    response = _post_validate(client, _workbook_bytes(fund_rows=[_base_fund_row()], lp_rows=lp_rows))
    assert response.status_code == 200
    payload = response.json()
    assert payload["success"] is False
    assert any("자동 생성" in error["reason"] for error in payload["errors"])


def test_import_normalizes_legacy_lp_type_alias(client, db_session):
    primary_gp = GPEntity(name="브이온 GP", entity_type="vc", is_primary=1)
    db_session.add(primary_gp)
    db_session.commit()

    lp_rows = _base_lp_rows()
    lp_rows[0]["type"] = "법인"

    import_response = _post_import(client, _workbook_bytes(fund_rows=[_base_fund_row()], lp_rows=lp_rows))
    assert import_response.status_code == 200
    payload = import_response.json()
    assert payload["success"] is True

    imported_lp = next(lp for lp in db_session.query(LP).all() if lp.name == "한국성장금융")
    assert imported_lp.type == LP_TYPE_INSTITUTIONAL
