from datetime import date, datetime, timedelta
import logging
import io
from urllib.parse import quote

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from dateutil.relativedelta import relativedelta
from sqlalchemy import case, func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from database import get_db
from models.fee import FeeConfig
from models.fund import Fund, LP, FundNoticePeriod, FundKeyTerm
from models.gp_entity import GPEntity
from models.lp_address_book import LPAddressBook
from models.lp_contribution import LPContribution
from models.investment import Investment
from models.phase3 import CapitalCall, CapitalCallItem
from models.task import Task
from models.workflow import Workflow
from models.workflow_instance import WorkflowInstance, WorkflowStepInstance
from schemas.fund import (
    FundFormationWorkflowAddRequest,
    FundFormationWorkflowAddResponse,
    FundMigrationImportResponse,
    FundMigrationValidateResponse,
    FundCreate,
    FundKeyTermCreate,
    FundKeyTermResponse,
    FundListItem,
    FundMigrationErrorItem,
    FundNoticePeriodCreate,
    FundNoticePeriodResponse,
    FundOverviewItem,
    FundOverviewResponse,
    FundOverviewTotals,
    FundResponse,
    FundUpdate,
    LPCreate,
    LPResponse,
    LPUpdate,
)
from services.workflow_service import calculate_business_days_before
from services.workflow_service import instantiate_workflow
from services.fund_integrity import recalculate_fund_stats, validate_lp_paid_in_pair
from services.compliance_engine import ComplianceEngine
from services.compliance_rule_engine import ComplianceRuleEngine
from services.lp_types import (
    LP_TYPE_GP,
    LP_TYPE_INSTITUTIONAL,
    MIGRATION_LP_TYPE_OPTIONS,
    coerce_lp_type,
    is_gp_lp_type,
    is_supported_migration_lp_type,
    normalize_lp_type,
)
from services.erp_backbone import (
    backbone_enabled,
    mark_subject_deleted,
    maybe_emit_mutation,
    record_snapshot,
    sync_fund_graph,
    sync_gp_entity_graph,
    sync_lp_graph,
)
from services.proposal_data import sync_fund_history

router = APIRouter(tags=["funds"])
logger = logging.getLogger(__name__)
OVERVIEW_UNIT = 1_000_000
LEGACY_PAID_STEP_TOKENS = (
    "\u7570\uc496\uc604",
    "\u2479\uc5ef",
    "\ub083\ud211",
)
LEGACY_CONFIRM_STEP_TOKENS = (
    "\ubea4\uc524",
)


def _percent_to_decimal(value: float | None) -> float | None:
    if value is None:
        return None
    return round(float(value) / 100, 6)


def _sync_fee_config_from_fund(db: Session, fund: Fund, changed_keys: set[str] | None = None) -> None:
    fee_keys = {"mgmt_fee_rate", "performance_fee_rate", "hurdle_rate"}
    if changed_keys is not None and not (changed_keys & fee_keys):
        return

    if (
        fund.mgmt_fee_rate is None
        and fund.performance_fee_rate is None
        and fund.hurdle_rate is None
    ):
        return

    config = db.query(FeeConfig).filter(FeeConfig.fund_id == fund.id).first()
    if config is None:
        config = FeeConfig(fund_id=fund.id)
        db.add(config)
        db.flush()

    if fund.mgmt_fee_rate is not None:
        config.mgmt_fee_rate = _percent_to_decimal(fund.mgmt_fee_rate)
    if fund.performance_fee_rate is not None:
        config.carry_rate = _percent_to_decimal(fund.performance_fee_rate)
    if fund.hurdle_rate is not None:
        config.hurdle_rate = _percent_to_decimal(fund.hurdle_rate)

FUND_TYPE_OPTIONS = (
    '투자조합',
    '벤처투자조합',
    '신기술투자조합',
    '사모투자합자회사(PEF)',
    '창업투자조합',
    '농림수산식품투자조합',
    '기타',
)

FUND_STATUS_OPTIONS = (
    ("forming", "결성예정"),
    ("active", "운용 중"),
    ("dissolved", "해산"),
    ("liquidated", "청산 완료"),
)

FUND_STATUS_LABEL_BY_CODE = {code: label for code, label in FUND_STATUS_OPTIONS}
FUND_STATUS_CODE_BY_KEY = {
    "결성예정": "forming",
    "운용중": "active",
    "해산": "dissolved",
    "청산완료": "liquidated",
}

MIGRATION_SHEET_NAME_FUND = "조합"
MIGRATION_SHEET_NAME_LP = "조합원(LP)"
MIGRATION_SHEET_NAME_GUIDE = "작성가이드"
MIGRATION_SHEET_NAME_COLUMN_GUIDE = "컬럼가이드"
LEGACY_MIGRATION_SHEET_NAMES = ("Funds", "LPs", "LPContributions")

MIGRATION_FUND_HEADERS = [
    "fund_key",
    "name",
    "type",
    "status",
    "formation_date",
    "registration_number",
    "registration_date",
    "fund_manager",
    "co_gp",
    "trustee",
    "commitment_total",
    "gp_commitment",
    "contribution_type",
    "investment_period_end",
    "maturity_date",
    "dissolution_date",
    "mgmt_fee_rate",
    "performance_fee_rate",
    "hurdle_rate",
    "account_number",
]

MIGRATION_FUND_REQUIRED_HEADERS = {"fund_key", "name", "type", "status", "formation_date"}

MIGRATION_FUND_HEADER_LABELS = {
    "fund_key": "조합번호",
    "name": "조합명",
    "type": "조합유형",
    "status": "상태",
    "formation_date": "결성일",
    "registration_number": "등록번호",
    "registration_date": "등록성립일",
    "fund_manager": "대표펀드매니저",
    "co_gp": "공동GP",
    "trustee": "수탁사",
    "commitment_total": "약정총액",
    "gp_commitment": "GP약정액",
    "contribution_type": "출자방식",
    "investment_period_end": "투자기간종료일",
    "maturity_date": "만기일",
    "dissolution_date": "해산일",
    "mgmt_fee_rate": "관리보수율(%)",
    "performance_fee_rate": "성과보수율(%)",
    "hurdle_rate": "허들레이트(%)",
    "account_number": "계좌번호",
}

MIGRATION_FUND_HEADER_ALIASES = {
    "fund_key": ["조합번호*", "조합번호"],
    "name": ["조합명*", "조합명"],
    "type": ["조합유형*", "조합유형", "fund_type"],
    "status": ["상태*", "상태"],
    "formation_date": ["결성일*", "결성일", "등록성립일", "formation_dt"],
    "registration_number": ["등록번호", "사업자번호"],
    "registration_date": ["등록성립일", "등록일"],
    "fund_manager": ["대표펀드매니저"],
    "co_gp": ["공동GP"],
    "trustee": ["수탁사"],
    "commitment_total": ["약정총액", "약정금액"],
    "gp_commitment": ["GP약정액"],
    "contribution_type": ["출자방식"],
    "investment_period_end": ["투자기간종료일"],
    "maturity_date": ["만기일"],
    "dissolution_date": ["해산일"],
    "mgmt_fee_rate": ["관리보수율(%)"],
    "performance_fee_rate": ["성과보수율(%)"],
    "hurdle_rate": ["허들레이트(%)"],
    "account_number": ["계좌번호"],
}

MIGRATION_LP_HEADERS = [
    "fund_key",
    "name",
    "type",
    "commitment",
    "paid_in",
    "contact",
    "business_number",
    "address",
    "due_date",
    "round_no",
    "commitment_ratio_percent",
    "actual_paid_date",
    "memo",
]

MIGRATION_LP_REQUIRED_HEADERS = {"fund_key", "name"}
MIGRATION_LP_BASE_REQUIRED_HEADERS = {"type", "commitment"}

MIGRATION_LP_HEADER_LABELS = {
    "fund_key": "조합번호",
    "name": "LP명",
    "type": "LP유형",
    "commitment": "약정총액",
    "paid_in": "누적납입액",
    "contact": "연락처",
    "business_number": "사업자번호",
    "address": "주소",
    "due_date": "납입기일",
    "round_no": "회차번호",
    "commitment_ratio_percent": "회차별납입비율(%)",
    "actual_paid_date": "실납입일",
    "memo": "비고",
}

MIGRATION_LP_HEADER_ALIASES = {
    "fund_key": ["조합번호*", "조합번호"],
    "name": ["LP명*", "LP명"],
    "type": ["LP유형", "LP유형*"],
    "commitment": ["약정총액", "약정금액", "commitment_amount"],
    "paid_in": ["누적납입액"],
    "contact": ["연락처"],
    "business_number": ["사업자번호"],
    "address": ["주소"],
    "due_date": ["납입기일"],
    "round_no": ["회차번호"],
    "commitment_ratio_percent": ["회차별납입비율(%)", "납입비율(%)"],
    "actual_paid_date": ["실납입일", "실제입금일"],
    "memo": ["비고", "memo"],
}

MIGRATION_FUND_TYPES = set(FUND_TYPE_OPTIONS)
MIGRATION_FUND_STATUS = {code for code, _ in FUND_STATUS_OPTIONS}
MIGRATION_LP_TYPES = set(MIGRATION_LP_TYPE_OPTIONS)
MIGRATION_LP_TYPE_GUIDE_TEXT = " / ".join(MIGRATION_LP_TYPE_OPTIONS)

FORMING_STATUS_KEYS = {
    'forming',
    'planned',
    '결성예정',
    '결성예정(planned)',
}


def _migration_header_label(
    key: str,
    labels: dict[str, str],
    required_fields: set[str] | None = None,
) -> str:
    label = labels.get(key, key)
    if required_fields and key in required_fields:
        return f"{label}*"
    return label


def _migration_column_label(key: str, labels: dict[str, str]) -> str:
    return labels.get(key, key)


def _parse_migration_status(
    value: object | None,
    *,
    row: int,
    column: str,
    errors: list[FundMigrationErrorItem],
) -> str | None:
    text = _to_str(value)
    if not text:
        errors.append(FundMigrationErrorItem(row=row, column=column, reason="필수값입니다"))
        return None
    normalized = _normalize_status_key(text)
    status = FUND_STATUS_CODE_BY_KEY.get(normalized)
    if status is None:
        supported = ", ".join(label for _, label in FUND_STATUS_OPTIONS)
        errors.append(
            FundMigrationErrorItem(
                row=row,
                column=column,
                reason=f"상태는 {supported} 중 하나여야 합니다",
            )
        )
        return None
    return status


def _is_blank(value: object | None) -> bool:
    return value is None or _to_str(value) == ""

FORMATION_WORKFLOW_NAME_MAP = {
    '고유번호증발급': '고유번호증 발급',
    '고유번호증 발급': '고유번호증 발급',
    '수탁계약체결': '수탁계약 체결',
    '수탁계약 체결': '수탁계약 체결',
    '결성총회개최': '결성총회 개최',
    '결성총회 개최': '결성총회 개최',
    '벤처투자조합등록': '벤처투자조합 등록',
    '벤처투자조합 등록': '벤처투자조합 등록',
    '투자조합등록': '벤처투자조합 등록',
    '투자조합 등록': '벤처투자조합 등록',
    '조합등록': '벤처투자조합 등록',
    '조합 등록': '벤처투자조합 등록',
}
FORMATION_SLOT_MEMO_PREFIX = 'formation_slot='


def _run_compliance_rule_checks(
    *,
    db: Session,
    fund_id: int,
    trigger_source: str,
    trigger_source_id: int | None = None,
) -> None:
    try:
        checks = ComplianceRuleEngine().evaluate_all(
            fund_id=fund_id,
            db=db,
            trigger_type="event",
            trigger_source=trigger_source,
            trigger_source_id=trigger_source_id,
        )
        violations = [check for check in checks if check.result in {"fail", "error"}]
        if violations:
            logger.warning(
                "compliance rule violations: fund_id=%s trigger_source=%s violations=%s",
                fund_id,
                trigger_source,
                len(violations),
            )
    except Exception as exc:  # noqa: BLE001 - rule checks must not break main flow
        db.rollback()
        logger.warning(
            "compliance rule engine failed: fund_id=%s trigger_source=%s trigger_source_id=%s error=%s",
            fund_id,
            trigger_source,
            trigger_source_id,
            exc,
        )


def _to_str(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_date_cell(
    value: object | None,
    row: int,
    column: str,
    errors: list[FundMigrationErrorItem],
) -> date | None:
    if value is None or _to_str(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            errors.append(
                FundMigrationErrorItem(
                    row=row,
                    column=column,
                    reason="날짜는 YYYY-MM-DD 형식이어야 합니다",
                )
            )
            return None
    errors.append(
        FundMigrationErrorItem(
            row=row,
            column=column,
            reason="날짜 형식이 올바르지 않습니다",
        )
    )
    return None


def _parse_number_cell(
    value: object | None,
    row: int,
    column: str,
    errors: list[FundMigrationErrorItem],
) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    try:
        if isinstance(value, str):
            parsed = float(value.replace(",", "").strip())
        else:
            parsed = float(value)
    except (TypeError, ValueError):
        errors.append(
            FundMigrationErrorItem(
                row=row,
                column=column,
                reason="숫자 형식이어야 합니다",
            )
        )
        return None
    if parsed < 0:
        errors.append(
            FundMigrationErrorItem(
                row=row,
                column=column,
                reason="0 이상이어야 합니다",
            )
        )
        return None
    return parsed


def _parse_int_cell(
    value: object | None,
    row: int,
    column: str,
    errors: list[FundMigrationErrorItem],
) -> int | None:
    number = _parse_number_cell(value, row, column, errors)
    if number is None:
        return None
    if int(number) != number:
        errors.append(
            FundMigrationErrorItem(
                row=row,
                column=column,
                reason="정수만 입력해야 합니다",
            )
        )
        return None
    return int(number)


def _normalize_lookup_text(value: str | None) -> str:
    return (value or "").strip().casefold()


def _find_existing_lp_by_identifiers(
    db: Session,
    fund_id: int,
    *,
    name: str,
    business_number: str | None,
) -> LP | None:
    normalized_name = _normalize_lp_text(name)
    normalized_business_number = _normalize_lp_optional(business_number)
    if normalized_business_number:
        existing = (
            db.query(LP)
            .filter(
                LP.fund_id == fund_id,
                LP.business_number == normalized_business_number,
            )
            .first()
        )
        if existing:
            return existing
    return (
        db.query(LP)
        .filter(
            LP.fund_id == fund_id,
            LP.name == normalized_name,
        )
        .first()
    )


def _is_initial_import_locked(db: Session, fund: Fund) -> bool:
    if fund.initial_import_completed_at is not None:
        return True
    existing_contribution = (
        db.query(LPContribution.id)
        .filter(LPContribution.fund_id == fund.id)
        .first()
    )
    return existing_contribution is not None


def _migration_primary_gp_validation_errors(db: Session) -> list[FundMigrationErrorItem]:
    primary_rows = (
        db.query(GPEntity)
        .filter(GPEntity.is_primary == 1)
        .order_by(GPEntity.id.asc())
        .all()
    )
    if len(primary_rows) == 1:
        return []
    if not primary_rows:
        reason = "대표 GP 법인이 없습니다. 고유계정에서 대표 GP를 먼저 등록하세요."
    else:
        reason = "대표 GP 법인이 여러 개입니다. 고유계정에서 대표 GP를 1개만 남겨주세요."
    return [FundMigrationErrorItem(row=1, column="대표 GP", reason=reason)]


def _require_migration_primary_gp(db: Session) -> GPEntity:
    primary_rows = (
        db.query(GPEntity)
        .filter(GPEntity.is_primary == 1)
        .order_by(GPEntity.id.asc())
        .all()
    )
    if len(primary_rows) != 1:
        raise HTTPException(status_code=400, detail="대표 GP 법인 설정이 올바르지 않습니다")
    return primary_rows[0]


def _text_conflicts(current: str | None, incoming: str | None) -> bool:
    return current is not None and incoming is not None and current != incoming


def _number_conflicts(current: float | None, incoming: float | None) -> bool:
    if current is None or incoming is None:
        return False
    return abs(float(current) - float(incoming)) > 0.5


def _derive_contribution_round_numbers(
    contribution_rows: list[dict],
    errors: list[FundMigrationErrorItem],
) -> None:
    grouped: dict[tuple[str, str], list[dict]] = {}
    for row in contribution_rows:
        grouped.setdefault(row["_lp_group_key"], []).append(row)

    for group_key, rows in grouped.items():
        used_rounds: set[int] = set()
        for row in rows:
            explicit_round = row.get("round_no")
            if explicit_round is None:
                continue
            if explicit_round in used_rounds:
                errors.append(
                    FundMigrationErrorItem(
                        row=row["__row"],
                        column="round_no",
                        reason="같은 LP에서 회차번호가 중복되었습니다",
                    )
                )
            else:
                used_rounds.add(explicit_round)

        next_round = 1
        for row in sorted(
            rows,
            key=lambda item: (
                item.get("due_date") or date.max,
                int(item.get("__row") or 0),
            ),
        ):
            assigned_round = row.get("round_no")
            if assigned_round is None:
                while next_round in used_rounds:
                    next_round += 1
                assigned_round = next_round
                used_rounds.add(assigned_round)
                next_round += 1
            else:
                next_round = max(next_round, int(assigned_round) + 1)
            row["round_no"] = int(assigned_round)


def _legacy_migration_template_error() -> list[FundMigrationErrorItem]:
    return [
        FundMigrationErrorItem(
            row=1,
            column="sheet",
            reason="구버전 템플릿입니다. 조합관리에서 새 템플릿을 다시 다운로드하세요.",
        )
    ]


def _missing_sheet_errors(sheet_names: tuple[str, ...]) -> list[FundMigrationErrorItem]:
    return [
        FundMigrationErrorItem(row=1, column="sheet", reason=f"필수 시트 누락: {name}")
        for name in sheet_names
    ]


def _merge_lp_text_field(
    group: dict,
    *,
    row_no: int,
    field: str,
    value: str | None,
    column: str,
    errors: list[FundMigrationErrorItem],
) -> None:
    if value is None:
        return
    if _text_conflicts(group.get(field), value):
        errors.append(
            FundMigrationErrorItem(
                row=row_no,
                column=column,
                reason="같은 LP의 기본정보는 모든 행에서 동일해야 합니다",
            )
        )
        return
    if group.get(field) is None:
        group[field] = value


def _merge_lp_number_field(
    group: dict,
    *,
    row_no: int,
    field: str,
    value: float | None,
    column: str,
    errors: list[FundMigrationErrorItem],
) -> None:
    if value is None:
        return
    if _number_conflicts(group.get(field), value):
        errors.append(
            FundMigrationErrorItem(
                row=row_no,
                column=column,
                reason="같은 LP의 기본정보는 모든 행에서 동일해야 합니다",
            )
        )
        return
    if group.get(field) is None:
        group[field] = value


def _lp_group_key(fund_key: str, name: str) -> tuple[str, str]:
    return fund_key, _normalize_lookup_text(name)


def _contribution_row_has_input(row: dict) -> bool:
    return any(
        not _is_blank(row.get(field))
        for field in ("due_date", "round_no", "commitment_ratio_percent", "actual_paid_date", "memo")
    )


def _count_contribution_rows(raw_rows: list[dict]) -> int:
    return sum(1 for row in raw_rows if _contribution_row_has_input(row))


def _build_lp_lookup_key(row: dict) -> tuple[str, str]:
    token = _normalize_lookup_text(row.get("business_number") or row["name"])
    return row["fund_key"], token


def _find_existing_lp_for_group(db: Session, existing_fund: Fund | None, row: dict) -> LP | None:
    if existing_fund is None:
        return None
    return _find_existing_lp_by_identifiers(
        db,
        existing_fund.id,
        name=row["name"],
        business_number=row.get("business_number"),
    )


def _validate_lp_group_required_fields(
    *,
    row_no: int,
    row: dict,
    existing_lp: LP | None,
    errors: list[FundMigrationErrorItem],
) -> None:
    if row.get("type") is None:
        if existing_lp is not None and existing_lp.type:
            row["type"] = normalize_lp_type(existing_lp.type) or existing_lp.type
        else:
            errors.append(FundMigrationErrorItem(row=row_no, column="type", reason="LP유형은 첫 행에서 입력해야 합니다"))
    else:
        normalized_type = normalize_lp_type(row["type"])
        if is_gp_lp_type(normalized_type):
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column="type",
                    reason="업무집행조합원(GP)은 LP 시트에서 입력하지 않습니다. 대표 GP 법인에서 자동 생성됩니다",
                )
            )
        elif not is_supported_migration_lp_type(normalized_type):
            errors.append(FundMigrationErrorItem(row=row_no, column="type", reason="지원하지 않는 LP 유형입니다"))
        else:
            row["type"] = normalized_type

    if row.get("commitment") is None:
        existing_commitment = float(existing_lp.commitment or 0) if existing_lp is not None else 0.0
        if existing_commitment > 0:
            row["commitment"] = existing_commitment
        else:
            errors.append(FundMigrationErrorItem(row=row_no, column="commitment", reason="약정총액은 첫 행에서 입력해야 합니다"))


def _read_sheet_rows(
    worksheet,
    headers: list[str],
    sheet_name: str,
    header_aliases: dict[str, list[str]] | None = None,
) -> tuple[list[dict], list[FundMigrationErrorItem]]:
    errors: list[FundMigrationErrorItem] = []
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
    header_map = {_to_str(header_values[idx]): idx for idx in range(len(header_values))}
    header_aliases = header_aliases or {}
    column_index_map: dict[str, int] = {}

    for header in headers:
        candidate_labels = [header, *header_aliases.get(header, [])]
        matched_index = next((header_map[label] for label in candidate_labels if label in header_map), None)
        if matched_index is None:
            errors.append(
                FundMigrationErrorItem(
                    row=1,
                    column=header,
                    reason=f"{sheet_name} 시트에 필수 컬럼이 없습니다",
                )
            )
            continue
        column_index_map[header] = matched_index

    if errors:
        return [], errors

    rows: list[dict] = []
    for row_no, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        first_value = _to_str(row_values[0] if row_values else None)
        if first_value.startswith("#"):
            continue
        if all(value is None or _to_str(value) == "" for value in row_values):
            continue

        item: dict = {"__row": row_no}
        for header in headers:
            idx = column_index_map[header]
            item[header] = row_values[idx] if idx < len(row_values) else None
        rows.append(item)
    return rows, errors


def _validate_migration_rows(
    raw_funds: list[dict],
    raw_lps: list[dict],
    raw_contributions: list[dict],
    db: Session,
) -> tuple[FundMigrationValidateResponse, list[dict], list[dict], list[dict]]:
    errors: list[FundMigrationErrorItem] = []
    warnings: list[FundMigrationErrorItem] = []
    fund_rows: list[dict] = []
    lp_rows: list[dict] = []
    contribution_rows: list[dict] = []
    fund_keys: set[str] = set()
    existing_fund_map: dict[str, Fund | None] = {}
    lp_business_keys: set[tuple[str, str]] = set()
    lp_name_keys: set[tuple[str, str]] = set()

    if not raw_funds:
        errors.append(
            FundMigrationErrorItem(
                row=2,
                column='fund_key',
                reason='Funds 시트에 데이터가 없습니다',
            )
        )

    for row in raw_funds:
        row_no = int(row.get('__row', 0))
        fund_key = _to_str(row.get('fund_key'))
        name = _to_str(row.get('name'))
        fund_type = _to_str(row.get('type'))
        status = _to_str(row.get('status')) or 'active'

        if not fund_key:
            errors.append(FundMigrationErrorItem(row=row_no, column='fund_key', reason='필수값입니다'))
        elif fund_key in fund_keys:
            errors.append(FundMigrationErrorItem(row=row_no, column='fund_key', reason='중복 fund_key 입니다'))
        else:
            fund_keys.add(fund_key)

        if not name:
            errors.append(FundMigrationErrorItem(row=row_no, column='name', reason='필수값입니다'))
        if not fund_type:
            errors.append(FundMigrationErrorItem(row=row_no, column='type', reason='필수값입니다'))
        elif fund_type not in MIGRATION_FUND_TYPES:
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column='type',
                    reason='지원하지 않는 조합 유형입니다',
                )
            )

        if status not in MIGRATION_FUND_STATUS:
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column='status',
                    reason='status는 forming/active/dissolved/liquidated 중 하나여야 합니다',
                )
            )

        parsed = {
            '__row': row_no,
            'fund_key': fund_key,
            'name': name,
            'type': fund_type,
            'status': status,
            'formation_date': _parse_date_cell(row.get('formation_date'), row_no, 'formation_date', errors),
            'registration_number': _to_str(row.get('registration_number')) or None,
            'registration_date': _parse_date_cell(row.get('registration_date'), row_no, 'registration_date', errors),
            'gp': _to_str(row.get('gp')) or None,
            'fund_manager': _to_str(row.get('fund_manager')) or None,
            'co_gp': _to_str(row.get('co_gp')) or None,
            'trustee': _to_str(row.get('trustee')) or None,
            'commitment_total': _parse_number_cell(row.get('commitment_total'), row_no, 'commitment_total', errors),
            'gp_commitment': _parse_number_cell(row.get('gp_commitment'), row_no, 'gp_commitment', errors),
            'contribution_type': _to_str(row.get('contribution_type')) or None,
            'investment_period_end': _parse_date_cell(row.get('investment_period_end'), row_no, 'investment_period_end', errors),
            'maturity_date': _parse_date_cell(row.get('maturity_date'), row_no, 'maturity_date', errors),
            'dissolution_date': _parse_date_cell(row.get('dissolution_date'), row_no, 'dissolution_date', errors),
            'mgmt_fee_rate': _parse_number_cell(row.get('mgmt_fee_rate'), row_no, 'mgmt_fee_rate', errors),
            'performance_fee_rate': _parse_number_cell(row.get('performance_fee_rate'), row_no, 'performance_fee_rate', errors),
            'hurdle_rate': _parse_number_cell(row.get('hurdle_rate'), row_no, 'hurdle_rate', errors),
            'account_number': _to_str(row.get('account_number')) or None,
        }
        existing_fund = _find_existing_fund(db, parsed)
        existing_fund_map[fund_key] = existing_fund
        if existing_fund is not None and _is_initial_import_locked(db, existing_fund):
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column='fund_key',
                    reason='초기 세팅 import가 이미 완료된 조합입니다. 이후 수정은 ERP 화면에서 진행하세요.',
                )
            )
        fund_rows.append(parsed)

    for row in raw_lps:
        row_no = int(row.get('__row', 0))
        fund_key = _to_str(row.get('fund_key'))
        name = _to_str(row.get('name'))
        lp_type = _to_str(row.get('type'))
        normalized_lp_type = normalize_lp_type(lp_type)

        if not fund_key:
            errors.append(FundMigrationErrorItem(row=row_no, column='fund_key', reason='필수값입니다'))
        elif fund_key not in fund_keys:
            errors.append(FundMigrationErrorItem(row=row_no, column='fund_key', reason='Funds 시트에 없는 fund_key 입니다'))

        if not name:
            errors.append(FundMigrationErrorItem(row=row_no, column='name', reason='필수값입니다'))
        if not lp_type:
            errors.append(FundMigrationErrorItem(row=row_no, column='type', reason='필수값입니다'))
        elif is_gp_lp_type(normalized_lp_type):
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column='type',
                    reason='업무집행조합원(GP)은 LP 시트에서 입력하지 않습니다. 대표 GP 법인에서 자동 생성됩니다',
                )
            )
        elif not is_supported_migration_lp_type(normalized_lp_type):
            errors.append(FundMigrationErrorItem(row=row_no, column='type', reason='지원하지 않는 LP 유형입니다'))

        parsed_lp = {
            '__row': row_no,
            'fund_key': fund_key,
            'name': name,
            'type': normalized_lp_type,
            'commitment': _parse_number_cell(row.get('commitment'), row_no, 'commitment', errors),
            'paid_in': _parse_number_cell(row.get('paid_in'), row_no, 'paid_in', errors),
            'contact': _to_str(row.get('contact')) or None,
            'business_number': _to_str(row.get('business_number')) or None,
            'address': _to_str(row.get('address')) or None,
        }

        business_number = _normalize_lookup_text(parsed_lp['business_number'])
        if business_number:
            business_key = (fund_key, business_number)
            if business_key in lp_business_keys:
                errors.append(
                    FundMigrationErrorItem(
                        row=row_no,
                        column='business_number',
                        reason='같은 조합 내 LP 사업자번호가 중복되었습니다',
                    )
                )
            else:
                lp_business_keys.add(business_key)

        name_key = (fund_key, _normalize_lookup_text(name))
        if name_key in lp_name_keys:
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column='name',
                    reason='같은 조합 내 LP명이 중복되었습니다',
                )
            )
        else:
            lp_name_keys.add(name_key)

        lp_rows.append(parsed_lp)

    contribution_amounts_by_lp: dict[tuple[str, str], float] = {}
    contribution_commitment_by_lp: dict[tuple[str, str], float] = {}
    contribution_seen_keys: set[tuple[str, str, date | None, int | None]] = set()

    for row in raw_contributions:
        row_no = int(row.get('__row', 0))
        fund_key = _to_str(row.get('fund_key'))
        lp_name = _to_str(row.get('lp_name'))
        lp_business_number = _to_str(row.get('lp_business_number')) or None
        due_date = _parse_date_cell(row.get('due_date'), row_no, 'due_date', errors)
        actual_paid_date = _parse_date_cell(row.get('actual_paid_date'), row_no, 'actual_paid_date', errors)
        ratio = _parse_number_cell(row.get('commitment_ratio_percent'), row_no, 'commitment_ratio_percent', errors)
        round_no = _parse_int_cell(row.get('round_no'), row_no, 'round_no', errors)

        if not fund_key:
            errors.append(FundMigrationErrorItem(row=row_no, column='fund_key', reason='필수값입니다'))
        elif fund_key not in fund_keys:
            errors.append(FundMigrationErrorItem(row=row_no, column='fund_key', reason='Funds 시트에 없는 fund_key 입니다'))

        if not lp_name:
            errors.append(FundMigrationErrorItem(row=row_no, column='lp_name', reason='필수값입니다'))
        if due_date is None:
            errors.append(FundMigrationErrorItem(row=row_no, column='due_date', reason='필수값입니다'))
        if ratio is None:
            errors.append(FundMigrationErrorItem(row=row_no, column='commitment_ratio_percent', reason='필수값입니다'))
        elif ratio <= 0 or ratio > 100:
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column='commitment_ratio_percent',
                    reason='0 초과 100 이하의 비율만 입력할 수 있습니다',
                )
            )
        if round_no is not None and round_no <= 0:
            errors.append(FundMigrationErrorItem(row=row_no, column='round_no', reason='1 이상의 정수만 입력할 수 있습니다'))

        contribution_rows.append(
            {
                '__row': row_no,
                'fund_key': fund_key,
                'lp_name': lp_name,
                'lp_business_number': lp_business_number,
                'due_date': due_date,
                'round_no': round_no,
                'commitment_ratio_percent': ratio,
                'actual_paid_date': actual_paid_date,
                'memo': _to_str(row.get('memo')) or None,
            }
        )

    _derive_contribution_round_numbers(contribution_rows, errors)

    for row in contribution_rows:
        fund_key = row['fund_key']
        existing_fund = existing_fund_map.get(fund_key)
        matched_lp = _find_lp_row_match(
            lp_rows,
            fund_key=fund_key,
            lp_name=row['lp_name'],
            lp_business_number=row['lp_business_number'],
        )
        if matched_lp is None and existing_fund is not None:
            matched_lp = _find_existing_lp_by_identifiers(
                db,
                existing_fund.id,
                name=row['lp_name'],
                business_number=row['lp_business_number'],
            )

        if matched_lp is None:
            errors.append(
                FundMigrationErrorItem(
                    row=row['__row'],
                    column='lp_name',
                    reason='매칭되는 LP를 찾을 수 없습니다',
                )
            )
            continue

        commitment = matched_lp['commitment'] if isinstance(matched_lp, dict) else float(matched_lp.commitment or 0)
        if commitment is None or float(commitment) <= 0:
            errors.append(
                FundMigrationErrorItem(
                    row=row['__row'],
                    column='commitment_ratio_percent',
                    reason='약정액이 있어야 납입비율을 금액으로 환산할 수 있습니다',
                )
            )
            continue

        lookup_token = _normalize_lookup_text(row['lp_business_number'] or row['lp_name'])
        duplicate_key = (fund_key, lookup_token, row['due_date'], row['round_no'])
        if duplicate_key in contribution_seen_keys:
            errors.append(
                FundMigrationErrorItem(
                    row=row['__row'],
                    column='round_no',
                    reason='같은 조합/LP/기일/회차 조합이 중복되었습니다',
                )
            )
        else:
            contribution_seen_keys.add(duplicate_key)

        amount = round(float(commitment) * float(row['commitment_ratio_percent'] or 0) / 100)
        row['amount'] = amount
        row['actual_paid_date'] = row['actual_paid_date'] or row['due_date']
        contribution_key = (fund_key, lookup_token)
        contribution_amounts_by_lp[contribution_key] = contribution_amounts_by_lp.get(contribution_key, 0.0) + float(amount)
        contribution_commitment_by_lp[contribution_key] = float(commitment)

    for contribution_key, total_amount in contribution_amounts_by_lp.items():
        commitment = contribution_commitment_by_lp.get(contribution_key, 0.0)
        if commitment > 0 and total_amount - commitment > 0.5:
            fund_key, lookup_token = contribution_key
            matching_row = next(
                (
                    row
                    for row in contribution_rows
                    if row['fund_key'] == fund_key
                    and _normalize_lookup_text(row['lp_business_number'] or row['lp_name']) == lookup_token
                ),
                None,
            )
            if matching_row is not None:
                errors.append(
                    FundMigrationErrorItem(
                        row=matching_row['__row'],
                        column='commitment_ratio_percent',
                        reason='회차별 납입비율 합계가 약정액을 초과합니다',
                    )
                )

    for lp_row in lp_rows:
        contribution_key = (
            lp_row['fund_key'],
            _normalize_lookup_text(lp_row['business_number'] or lp_row['name']),
        )
        if contribution_key not in contribution_amounts_by_lp:
            continue
        if lp_row.get('paid_in') is None:
            continue
        contribution_total = contribution_amounts_by_lp[contribution_key]
        if abs(float(lp_row['paid_in']) - float(contribution_total)) > 0.5:
            warnings.append(
                FundMigrationErrorItem(
                    row=lp_row['__row'],
                    column='paid_in',
                    reason='LPContributions 시트가 있으면 누적납입액은 납입이력 합계가 우선 적용됩니다',
                )
            )

    validation = FundMigrationValidateResponse(
        success=len(errors) == 0,
        fund_rows=len(fund_rows),
        lp_rows=len(lp_rows),
        contribution_rows=len(contribution_rows),
        warnings=warnings,
        errors=errors,
    )
    return validation, fund_rows, lp_rows, contribution_rows


def _parse_and_validate_migration(
    file_content: bytes,
    db: Session,
) -> tuple[FundMigrationValidateResponse, list[dict], list[dict], list[dict]]:
    return _parse_and_validate_migration_v2(file_content, db)

    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail='openpyxl not installed') from exc

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail='유효한 엑셀 파일을 읽을 수 없습니다') from exc

    missing_sheets = [name for name in ('Funds', 'LPs') if name not in workbook.sheetnames]
    if missing_sheets:
        errors = [
            FundMigrationErrorItem(row=1, column='sheet', reason=f'필수 시트 누락: {name}')
            for name in missing_sheets
        ]
        return (
            FundMigrationValidateResponse(success=False, fund_rows=0, lp_rows=0, contribution_rows=0, errors=errors),
            [],
            [],
            [],
        )

    fund_sheet = workbook['Funds']
    lp_sheet = workbook['LPs']
    contribution_sheet = workbook['LPContributions'] if 'LPContributions' in workbook.sheetnames else None
    raw_funds, fund_sheet_errors = _read_sheet_rows(
        fund_sheet,
        MIGRATION_FUND_HEADERS,
        'Funds',
        header_aliases=MIGRATION_FUND_HEADER_ALIASES,
    )
    raw_lps, lp_sheet_errors = _read_sheet_rows(
        lp_sheet,
        MIGRATION_LP_HEADERS,
        'LPs',
        header_aliases=MIGRATION_LP_HEADER_ALIASES,
    )
    raw_contributions: list[dict] = []
    contribution_sheet_errors: list[FundMigrationErrorItem] = []
    if contribution_sheet is not None:
        raw_contributions, contribution_sheet_errors = _read_sheet_rows(
            contribution_sheet,
            MIGRATION_CONTRIBUTION_HEADERS,
            'LPContributions',
            header_aliases=MIGRATION_CONTRIBUTION_HEADER_ALIASES,
        )

    if fund_sheet_errors or lp_sheet_errors or contribution_sheet_errors:
        errors = [*fund_sheet_errors, *lp_sheet_errors, *contribution_sheet_errors]
        return (
            FundMigrationValidateResponse(success=False, fund_rows=0, lp_rows=0, contribution_rows=0, errors=errors),
            [],
            [],
            [],
        )

    return _validate_migration_rows(raw_funds, raw_lps, raw_contributions, db)


def _validate_migration_rows_v2(
    raw_funds: list[dict],
    raw_lps: list[dict],
    db: Session,
) -> tuple[FundMigrationValidateResponse, list[dict], list[dict], list[dict]]:
    errors: list[FundMigrationErrorItem] = []
    warnings: list[FundMigrationErrorItem] = []
    fund_rows: list[dict] = []
    lp_rows: list[dict] = []
    contribution_rows: list[dict] = []
    fund_keys: set[str] = set()
    existing_fund_map: dict[str, Fund | None] = {}
    lp_groups: dict[tuple[str, str], dict] = {}
    lp_business_keys: dict[tuple[str, str], tuple[str, str]] = {}

    fund_key_column = _migration_column_label("fund_key", MIGRATION_FUND_HEADER_LABELS)
    fund_name_column = _migration_column_label("name", MIGRATION_FUND_HEADER_LABELS)
    fund_type_column = _migration_column_label("type", MIGRATION_FUND_HEADER_LABELS)
    fund_status_column = _migration_column_label("status", MIGRATION_FUND_HEADER_LABELS)
    formation_date_column = _migration_column_label("formation_date", MIGRATION_FUND_HEADER_LABELS)

    if not raw_funds:
        errors.append(
            FundMigrationErrorItem(
                row=2,
                column=fund_key_column,
                reason=f"{MIGRATION_SHEET_NAME_FUND} 시트에 데이터가 없습니다",
            )
        )

    for row in raw_funds:
        row_no = int(row.get("__row", 0))
        fund_key = _to_str(row.get("fund_key"))
        name = _to_str(row.get("name"))
        fund_type = _to_str(row.get("type"))
        status = _parse_migration_status(
            row.get("status"),
            row=row_no,
            column=fund_status_column,
            errors=errors,
        )

        if not fund_key:
            errors.append(FundMigrationErrorItem(row=row_no, column=fund_key_column, reason="필수값입니다"))
        elif fund_key in fund_keys:
            errors.append(FundMigrationErrorItem(row=row_no, column=fund_key_column, reason="중복된 조합번호입니다"))
        else:
            fund_keys.add(fund_key)

        if not name:
            errors.append(FundMigrationErrorItem(row=row_no, column=fund_name_column, reason="필수값입니다"))
        if not fund_type:
            errors.append(FundMigrationErrorItem(row=row_no, column=fund_type_column, reason="필수값입니다"))
        elif fund_type not in MIGRATION_FUND_TYPES:
            errors.append(FundMigrationErrorItem(row=row_no, column=fund_type_column, reason="지원하지 않는 조합 유형입니다"))

        if _is_blank(row.get("formation_date")):
            errors.append(FundMigrationErrorItem(row=row_no, column=formation_date_column, reason="필수값입니다"))

        parsed = {
            "__row": row_no,
            "fund_key": fund_key,
            "name": name,
            "type": fund_type,
            "status": status,
            "formation_date": _parse_date_cell(row.get("formation_date"), row_no, formation_date_column, errors),
            "registration_number": _to_str(row.get("registration_number")) or None,
            "registration_date": _parse_date_cell(
                row.get("registration_date"),
                row_no,
                _migration_column_label("registration_date", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "fund_manager": _to_str(row.get("fund_manager")) or None,
            "co_gp": _to_str(row.get("co_gp")) or None,
            "trustee": _to_str(row.get("trustee")) or None,
            "commitment_total": _parse_number_cell(
                row.get("commitment_total"),
                row_no,
                _migration_column_label("commitment_total", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "gp_commitment": _parse_number_cell(
                row.get("gp_commitment"),
                row_no,
                _migration_column_label("gp_commitment", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "contribution_type": _to_str(row.get("contribution_type")) or None,
            "investment_period_end": _parse_date_cell(
                row.get("investment_period_end"),
                row_no,
                _migration_column_label("investment_period_end", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "maturity_date": _parse_date_cell(
                row.get("maturity_date"),
                row_no,
                _migration_column_label("maturity_date", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "dissolution_date": _parse_date_cell(
                row.get("dissolution_date"),
                row_no,
                _migration_column_label("dissolution_date", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "mgmt_fee_rate": _parse_number_cell(
                row.get("mgmt_fee_rate"),
                row_no,
                _migration_column_label("mgmt_fee_rate", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "performance_fee_rate": _parse_number_cell(
                row.get("performance_fee_rate"),
                row_no,
                _migration_column_label("performance_fee_rate", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "hurdle_rate": _parse_number_cell(
                row.get("hurdle_rate"),
                row_no,
                _migration_column_label("hurdle_rate", MIGRATION_FUND_HEADER_LABELS),
                errors,
            ),
            "account_number": _to_str(row.get("account_number")) or None,
        }
        existing_fund = _find_existing_fund(db, parsed)
        existing_fund_map[fund_key] = existing_fund
        if existing_fund is not None and _is_initial_import_locked(db, existing_fund):
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column=fund_key_column,
                    reason="초기 세팅 import가 이미 완료된 조합입니다. 이후 수정은 ERP 화면에서 진행하세요.",
                )
            )
        fund_rows.append(parsed)

    for row in raw_lps:
        row_no = int(row.get("__row", 0))
        fund_key = _to_str(row.get("fund_key"))
        name = _to_str(row.get("name"))
        type_value = _to_str(row.get("type")) or None
        commitment = _parse_number_cell(row.get("commitment"), row_no, _migration_column_label("commitment", MIGRATION_LP_HEADER_LABELS), errors)
        paid_in = _parse_number_cell(row.get("paid_in"), row_no, _migration_column_label("paid_in", MIGRATION_LP_HEADER_LABELS), errors)
        business_number = _to_str(row.get("business_number")) or None
        due_date = _parse_date_cell(row.get("due_date"), row_no, _migration_column_label("due_date", MIGRATION_LP_HEADER_LABELS), errors)
        round_no = _parse_int_cell(row.get("round_no"), row_no, _migration_column_label("round_no", MIGRATION_LP_HEADER_LABELS), errors)
        commitment_ratio = _parse_number_cell(
            row.get("commitment_ratio_percent"),
            row_no,
            _migration_column_label("commitment_ratio_percent", MIGRATION_LP_HEADER_LABELS),
            errors,
        )
        actual_paid_date = _parse_date_cell(
            row.get("actual_paid_date"),
            row_no,
            _migration_column_label("actual_paid_date", MIGRATION_LP_HEADER_LABELS),
            errors,
        )

        if round_no is not None and round_no <= 0:
            errors.append(FundMigrationErrorItem(row=row_no, column=_migration_column_label("round_no", MIGRATION_LP_HEADER_LABELS), reason="1 이상의 정수만 입력할 수 있습니다"))
        if commitment_ratio is not None and (commitment_ratio <= 0 or commitment_ratio > 100):
            errors.append(
                FundMigrationErrorItem(
                    row=row_no,
                    column=_migration_column_label("commitment_ratio_percent", MIGRATION_LP_HEADER_LABELS),
                    reason="0 초과 100 이하의 비율만 입력할 수 있습니다",
                )
            )

        if not fund_key:
            errors.append(FundMigrationErrorItem(row=row_no, column=_migration_column_label("fund_key", MIGRATION_LP_HEADER_LABELS), reason="필수값입니다"))
        elif fund_key not in fund_keys:
            errors.append(FundMigrationErrorItem(row=row_no, column=_migration_column_label("fund_key", MIGRATION_LP_HEADER_LABELS), reason=f"{MIGRATION_SHEET_NAME_FUND} 시트에 없는 조합번호입니다"))
        if not name:
            errors.append(FundMigrationErrorItem(row=row_no, column=_migration_column_label("name", MIGRATION_LP_HEADER_LABELS), reason="필수값입니다"))

        if not fund_key or not name:
            continue

        group_key = _lp_group_key(fund_key, name)
        group = lp_groups.setdefault(
            group_key,
            {
                "__row": row_no,
                "fund_key": fund_key,
                "name": name,
                "type": None,
                "commitment": None,
                "paid_in": None,
                "contact": None,
                "business_number": None,
                "address": None,
            },
        )
        _merge_lp_text_field(group, row_no=row_no, field="type", value=type_value, column=_migration_column_label("type", MIGRATION_LP_HEADER_LABELS), errors=errors)
        _merge_lp_number_field(group, row_no=row_no, field="commitment", value=commitment, column=_migration_column_label("commitment", MIGRATION_LP_HEADER_LABELS), errors=errors)
        _merge_lp_number_field(group, row_no=row_no, field="paid_in", value=paid_in, column=_migration_column_label("paid_in", MIGRATION_LP_HEADER_LABELS), errors=errors)
        _merge_lp_text_field(group, row_no=row_no, field="contact", value=_to_str(row.get("contact")) or None, column=_migration_column_label("contact", MIGRATION_LP_HEADER_LABELS), errors=errors)
        _merge_lp_text_field(group, row_no=row_no, field="business_number", value=business_number, column=_migration_column_label("business_number", MIGRATION_LP_HEADER_LABELS), errors=errors)
        _merge_lp_text_field(group, row_no=row_no, field="address", value=_to_str(row.get("address")) or None, column=_migration_column_label("address", MIGRATION_LP_HEADER_LABELS), errors=errors)

        if _contribution_row_has_input(row):
            if _is_blank(row.get("due_date")):
                errors.append(FundMigrationErrorItem(row=row_no, column=_migration_column_label("due_date", MIGRATION_LP_HEADER_LABELS), reason="납입회차를 입력한 행은 납입기일이 필요합니다"))
            if _is_blank(row.get("commitment_ratio_percent")):
                errors.append(FundMigrationErrorItem(row=row_no, column=_migration_column_label("commitment_ratio_percent", MIGRATION_LP_HEADER_LABELS), reason="납입회차를 입력한 행은 회차별납입비율이 필요합니다"))
            contribution_rows.append(
                {
                    "__row": row_no,
                    "fund_key": fund_key,
                    "lp_name": name,
                    "lp_business_number": business_number,
                    "due_date": due_date,
                    "round_no": round_no,
                    "commitment_ratio_percent": commitment_ratio,
                    "actual_paid_date": actual_paid_date,
                    "memo": _to_str(row.get("memo")) or None,
                    "_lp_group_key": group_key,
                }
            )

    for group_key, lp_row in lp_groups.items():
        existing_lp = _find_existing_lp_for_group(db, existing_fund_map.get(lp_row["fund_key"]), lp_row)
        if existing_lp is not None:
            if lp_row.get("type") is None and existing_lp.type:
                lp_row["type"] = normalize_lp_type(existing_lp.type) or existing_lp.type
            if lp_row.get("commitment") is None and existing_lp.commitment is not None:
                lp_row["commitment"] = float(existing_lp.commitment)
            if lp_row.get("paid_in") is None and existing_lp.paid_in is not None:
                lp_row["paid_in"] = float(existing_lp.paid_in)
            if lp_row.get("contact") is None and existing_lp.contact:
                lp_row["contact"] = existing_lp.contact
            if lp_row.get("business_number") is None and existing_lp.business_number:
                lp_row["business_number"] = existing_lp.business_number
            if lp_row.get("address") is None and existing_lp.address:
                lp_row["address"] = existing_lp.address

        _validate_lp_group_required_fields(row_no=lp_row["__row"], row=lp_row, existing_lp=existing_lp, errors=errors)

        if lp_row.get("commitment") is not None and lp_row.get("paid_in") is not None and float(lp_row["paid_in"]) > float(lp_row["commitment"]):
            errors.append(FundMigrationErrorItem(row=lp_row["__row"], column=_migration_column_label("paid_in", MIGRATION_LP_HEADER_LABELS), reason="누적납입액은 약정총액을 초과할 수 없습니다"))

        business_token = _normalize_lookup_text(lp_row.get("business_number"))
        if business_token:
            business_key = (lp_row["fund_key"], business_token)
            previous_group_key = lp_business_keys.get(business_key)
            if previous_group_key and previous_group_key != group_key:
                errors.append(FundMigrationErrorItem(row=lp_row["__row"], column=_migration_column_label("business_number", MIGRATION_LP_HEADER_LABELS), reason="같은 조합 내 LP 사업자번호가 중복되었습니다"))
            else:
                lp_business_keys[business_key] = group_key

        lp_rows.append(lp_row)

    _derive_contribution_round_numbers(contribution_rows, errors)
    lp_group_map = {group_key: row for group_key, row in lp_groups.items()}
    contribution_amounts_by_lp: dict[tuple[str, str], float] = {}
    contribution_commitment_by_lp: dict[tuple[str, str], float] = {}
    contribution_seen_keys: set[tuple[str, str, date | None, int | None]] = set()

    for row in contribution_rows:
        lp_row = lp_group_map.get(row["_lp_group_key"])
        if lp_row is None:
            errors.append(FundMigrationErrorItem(row=row["__row"], column=_migration_column_label("name", MIGRATION_LP_HEADER_LABELS), reason="매칭되는 LP를 찾을 수 없습니다"))
            continue

        row["lp_name"] = lp_row["name"]
        row["lp_business_number"] = lp_row.get("business_number")
        row["actual_paid_date"] = row["actual_paid_date"] or row["due_date"]
        commitment = lp_row.get("commitment")
        if commitment is None or float(commitment) <= 0:
            errors.append(FundMigrationErrorItem(row=row["__row"], column=_migration_column_label("commitment_ratio_percent", MIGRATION_LP_HEADER_LABELS), reason="약정총액이 있어야 납입회차를 금액으로 환산할 수 있습니다"))
            continue

        lookup_token = _normalize_lookup_text(row["lp_business_number"] or row["lp_name"])
        duplicate_key = (row["fund_key"], lookup_token, row["due_date"], row["round_no"])
        if duplicate_key in contribution_seen_keys:
            errors.append(FundMigrationErrorItem(row=row["__row"], column=_migration_column_label("round_no", MIGRATION_LP_HEADER_LABELS), reason="같은 조합/LP/납입기일/회차 조합이 중복되었습니다"))
        else:
            contribution_seen_keys.add(duplicate_key)

        amount = round(float(commitment) * float(row["commitment_ratio_percent"] or 0) / 100)
        row["amount"] = float(amount)
        contribution_key = _build_lp_lookup_key(lp_row)
        contribution_amounts_by_lp[contribution_key] = contribution_amounts_by_lp.get(contribution_key, 0.0) + float(amount)
        contribution_commitment_by_lp[contribution_key] = float(commitment)

    for contribution_key, total_amount in contribution_amounts_by_lp.items():
        commitment = contribution_commitment_by_lp.get(contribution_key, 0.0)
        if commitment > 0 and total_amount - commitment > 0.5:
            fund_key, lookup_token = contribution_key
            matching_row = next(
                (
                    row
                    for row in contribution_rows
                    if row["fund_key"] == fund_key
                    and _normalize_lookup_text(row["lp_business_number"] or row["lp_name"]) == lookup_token
                ),
                None,
            )
            if matching_row is not None:
                errors.append(FundMigrationErrorItem(row=matching_row["__row"], column=_migration_column_label("commitment_ratio_percent", MIGRATION_LP_HEADER_LABELS), reason="납입회차 합계가 약정총액을 초과합니다"))

    for lp_row in lp_rows:
        contribution_key = _build_lp_lookup_key(lp_row)
        if contribution_key not in contribution_amounts_by_lp or lp_row.get("paid_in") is None:
            continue
        contribution_total = contribution_amounts_by_lp[contribution_key]
        if abs(float(lp_row["paid_in"]) - float(contribution_total)) > 0.5:
            warnings.append(
                FundMigrationErrorItem(
                    row=lp_row["__row"],
                    column=_migration_column_label("paid_in", MIGRATION_LP_HEADER_LABELS),
                    reason="납입회차가 있으면 누적납입액은 납입회차 합계가 우선 적용됩니다",
                )
            )

    validation = FundMigrationValidateResponse(
        success=len(errors) == 0,
        fund_rows=len(fund_rows),
        lp_rows=len(lp_rows),
        contribution_rows=len(contribution_rows),
        warnings=warnings,
        errors=errors,
    )
    return validation, fund_rows, lp_rows, contribution_rows


def _parse_and_validate_migration_v2(
    file_content: bytes,
    db: Session,
) -> tuple[FundMigrationValidateResponse, list[dict], list[dict], list[dict]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from exc

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="유효한 엑셀 파일을 읽을 수 없습니다") from exc

    if any(sheet_name in workbook.sheetnames for sheet_name in LEGACY_MIGRATION_SHEET_NAMES):
        errors = _legacy_migration_template_error()
        return FundMigrationValidateResponse(success=False, fund_rows=0, lp_rows=0, contribution_rows=0, errors=errors), [], [], []

    missing_sheets = [
        sheet_name
        for sheet_name in (MIGRATION_SHEET_NAME_FUND, MIGRATION_SHEET_NAME_LP)
        if sheet_name not in workbook.sheetnames
    ]
    if missing_sheets:
        errors = _missing_sheet_errors(tuple(missing_sheets))
        return FundMigrationValidateResponse(success=False, fund_rows=0, lp_rows=0, contribution_rows=0, errors=errors), [], [], []

    raw_funds, fund_sheet_errors = _read_sheet_rows(
        workbook[MIGRATION_SHEET_NAME_FUND],
        MIGRATION_FUND_HEADERS,
        MIGRATION_SHEET_NAME_FUND,
        header_aliases=MIGRATION_FUND_HEADER_ALIASES,
    )
    raw_lps, lp_sheet_errors = _read_sheet_rows(
        workbook[MIGRATION_SHEET_NAME_LP],
        MIGRATION_LP_HEADERS,
        MIGRATION_SHEET_NAME_LP,
        header_aliases=MIGRATION_LP_HEADER_ALIASES,
    )
    primary_gp_errors = _migration_primary_gp_validation_errors(db)
    if fund_sheet_errors or lp_sheet_errors:
        errors = [*primary_gp_errors, *fund_sheet_errors, *lp_sheet_errors]
        return FundMigrationValidateResponse(success=False, fund_rows=0, lp_rows=0, contribution_rows=0, errors=errors), [], [], []

    validation, fund_rows, lp_rows, contribution_rows = _validate_migration_rows_v2(raw_funds, raw_lps, db)
    if primary_gp_errors:
        validation = FundMigrationValidateResponse(
            success=False,
            fund_rows=validation.fund_rows,
            lp_rows=validation.lp_rows,
            contribution_rows=validation.contribution_rows,
            warnings=validation.warnings,
            errors=[*primary_gp_errors, *validation.errors],
        )
    return validation, fund_rows, lp_rows, contribution_rows


def calculate_paid_in_as_of(
    db: Session,
    fund_id: int,
    reference_date: date,
    fallback_gp_commitment: float | None = None,
) -> tuple[float, float]:
    lps = db.query(LP).filter(LP.fund_id == fund_id).all()
    gp_lp_ids = {
        lp.id
        for lp in lps
        if is_gp_lp_type(lp.type)
    }

    has_call_items = (
        db.query(CapitalCallItem.id)
        .join(CapitalCall, CapitalCall.id == CapitalCallItem.capital_call_id)
        .filter(CapitalCall.fund_id == fund_id)
        .first()
    )

    if has_call_items is None:
        total_paid_in = sum(float(lp.paid_in or 0) for lp in lps)
        gp_paid_in = sum(float(lp.paid_in or 0) for lp in lps if lp.id in gp_lp_ids)
        if not gp_lp_ids and fallback_gp_commitment is not None:
            gp_paid_in = float(fallback_gp_commitment or 0)
        return round(total_paid_in, 2), round(gp_paid_in, 2)

    paid_items = (
        db.query(CapitalCallItem.lp_id, CapitalCallItem.amount)
        .join(CapitalCall, CapitalCall.id == CapitalCallItem.capital_call_id)
        .filter(
            CapitalCall.fund_id == fund_id,
            CapitalCallItem.paid == 1,
            CapitalCallItem.paid_date.isnot(None),
            CapitalCallItem.paid_date <= reference_date,
        )
        .all()
    )
    total_paid_in = sum(float(item.amount or 0) for item in paid_items)
    gp_paid_in = sum(float(item.amount or 0) for item in paid_items if item.lp_id in gp_lp_ids)
    if not gp_lp_ids and fallback_gp_commitment is not None:
        gp_paid_in = float(fallback_gp_commitment or 0)
    return round(total_paid_in, 2), round(gp_paid_in, 2)


def calculate_lp_paid_in_from_calls(db: Session, fund_id: int, lp_id: int) -> tuple[bool, int]:
    item_count = (
        db.query(func.count(CapitalCallItem.id))
        .join(CapitalCall, CapitalCall.id == CapitalCallItem.capital_call_id)
        .filter(
            CapitalCall.fund_id == fund_id,
            CapitalCallItem.lp_id == lp_id,
        )
        .scalar()
    )
    if int(item_count or 0) <= 0:
        return False, 0

    paid_in_total = (
        db.query(func.coalesce(func.sum(CapitalCallItem.amount), 0))
        .join(CapitalCall, CapitalCall.id == CapitalCallItem.capital_call_id)
        .filter(
            CapitalCall.fund_id == fund_id,
            CapitalCallItem.lp_id == lp_id,
            CapitalCallItem.paid == 1,
        )
        .scalar()
    )
    return True, int(paid_in_total or 0)


def _normalize_status_key(status: str | None) -> str:
    if not status:
        return ""
    return "".join(status.split()).lower()


def _is_forming_status(status: str | None) -> bool:
    return _normalize_status_key(status) in FORMING_STATUS_KEYS


def _normalize_template_identifier(value: str | None) -> str:
    return "".join((value or "").split()).replace("?", "").lower()


def _template_has_formation_paid_in_step(template: Workflow) -> bool:
    for step in template.steps or []:
        normalized = _normalize_template_identifier(step.name)
        has_paid_token = any(
            token in normalized
            for token in (
                "출자",
                "납입",
                "입금",
                "납부",
                "payment",
                *LEGACY_PAID_STEP_TOKENS,
            )
        )
        has_confirm_token = any(
            token in normalized
            for token in (
                "확인",
                "confirm",
                "check",
                *LEGACY_CONFIRM_STEP_TOKENS,
            )
        )
        if has_paid_token and has_confirm_token:
            return True
    return False


def _resolve_formation_template_name(template_category_or_name: str) -> str | None:
    if not template_category_or_name:
        return None
    normalized = _normalize_template_identifier(template_category_or_name)
    if not normalized:
        return None
    return FORMATION_WORKFLOW_NAME_MAP.get(normalized) or FORMATION_WORKFLOW_NAME_MAP.get(
        template_category_or_name
    )


def _find_formation_template(
    db: Session,
    template_category_or_name: str,
) -> Workflow | None:
    target_name = _resolve_formation_template_name(template_category_or_name)
    if not target_name:
        return None
    return (
        db.query(Workflow)
        .filter(Workflow.name == target_name)
        .order_by(Workflow.id.asc())
        .first()
    )


def _find_formation_template_by_id(db: Session, template_id: int | None) -> Workflow | None:
    if template_id is None:
        return None
    return db.get(Workflow, template_id)


def _find_existing_instance_by_formation_slot(
    db: Session,
    *,
    fund_id: int,
    formation_slot_name: str,
) -> WorkflowInstance | None:
    slot_token = _normalize_template_identifier(formation_slot_name)
    if not slot_token:
        return None

    current = (
        db.query(WorkflowInstance)
        .filter(
            WorkflowInstance.fund_id == fund_id,
            func.lower(func.coalesce(WorkflowInstance.status, "")) != "cancelled",
            func.lower(func.coalesce(WorkflowInstance.memo, "")).like(
                f"%{FORMATION_SLOT_MEMO_PREFIX}{slot_token}%"
            ),
        )
        .order_by(WorkflowInstance.id.desc())
        .first()
    )
    if current:
        return current

    # Backward compatibility for old instances created before slot tokens were stored in memo.
    return (
        db.query(WorkflowInstance)
        .join(Workflow, Workflow.id == WorkflowInstance.workflow_id)
        .filter(
            WorkflowInstance.fund_id == fund_id,
            func.lower(func.coalesce(WorkflowInstance.status, "")) != "cancelled",
            Workflow.name == formation_slot_name,
        )
        .order_by(WorkflowInstance.id.desc())
        .first()
    )


def _to_lp_commitment_amount(value: float | None) -> int | None:
    if value is None:
        return None
    parsed = float(value)
    if parsed < 0:
        return 0
    return int(round(parsed))


def _resolve_gp_entity(
    db: Session,
    *,
    gp_entity_id: int | None,
    gp_name: str | None,
) -> GPEntity | None:
    if gp_entity_id is not None:
        entity = db.get(GPEntity, gp_entity_id)
        if entity is None:
            raise HTTPException(status_code=400, detail="선택한 GP 법인을 찾을 수 없습니다")
        return entity

    normalized_name = (gp_name or "").strip()
    if not normalized_name:
        return None

    return (
        db.query(GPEntity)
        .filter(func.lower(GPEntity.name) == normalized_name.lower())
        .order_by(GPEntity.is_primary.desc(), GPEntity.id.asc())
        .first()
    )


def _ensure_gp_lp_record(
    db: Session,
    *,
    fund: Fund,
    gp_name: str | None,
    gp_commitment: float | None,
    gp_entity: GPEntity | None = None,
) -> None:
    normalized_name = (gp_name or "").strip()
    if not normalized_name:
        return

    existing_lp = (
        db.query(LP)
        .filter(
            LP.fund_id == fund.id,
            or_(
                LP.name == normalized_name,
                func.lower(func.coalesce(LP.type, "")) == "gp",
            ),
        )
        .order_by(LP.id.asc())
        .first()
    )
    matched_gp_entity = gp_entity or _resolve_gp_entity(db, gp_entity_id=None, gp_name=normalized_name)

    commitment_amount = _to_lp_commitment_amount(gp_commitment)
    if existing_lp:
        existing_lp.name = normalized_name
        existing_lp.type = LP_TYPE_GP
        if commitment_amount is not None:
            existing_lp.commitment = commitment_amount
        if matched_gp_entity:
            existing_lp.business_number = matched_gp_entity.business_number
            existing_lp.contact = matched_gp_entity.representative
            existing_lp.address = matched_gp_entity.address
        return

    db.add(
        LP(
            fund_id=fund.id,
            name=normalized_name,
            type=LP_TYPE_GP,
            commitment=commitment_amount,
            paid_in=0,
            contact=(matched_gp_entity.representative if matched_gp_entity else None),
            business_number=(matched_gp_entity.business_number if matched_gp_entity else None),
            address=(matched_gp_entity.address if matched_gp_entity else None),
        )
    )


def to_overview_unit(value: float | None) -> float | None:
    if value is None:
        return None
    return round(float(value) / OVERVIEW_UNIT, 2)


def format_remaining_period(maturity_date: date | None, ref_date: date) -> str:
    if maturity_date is None:
        return "-"
    if ref_date > maturity_date:
        return "만기 경과"

    delta = relativedelta(maturity_date, ref_date)
    parts: list[str] = []
    if delta.years:
        parts.append(f"{delta.years}년")
    if delta.months:
        parts.append(f"{delta.months}개월")

    # For less than 1 month remaining, show days instead of 0개월.
    if not parts and delta.days:
        parts.append(f"{delta.days}일")

    return "".join(parts) if parts else "만기일"


def build_fund_overview(
    db: Session,
    ref_date: date,
) -> tuple[list[FundOverviewItem], FundOverviewTotals]:
    funds = (
        db.query(Fund)
        .filter(
            or_(
                Fund.formation_date.is_(None),
                Fund.formation_date <= ref_date,
            )
        )
        .order_by(Fund.id.asc())
        .all()
    )

    investment_rows = (
        db.query(
            Investment.fund_id.label("fund_id"),
            func.coalesce(func.sum(Investment.amount), 0).label("total_invested"),
            func.coalesce(
                func.sum(case((Investment.status == "active", Investment.amount), else_=0)),
                0,
            ).label("investment_assets"),
            func.count(func.distinct(Investment.company_id)).label("company_count"),
        )
        .filter(
            Investment.investment_date.isnot(None),
            Investment.investment_date <= ref_date,
        )
        .group_by(Investment.fund_id)
        .all()
    )
    investment_by_fund = {
        int(row.fund_id): {
            "total_invested": float(row.total_invested or 0),
            "investment_assets": float(row.investment_assets or 0),
            "company_count": int(row.company_count or 0),
        }
        for row in investment_rows
    }
    fund_ids = [fund.id for fund in funds]
    active_workflow_by_fund: dict[int, int] = {}
    pending_task_by_fund: dict[int, int] = {}
    if fund_ids:
        active_workflow_by_fund = {
            int(row.fund_id): int(row.count or 0)
            for row in (
                db.query(
                    WorkflowInstance.fund_id.label("fund_id"),
                    func.count(WorkflowInstance.id).label("count"),
                )
                .filter(
                    WorkflowInstance.fund_id.in_(fund_ids),
                    WorkflowInstance.status == "active",
                )
                .group_by(WorkflowInstance.fund_id)
                .all()
            )
        }
        pending_task_by_fund = {
            int(row.fund_id): int(row.count or 0)
            for row in (
                db.query(
                    Task.fund_id.label("fund_id"),
                    func.count(Task.id).label("count"),
                )
                .filter(
                    Task.fund_id.in_(fund_ids),
                    Task.status.in_(["pending", "in_progress"]),
                )
                .group_by(Task.fund_id)
                .all()
            )
        }

    totals = {
        "commitment_total": 0.0,
        "total_paid_in": 0.0,
        "gp_commitment": 0.0,
        "total_invested": 0.0,
        "uninvested": 0.0,
        "investment_assets": 0.0,
        "company_count": 0,
        "active_workflow_count": 0,
        "pending_task_count": 0,
    }

    items: list[FundOverviewItem] = []
    for index, fund in enumerate(funds, start=1):
        agg = investment_by_fund.get(fund.id, {})
        total_paid_in, gp_paid_in = calculate_paid_in_as_of(
            db,
            fund.id,
            ref_date,
            fallback_gp_commitment=fund.gp_commitment,
        )
        total_invested = float(agg.get("total_invested", 0.0))
        investment_assets = float(agg.get("investment_assets", total_invested))
        company_count = int(agg.get("company_count", 0))

        paid_in_ratio = (
            round((total_paid_in / fund.commitment_total) * 100, 2)
            if fund.commitment_total
            else None
        )
        uninvested = (
            round((fund.commitment_total or 0) - total_invested, 2)
            if fund.commitment_total is not None
            else None
        )

        progress: float | None = None
        if fund.formation_date and fund.investment_period_end:
            total_days = (fund.investment_period_end - fund.formation_date).days
            if total_days > 0:
                elapsed_days = (min(ref_date, fund.investment_period_end) - fund.formation_date).days
                elapsed_days = max(0, elapsed_days)
                progress = round(max(0, min(100, (elapsed_days / total_days) * 100)), 2)
            else:
                progress = 0.0

        remaining = format_remaining_period(fund.maturity_date, ref_date)

        item = FundOverviewItem(
            no=index,
            id=fund.id,
            name=fund.name,
            fund_type=fund.type,
            fund_manager=fund.fund_manager,
            formation_date=fund.formation_date.isoformat() if fund.formation_date else None,
            registration_date=fund.registration_date.isoformat() if fund.registration_date else None,
            investment_period_end=fund.investment_period_end.isoformat() if fund.investment_period_end else None,
            investment_period_progress=progress,
            maturity_date=fund.maturity_date.isoformat() if fund.maturity_date else None,
            commitment_total=to_overview_unit(fund.commitment_total),
            total_paid_in=to_overview_unit(total_paid_in),
            paid_in_ratio=paid_in_ratio,
            gp_commitment=to_overview_unit(gp_paid_in),
            total_invested=to_overview_unit(total_invested),
            uninvested=to_overview_unit(uninvested),
            investment_assets=to_overview_unit(investment_assets),
            company_count=company_count,
            active_workflow_count=active_workflow_by_fund.get(fund.id, 0),
            pending_task_count=pending_task_by_fund.get(fund.id, 0),
            hurdle_rate=fund.hurdle_rate,
            remaining_period=remaining,
        )
        items.append(item)

        totals["commitment_total"] += float(fund.commitment_total or 0)
        totals["total_paid_in"] += float(total_paid_in or 0)
        totals["gp_commitment"] += float(gp_paid_in or 0)
        totals["total_invested"] += float(total_invested or 0)
        totals["uninvested"] += float(uninvested or 0)
        totals["investment_assets"] += float(investment_assets or 0)
        totals["active_workflow_count"] += int(active_workflow_by_fund.get(fund.id, 0))
        totals["pending_task_count"] += int(pending_task_by_fund.get(fund.id, 0))
    if fund_ids:
        totals["company_count"] = int(
            db.query(func.count(func.distinct(Investment.company_id)))
            .filter(
                Investment.fund_id.in_(fund_ids),
                Investment.investment_date.isnot(None),
                Investment.investment_date <= ref_date,
            )
            .scalar()
            or 0
        )

    return (
        items,
        FundOverviewTotals(
            commitment_total=to_overview_unit(totals["commitment_total"]) or 0,
            total_paid_in=to_overview_unit(totals["total_paid_in"]) or 0,
            gp_commitment=to_overview_unit(totals["gp_commitment"]) or 0,
            total_invested=to_overview_unit(totals["total_invested"]) or 0,
            uninvested=to_overview_unit(totals["uninvested"]) or 0,
            investment_assets=to_overview_unit(totals["investment_assets"]) or 0,
            company_count=totals["company_count"],
            active_workflow_count=totals["active_workflow_count"],
            pending_task_count=totals["pending_task_count"],
        ),
    )


@router.get("/api/funds", response_model=list[FundListItem])
def list_funds(db: Session = Depends(get_db)):
    funds = db.query(Fund).order_by(Fund.id.desc()).all()
    investment_counts = {
        int(fund_id): int(count)
        for fund_id, count in (
            db.query(Investment.fund_id, func.count(Investment.id))
            .group_by(Investment.fund_id)
            .all()
        )
    }
    paid_in_totals = {
        int(fund_id): float(total or 0)
        for fund_id, total in (
            db.query(LP.fund_id, func.coalesce(func.sum(LP.paid_in), 0))
            .group_by(LP.fund_id)
            .all()
        )
    }
    return [
        FundListItem(
            id=f.id,
            name=f.name,
            type=f.type,
            business_number=f.business_number,
            status=f.status,
            gp_entity_id=f.gp_entity_id,
            gp=f.gp,
            formation_date=f.formation_date,
            registration_number=f.registration_number,
            registration_date=f.registration_date,
            maturity_date=f.maturity_date,
            dissolution_date=f.dissolution_date,
            commitment_total=f.commitment_total,
            aum=f.aum,
            paid_in_total=paid_in_totals.get(f.id, 0),
            lp_count=len(f.lps),
            investment_count=investment_counts.get(f.id, 0),
        )
        for f in funds
    ]


@router.get("/api/funds/overview", response_model=FundOverviewResponse)
def fund_overview(reference_date: date | None = None, db: Session = Depends(get_db)):
    ref_date = reference_date or date.today()
    items, totals = build_fund_overview(db, ref_date)
    return FundOverviewResponse(reference_date=ref_date.isoformat(), funds=items, totals=totals)


@router.get("/api/funds/overview/export")
def export_fund_overview(reference_date: date | None = None, db: Session = Depends(get_db)):
    ref_date = reference_date or date.today()
    overview_items, totals = build_fund_overview(db, ref_date)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "조합비교표"

    headers = [
        "NO",
        "조합명",
        "조합 구분",
        "대표 펀드매니저",
        "등록(성립)일",
        "투자기간 종료일",
        "투자기간 경과율",
        "청산시기(예정)",
        "약정총액",
        "납입총액",
        "납입비율",
        "GP출자금",
        "투자총액",
        "미투자액",
        "투자자산",
        "투자업체수",
        "기준수익률(규약)",
        "잔존기간",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, item in enumerate(overview_items, start=2):
        values = [
            item.no,
            item.name,
            item.fund_type,
            item.fund_manager,
            item.formation_date,
            item.investment_period_end,
            item.investment_period_progress,
            item.maturity_date,
            item.commitment_total,
            item.total_paid_in,
            item.paid_in_ratio,
            item.gp_commitment,
            item.total_invested,
            item.uninvested,
            item.investment_assets,
            item.company_count,
            item.hurdle_rate,
            item.remaining_period,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in (9, 10, 12, 13, 14, 15):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (7, 11, 17):
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 16:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    totals_row = len(overview_items) + 2
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=8)
    ws.cell(row=totals_row, column=1, value="합계")
    ws.cell(row=totals_row, column=9, value=totals.commitment_total)
    ws.cell(row=totals_row, column=10, value=totals.total_paid_in)
    ws.cell(row=totals_row, column=12, value=totals.gp_commitment)
    ws.cell(row=totals_row, column=13, value=totals.total_invested)
    ws.cell(row=totals_row, column=14, value=totals.uninvested)
    ws.cell(row=totals_row, column=15, value=totals.investment_assets)
    ws.cell(row=totals_row, column=16, value=totals.company_count)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = Font(bold=True)
        if col_idx in (9, 10, 12, 13, 14, 15):
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 16:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for idx in range(1, len(headers) + 1):
        letter = openpyxl.utils.get_column_letter(idx)
        max_len = 0
        for row in range(1, totals_row + 1):
            value = ws.cell(row=row, column=idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[letter].width = min(max_len + 4, 28)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"fund_overview_{ref_date.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _download_migration_template_v2():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from exc

    workbook = openpyxl.Workbook()
    header_fill = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
    helper_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    guide_fill = PatternFill(start_color="EEF6FF", end_color="EEF6FF", fill_type="solid")
    header_font = Font(bold=True, color="1F2937")
    helper_font = Font(color="6B7280", italic=True, size=10)

    def style_sheet(sheet, header_count: int, freeze_panes: str, helper_rows: tuple[int, ...]) -> None:
        for column_index in range(1, header_count + 1):
            header_cell = sheet.cell(row=1, column=column_index)
            header_cell.fill = header_fill
            header_cell.font = header_font
            for helper_row in helper_rows:
                helper_cell = sheet.cell(row=helper_row, column=column_index)
                helper_cell.fill = helper_fill
                helper_cell.font = helper_font
                helper_cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = freeze_panes

    def style_guide_sheet(sheet, header_count: int) -> None:
        for column_index in range(1, header_count + 1):
            header_cell = sheet.cell(row=1, column=column_index)
            header_cell.fill = guide_fill
            header_cell.font = header_font
        sheet.freeze_panes = "A2"

    def autofit_sheet(sheet) -> None:
        for column_cells in sheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = max((len(value) for value in values), default=10)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 4, 14), 40)

    fund_sheet = workbook.active
    fund_sheet.title = MIGRATION_SHEET_NAME_FUND
    fund_sheet.append([
        _migration_header_label(key, MIGRATION_FUND_HEADER_LABELS, MIGRATION_FUND_REQUIRED_HEADERS)
        for key in MIGRATION_FUND_HEADERS
    ])
    fund_sheet.append([
        "# 작성가이드",
        "조합의 공식 명칭",
        "조합 상세 카드와 같은 조합유형만 입력",
        "결성예정 / 운용 중 / 해산 / 청산 완료",
        "YYYY-MM-DD",
        "실제 등록번호 또는 사업자번호",
        "YYYY-MM-DD",
        "대표 펀드매니저명",
        "공동 GP가 있으면 입력",
        "수탁사명",
        "숫자만 입력",
        "대표 GP 약정액",
        "일시 / 분할 / 수시",
        "YYYY-MM-DD",
        "YYYY-MM-DD",
        "YYYY-MM-DD",
        "숫자만 입력",
        "숫자만 입력",
        "숫자만 입력",
        "실계좌번호가 있으면 입력",
    ])
    fund_sheet.append([
        "# 예시",
        "브이온 벤처성장조합 1호",
        "벤처투자조합",
        "운용 중",
        "2026-01-15",
        "123-45-67890",
        "2026-02-01",
        "김펀드매니저",
        "",
        "OO은행",
        10000000000,
        1000000000,
        "분할",
        "2030-01-15",
        "2034-01-15",
        "",
        2.0,
        20.0,
        6.0,
        "110-123-456789",
    ])
    fund_sheet.append(["# 조합번호 예시", "1, 2, 3처럼 파일 안에서만 쓰는 번호입니다. 두 시트에서 같은 조합은 같은 번호를 씁니다."])
    style_sheet(fund_sheet, len(MIGRATION_FUND_HEADERS), "A4", (2, 3, 4))

    lp_sheet = workbook.create_sheet(MIGRATION_SHEET_NAME_LP)
    lp_sheet.append([
        _migration_header_label(key, MIGRATION_LP_HEADER_LABELS, MIGRATION_LP_REQUIRED_HEADERS)
        for key in MIGRATION_LP_HEADERS
    ])
    lp_sheet.append([
        "# 작성가이드",
        "LP의 공식 명칭",
        MIGRATION_LP_TYPE_GUIDE_TEXT,
        "같은 LP 첫 행에서 입력",
        "선택 입력",
        "선택 입력",
        "선택 입력",
        "선택 입력",
        "납입회차가 있으면 YYYY-MM-DD",
        "비워두면 자동으로 1,2,3",
        "해당 회차분 비율만 입력",
        "비우면 납입기일과 동일 처리",
        "선택 입력",
    ])
    lp_sheet.append([
        "# 예시-1",
        "한국성장금융",
        LP_TYPE_INSTITUTIONAL,
        3000000000,
        300000000,
        "02-0000-0000",
        "111-22-33333",
        "서울시 강남구",
        "2026-02-10",
        1,
        10,
        "2026-02-10",
        "1차 납입",
    ])
    lp_sheet.append([
        "# 예시-2",
        "한국성장금융",
        "",
        "",
        "",
        "",
        "",
        "",
        "2026-05-10",
        "",
        15,
        "2026-05-12",
        "2차 납입, 기본정보 생략 가능",
    ])
    style_sheet(lp_sheet, len(MIGRATION_LP_HEADERS), "A5", (2, 3, 4))

    guide_sheet = workbook.create_sheet(MIGRATION_SHEET_NAME_GUIDE)
    guide_sheet.append(["항목", "설명"])
    guide_sheet.append(["사용 순서", "1) 조합 시트 입력 -> 2) 조합원(LP) 시트 입력 -> 3) 검증 -> 4) Import"])
    guide_sheet.append(["조합번호", "파일 내부 연결용 번호입니다. 1, 2, 3처럼 간단히 입력하고 같은 조합은 두 시트에서 같은 번호를 사용합니다."])
    guide_sheet.append(["등록번호", "실제 고유번호증/사업자번호입니다. 조합번호와 다른 값입니다."])
    guide_sheet.append(["회차번호", "같은 LP의 납입 순번입니다. 비워두면 납입기일 순서대로 1, 2, 3이 자동 부여됩니다."])
    guide_sheet.append(["상태", "결성예정 / 운용 중 / 해산 / 청산 완료만 입력할 수 있습니다."])
    guide_sheet.append(["대표 GP", "엑셀에는 입력하지 않습니다. Import 시 고유계정의 대표 GP 법인 1개에 자동 연결됩니다."])
    guide_sheet.append(["LP유형", f"허용값: {MIGRATION_LP_TYPE_GUIDE_TEXT}. 업무집행조합원(GP)은 엑셀에 입력하지 않고 대표 GP 법인에서 자동 생성합니다."])
    guide_sheet.append(["LP 기본정보", "같은 LP가 여러 회차를 납입하면 첫 행에만 LP유형/약정총액/사업자번호 등을 입력하고 다음 행은 비워둘 수 있습니다."])
    guide_sheet.append(["필수값", "조합: 조합번호/조합명/조합유형/상태/결성일, 조합원(LP): 조합번호/LP명, 납입회차 행: 납입기일/회차별납입비율"])
    guide_sheet.append(["숫자/날짜 형식", "금액/비율은 숫자만 입력하고, 날짜는 YYYY-MM-DD 형식으로 입력합니다."])
    style_guide_sheet(guide_sheet, 2)

    column_guide_sheet = workbook.create_sheet(MIGRATION_SHEET_NAME_COLUMN_GUIDE)
    column_guide_sheet.append(["시트", "컬럼", "필수", "입력 가이드", "예시"])
    for row in [
        (MIGRATION_SHEET_NAME_FUND, "조합번호", "Y", "파일 내부 연결용 번호. 1, 2, 3처럼 입력", "1"),
        (MIGRATION_SHEET_NAME_FUND, "조합명", "Y", "조합의 공식 명칭", "브이온 벤처성장조합 1호"),
        (MIGRATION_SHEET_NAME_FUND, "조합유형", "Y", "조합 상세 카드와 동일한 유형만 입력", "벤처투자조합"),
        (MIGRATION_SHEET_NAME_FUND, "상태", "Y", "결성예정 / 운용 중 / 해산 / 청산 완료", "운용 중"),
        (MIGRATION_SHEET_NAME_FUND, "결성일", "Y", "조합 결성일", "2026-01-15"),
        (MIGRATION_SHEET_NAME_FUND, "등록번호", "N", "실제 고유번호증/사업자번호", "123-45-67890"),
        (MIGRATION_SHEET_NAME_FUND, "GP약정액", "N", "대표 GP 약정금액", "1000000000"),
        (MIGRATION_SHEET_NAME_LP, "조합번호", "Y", "조합 시트와 동일한 번호", "1"),
        (MIGRATION_SHEET_NAME_LP, "LP명", "Y", "LP의 공식 명칭", "한국성장금융"),
        (MIGRATION_SHEET_NAME_LP, "LP유형", "조건부", f"같은 LP의 첫 행에서 입력. 허용값: {MIGRATION_LP_TYPE_GUIDE_TEXT}", LP_TYPE_INSTITUTIONAL),
        (MIGRATION_SHEET_NAME_LP, "약정총액", "조건부", "같은 LP의 첫 행에서 입력", "3000000000"),
        (MIGRATION_SHEET_NAME_LP, "누적납입액", "N", "현재까지 실제 납입 누계", "300000000"),
        (MIGRATION_SHEET_NAME_LP, "사업자번호", "N", "있으면 입력, 없으면 생략 가능", "111-22-33333"),
        (MIGRATION_SHEET_NAME_LP, "납입기일", "납입회차 행", "납입회차를 적는 행이면 필수", "2026-02-10"),
        (MIGRATION_SHEET_NAME_LP, "회차번호", "N", "비워두면 자동 부여", "1"),
        (MIGRATION_SHEET_NAME_LP, "회차별납입비율(%)", "납입회차 행", "해당 회차분 비율만 입력", "10"),
        (MIGRATION_SHEET_NAME_LP, "실납입일", "N", "비우면 납입기일과 동일 처리", "2026-02-10"),
    ]:
        column_guide_sheet.append(list(row))
    style_guide_sheet(column_guide_sheet, 5)

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        autofit_sheet(sheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = "조합_마이그레이션_템플릿.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"fund_migration_template.xlsx\"; filename*=UTF-8''{encoded_filename}"
        },
    )


@router.get("/api/funds/migration-template")
def download_migration_template():
    return _download_migration_template_v2()

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from exc

    workbook = openpyxl.Workbook()
    header_fill = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
    helper_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    guide_fill = PatternFill(start_color="EEF6FF", end_color="EEF6FF", fill_type="solid")
    header_font = Font(bold=True, color="1F2937")
    helper_font = Font(color="6B7280", italic=True, size=10)

    def style_sheet(sheet, header_count: int) -> None:
        for column_index in range(1, header_count + 1):
            header_cell = sheet.cell(row=1, column=column_index)
            header_cell.fill = header_fill
            header_cell.font = header_font
            helper_cell = sheet.cell(row=2, column=column_index)
            helper_cell.fill = helper_fill
            helper_cell.font = helper_font
            helper_cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A3"

    def autofit_sheet(sheet) -> None:
        for column_cells in sheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = max((len(value) for value in values), default=10)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 4, 14), 44)

    funds_sheet = workbook.active
    funds_sheet.title = "Funds"
    funds_sheet.append([MIGRATION_FUND_HEADER_LABELS.get(key, key) for key in MIGRATION_FUND_HEADERS])
    funds_sheet.append(
        [
            "고유 식별키. Funds/LPs/LPContributions에서 반드시 같은 값을 반복 사용",
            "조합의 공식 명칭 입력",
            "벤처투자조합 / 투자조합 / 신기술투자조합 등",
            "forming / active / dissolved / liquidated",
            "조합 결성일, YYYY-MM-DD",
            "실제 등록번호 또는 사업자번호. 조합키와 다름",
            "등록일, 없으면 비움",
            "대표 GP명",
            "대표 운용역 또는 총괄 담당자",
            "공동 GP가 있으면 입력",
            "수탁사명",
            "숫자만 입력, 콤마 없이",
            "숫자만 입력, 콤마 없이",
            "분할 / 일시 / 기타",
            "투자기간 종료일, YYYY-MM-DD",
            "만기일, YYYY-MM-DD",
            "청산 종료 예정일이 있으면 입력",
            "관리보수율(%) 숫자",
            "성과보수율(%) 숫자",
            "허들레이트(%) 숫자",
            "실제 계좌번호가 있으면 입력",
        ]
    )
    funds_sheet.append(
        [
            "FUND-001",
            "예시 벤처성장조합 1호",
            "벤처투자조합",
            "active",
            "2026-01-15",
            "123-45-67890",
            "2026-02-01",
            "예시 GP",
            "김펀드매니저",
            "",
            "OO은행",
            10000000000,
            1000000000,
            "분할",
            "2030-01-15",
            "2034-01-15",
            "",
            2.0,
            20.0,
            6.0,
            "110-123-456789",
        ]
    )
    style_sheet(funds_sheet, len(MIGRATION_FUND_HEADERS))

    lp_sheet = workbook.create_sheet("LPs")
    lp_sheet.append([MIGRATION_LP_HEADER_LABELS.get(key, key) for key in MIGRATION_LP_HEADERS])
    lp_sheet.append(
        [
            "Funds 시트의 조합키와 동일하게 입력",
            "LP의 공식 명칭 입력",
            "기관투자자 / 개인투자자 / GP",
            "약정총액 숫자만 입력",
            "현재까지 납입한 금액 숫자만 입력",
            "담당자 연락처",
            "사업자번호가 있으면 입력, LPContributions 매칭에 우선 사용",
            "우편주소 또는 소재지",
        ]
    )
    lp_sheet.append(
        [
            "FUND-001",
            "예시 기관 LP",
            "기관투자자",
            3000000000,
            500000000,
            "02-0000-0000",
            "111-22-33333",
            "서울시 강남구",
        ]
    )
    lp_sheet.append(
        [
            "FUND-001",
            "예시 GP",
            "GP",
            1000000000,
            300000000,
            "02-1111-2222",
            "123-45-67890",
            "서울시 서초구",
        ]
    )
    style_sheet(lp_sheet, len(MIGRATION_LP_HEADERS))

    contribution_sheet = workbook.create_sheet("LPContributions")
    contribution_sheet.append(
        [MIGRATION_CONTRIBUTION_HEADER_LABELS.get(key, key) for key in MIGRATION_CONTRIBUTION_HEADERS]
    )
    contribution_sheet.append(
        [
            "Funds 시트의 조합키와 동일하게 입력",
            "LPs 시트의 LP명과 동일하게 입력",
            "사업자번호가 있으면 우선 매칭, 없으면 LP명으로 매칭",
            "납입기일, YYYY-MM-DD",
            "회차번호. 같은 조합 안에서 1부터 순서대로 증가",
            "해당 회차에 실제 납입한 비율(%). 누적 비율 아님",
            "실납입일, 비우면 납입기일과 동일하게 처리",
            "비고 또는 메모",
        ]
    )
    contribution_sheet.append(
        [
            "FUND-001",
            "예시 기관 LP",
            "111-22-33333",
            "2026-02-10",
            1,
            10,
            "2026-02-10",
            "최초 출자",
        ]
    )
    contribution_sheet.append(
        [
            "FUND-001",
            "예시 기관 LP",
            "111-22-33333",
            "2026-05-10",
            2,
            15,
            "2026-05-12",
            "추가 출자",
        ]
    )
    style_sheet(contribution_sheet, len(MIGRATION_CONTRIBUTION_HEADERS))

    guide_sheet = workbook.create_sheet("Guide")
    guide_sheet.append(["항목", "설명"])
    guide_sheet.append(["사용 순서", "1) Funds 입력 -> 2) LPs 입력 -> 3) LPContributions 입력 -> 4) Validate -> 5) Import"])
    guide_sheet.append(["조합키", "내부 연결용 키입니다. 실제 등록번호가 아니며, 세 시트에서 같은 조합은 동일한 조합키를 사용해야 합니다."])
    guide_sheet.append(["등록번호", "실제 등록번호/사업자번호입니다. 조합키와 별개이며 조합 식별 보조값으로 사용됩니다."])
    guide_sheet.append(["회차번호", "같은 조합 내 납입 순번입니다. 1부터 시작해 회차별로 증가시켜 입력하세요."])
    guide_sheet.append(["회차별납입비율(%)", "누적 비율이 아니라 해당 회차에 실제 납입한 비율입니다. 총합이 100%를 넘지 않도록 입력하세요."])
    guide_sheet.append(["날짜 형식", "모든 날짜는 YYYY-MM-DD 형식으로 입력하세요."])
    guide_sheet.append(["숫자 형식", "금액/비율은 숫자만 입력하세요. 콤마, 원, %, 공백은 넣지 않습니다."])
    guide_sheet.append(["필수값", "Funds: 조합키/조합명/조합유형, LPs: 조합키/LP명/LP유형, LPContributions: 조합키/LP명/납입기일/회차번호/회차별납입비율(%)"])
    guide_sheet.append(["주의사항", "Import 성공 후 같은 조합은 재등록이 제한될 수 있으니, 먼저 Validate로 오류를 확인한 뒤 Import하세요."])
    for column_index in range(1, 3):
        cell = guide_sheet.cell(row=1, column=column_index)
        cell.fill = guide_fill
        cell.font = header_font
    guide_sheet.freeze_panes = "A2"

    column_guide_sheet = workbook.create_sheet("ColumnGuide")
    column_guide_sheet.append(["시트", "컬럼", "필수", "입력 가이드", "예시"])
    column_guide_rows = [
        ("Funds", "조합키", "Y", "세 시트 연결용 내부 식별값. 조합당 하나의 고정값을 사용", "FUND-001"),
        ("Funds", "조합명", "Y", "조합의 공식 명칭", "예시 벤처성장조합 1호"),
        ("Funds", "조합유형", "Y", "사전에 정의된 조합 유형 사용", "벤처투자조합"),
        ("Funds", "상태", "Y", "forming / active / dissolved / liquidated 중 하나", "active"),
        ("Funds", "결성일", "Y", "조합 결성일", "2026-01-15"),
        ("Funds", "등록번호", "N", "실제 등록번호 또는 사업자번호. 조합키와 다름", "123-45-67890"),
        ("Funds", "약정총액", "N", "총 약정금액 숫자만 입력", "10000000000"),
        ("Funds", "GP출자금", "N", "GP 납입 또는 약정 금액 숫자만 입력", "1000000000"),
        ("Funds", "관리보수율(%)", "N", "백분율 숫자만 입력", "2.0"),
        ("Funds", "성과보수율(%)", "N", "백분율 숫자만 입력", "20.0"),
        ("Funds", "허들레이트(%)", "N", "백분율 숫자만 입력", "6.0"),
        ("LPs", "조합키", "Y", "Funds 시트와 동일한 조합키 입력", "FUND-001"),
        ("LPs", "LP명", "Y", "LP의 공식 명칭 입력", "예시 기관 LP"),
        ("LPs", "LP유형", "Y", "기관투자자 / 개인투자자 / GP", "기관투자자"),
        ("LPs", "약정총액", "N", "약정금액 숫자만 입력", "3000000000"),
        ("LPs", "납입총액", "N", "현재까지 실제 납입한 금액", "500000000"),
        ("LPs", "사업자번호", "N", "LPContributions 매칭 정확도를 높이기 위해 권장", "111-22-33333"),
        ("LPContributions", "조합키", "Y", "Funds 시트와 동일한 조합키 입력", "FUND-001"),
        ("LPContributions", "LP명", "Y", "LPs 시트의 LP명과 동일하게 입력", "예시 기관 LP"),
        ("LPContributions", "LP사업자번호", "N", "있으면 우선 매칭, 없으면 LP명으로 매칭", "111-22-33333"),
        ("LPContributions", "납입기일", "Y", "회차별 예정 또는 기준 납입일", "2026-02-10"),
        ("LPContributions", "회차번호", "Y", "같은 조합 내에서 1, 2, 3 순서로 증가", "1"),
        ("LPContributions", "회차별납입비율(%)", "Y", "해당 회차분만 입력. 누적값 금지", "10"),
        ("LPContributions", "실납입일", "N", "실제 납입일. 비우면 납입기일과 동일하게 처리", "2026-02-10"),
    ]
    for row in column_guide_rows:
        column_guide_sheet.append(list(row))
    for column_index in range(1, 6):
        cell = column_guide_sheet.cell(row=1, column=column_index)
        cell.fill = guide_fill
        cell.font = header_font
    column_guide_sheet.freeze_panes = "A2"

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        autofit_sheet(sheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = "조합_LP_일괄등록_템플릿.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"Fund_LP_bulk_upload_template.xlsx\"; filename*=UTF-8''{encoded_filename}"
            )
        },
    )


@router.post("/api/funds/migration-validate", response_model=FundMigrationValidateResponse)
async def validate_migration(
    payload: bytes = Body(..., media_type="application/octet-stream"),
    db: Session = Depends(get_db),
):
    if not payload:
        raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다")
    validation, _, _, _ = _parse_and_validate_migration_v2(payload, db)
    return validation


def _find_existing_fund(db: Session, row: dict) -> Fund | None:
    registration_number = row.get("registration_number")
    if registration_number:
        existing = db.query(Fund).filter(Fund.registration_number == registration_number).first()
        if existing:
            return existing
    if row.get("name") and row.get("formation_date"):
        existing = (
            db.query(Fund)
            .filter(
                Fund.name == row["name"],
                Fund.formation_date == row["formation_date"],
            )
            .first()
        )
        if existing:
            return existing
    return None


def _find_existing_lp(db: Session, fund_id: int, row: dict) -> LP | None:
    business_number = row.get("business_number")
    if business_number:
        existing = (
            db.query(LP)
            .filter(
                LP.fund_id == fund_id,
                LP.business_number == business_number,
            )
            .first()
        )
        if existing:
            return existing
    return (
        db.query(LP)
        .filter(
            LP.fund_id == fund_id,
            LP.name == row["name"],
        )
        .first()
    )


def _normalize_lp_text(value: str | None) -> str:
    return (value or "").strip()


def _normalize_lp_optional(value: str | None) -> str | None:
    text = _normalize_lp_text(value)
    return text or None


def _get_lp_address_book(db: Session, address_book_id: int | None) -> LPAddressBook | None:
    if address_book_id is None:
        return None
    row = db.get(LPAddressBook, address_book_id)
    if not row:
        raise HTTPException(status_code=404, detail="LP address book not found")
    return row


def _ensure_lp_unique_in_fund(
    db: Session,
    fund_id: int,
    *,
    name: str,
    business_number: str | None,
    address_book_id: int | None,
    current_lp_id: int | None = None,
) -> None:
    normalized_name = _normalize_lp_text(name)
    normalized_business_number = _normalize_lp_optional(business_number)

    if address_book_id is not None:
        query = db.query(LP).filter(
            LP.fund_id == fund_id,
            LP.address_book_id == address_book_id,
        )
        if current_lp_id is not None:
            query = query.filter(LP.id != current_lp_id)
        if query.first():
            raise HTTPException(
                status_code=409,
                detail="This address-book entity is already registered in the same fund",
            )

    if normalized_business_number:
        query = db.query(LP).filter(
            LP.fund_id == fund_id,
            LP.business_number == normalized_business_number,
        )
        if current_lp_id is not None:
            query = query.filter(LP.id != current_lp_id)
        if query.first():
            raise HTTPException(
                status_code=409,
                detail="LP with the same business number already exists in this fund",
            )

    query = db.query(LP).filter(
        LP.fund_id == fund_id,
        LP.name == normalized_name,
    )
    if current_lp_id is not None:
        query = query.filter(LP.id != current_lp_id)
    if query.first():
        raise HTTPException(
            status_code=409,
            detail="LP with the same name already exists in this fund",
        )


def _upsert_lp_address_book(db: Session, row: dict) -> int:
    business_number = row.get("business_number")
    existing = None
    if business_number:
        existing = db.query(LPAddressBook).filter(LPAddressBook.business_number == business_number).first()
    if existing is None:
        existing = (
            db.query(LPAddressBook)
            .filter(
                LPAddressBook.name == row["name"],
                LPAddressBook.type == row["type"],
            )
            .first()
        )

    if existing is None:
        db.add(
            LPAddressBook(
                name=row["name"],
                type=row["type"],
                business_number=row.get("business_number"),
                contact=row.get("contact"),
                address=row.get("address"),
                memo=None,
                gp_entity_id=None,
                is_active=1,
            )
        )
        return 1

    existing.name = row["name"]
    existing.type = row["type"]
    existing.business_number = row.get("business_number")
    existing.contact = row.get("contact")
    existing.address = row.get("address")
    existing.is_active = 1
    return 1


@router.post("/api/funds/migration-import", response_model=FundMigrationImportResponse)
async def import_migration(
    payload: bytes = Body(..., media_type="application/octet-stream"),
    mode: str = Query("upsert"),
    sync_address_book: bool = Query(False),
    db: Session = Depends(get_db),
):
    if mode not in {"insert", "upsert"}:
        raise HTTPException(status_code=400, detail="mode는 insert/upsert 중 하나여야 합니다")

    if not payload:
        raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다")

    validation, fund_rows, lp_rows, contribution_rows = _parse_and_validate_migration_v2(payload, db)
    if validation.errors:
        return FundMigrationImportResponse(
            success=False,
            mode=mode,
            fund_rows=validation.fund_rows,
            lp_rows=validation.lp_rows,
            contribution_rows=validation.contribution_rows,
            warnings=validation.warnings,
            errors=validation.errors,
            validation=validation,
        )

    import_errors: list[FundMigrationErrorItem] = []
    created_funds = 0
    updated_funds = 0
    created_lps = 0
    updated_lps = 0
    created_contributions = 0
    synced_address_books = 0

    try:
        primary_gp = _require_migration_primary_gp(db)
        fund_map: dict[str, Fund] = {}
        contribution_fund_ids: set[int] = set()
        for row in fund_rows:
            existing = _find_existing_fund(db, row)

            if mode == "insert" and existing is not None:
                import_errors.append(
                    FundMigrationErrorItem(
                        row=row["__row"],
                        column="registration_number",
                        reason="insert 모드에서 이미 존재하는 조합입니다",
                    )
                )
                continue

            payload = {
                "name": row["name"],
                "type": row["type"],
                "status": row["status"],
                "formation_date": row["formation_date"],
                "registration_number": row["registration_number"],
                "registration_date": row["registration_date"],
                "gp_entity_id": primary_gp.id,
                "gp": primary_gp.name,
                "fund_manager": row["fund_manager"],
                "co_gp": row["co_gp"],
                "trustee": row["trustee"],
                "commitment_total": row["commitment_total"],
                "gp_commitment": row["gp_commitment"],
                "contribution_type": row["contribution_type"],
                "investment_period_end": row["investment_period_end"],
                "maturity_date": row["maturity_date"],
                "dissolution_date": row["dissolution_date"],
                "mgmt_fee_rate": row["mgmt_fee_rate"],
                "performance_fee_rate": row["performance_fee_rate"],
                "hurdle_rate": row["hurdle_rate"],
                "account_number": row["account_number"],
            }

            if existing is None:
                fund = Fund(**payload)
                db.add(fund)
                db.flush()
                created_funds += 1
            else:
                fund = existing
                for key, value in payload.items():
                    setattr(fund, key, value)
                updated_funds += 1

            _sync_fee_config_from_fund(db, fund, {"mgmt_fee_rate", "performance_fee_rate", "hurdle_rate"})
            _ensure_gp_lp_record(
                db,
                fund=fund,
                gp_name=primary_gp.name,
                gp_commitment=row["gp_commitment"],
                gp_entity=primary_gp,
            )

            fund_map[row["fund_key"]] = fund

        lp_entity_map: dict[tuple[str, str], LP] = {}
        for row in lp_rows:
            fund = fund_map.get(row["fund_key"])
            if fund is None:
                import_errors.append(
                    FundMigrationErrorItem(
                        row=row["__row"],
                        column=_migration_column_label("fund_key", MIGRATION_LP_HEADER_LABELS),
                        reason="매칭된 조합을 찾을 수 없습니다",
                    )
                )
                continue

            existing_lp = _find_existing_lp_by_identifiers(
                db,
                fund.id,
                name=row["name"],
                business_number=row.get("business_number"),
            )
            if mode == "insert" and existing_lp is not None:
                import_errors.append(
                    FundMigrationErrorItem(
                        row=row["__row"],
                        column=_migration_column_label("business_number", MIGRATION_LP_HEADER_LABELS),
                        reason="insert 모드에서 이미 존재하는 LP입니다",
                    )
                )
                continue

            lp_payload = {
                "name": row["name"],
                "type": row["type"],
                "commitment": row["commitment"],
                "paid_in": row["paid_in"],
                "contact": row["contact"],
                "business_number": row["business_number"],
                "address": row["address"],
            }

            if existing_lp is None:
                existing_lp = LP(fund_id=fund.id, **lp_payload)
                db.add(existing_lp)
                db.flush()
                created_lps += 1
            else:
                for key, value in lp_payload.items():
                    if value is None and key in {"paid_in", "contact", "business_number", "address"}:
                        continue
                    setattr(existing_lp, key, value)
                updated_lps += 1

            if row.get("business_number"):
                lp_entity_map[(row["fund_key"], _normalize_lookup_text(row["business_number"]))] = existing_lp
            lp_entity_map[(row["fund_key"], _normalize_lookup_text(row["name"]))] = existing_lp

            if sync_address_book:
                synced_address_books += _upsert_lp_address_book(db, row)

        for row in contribution_rows:
            fund = fund_map.get(row["fund_key"])
            if fund is None:
                import_errors.append(
                    FundMigrationErrorItem(
                        row=row["__row"],
                        column=_migration_column_label("fund_key", MIGRATION_LP_HEADER_LABELS),
                        reason="매칭된 조합을 찾을 수 없습니다",
                    )
                )
                continue

            lookup_token = _normalize_lookup_text(row["lp_business_number"] or row["lp_name"])
            lp = lp_entity_map.get((row["fund_key"], lookup_token))
            if lp is None:
                lp = _find_existing_lp_by_identifiers(
                    db,
                    fund.id,
                    name=row["lp_name"],
                    business_number=row["lp_business_number"],
                )
            if lp is None:
                import_errors.append(
                    FundMigrationErrorItem(
                        row=row["__row"],
                        column=_migration_column_label("name", MIGRATION_LP_HEADER_LABELS),
                        reason="매칭된 LP를 찾을 수 없습니다",
                    )
                )
                continue

            contribution = LPContribution(
                fund_id=fund.id,
                lp_id=lp.id,
                due_date=row["due_date"],
                amount=float(row["amount"]),
                commitment_ratio=float(row["commitment_ratio_percent"]),
                round_no=row["round_no"],
                actual_paid_date=row["actual_paid_date"],
                memo=row["memo"],
                source="migration",
            )
            db.add(contribution)
            created_contributions += 1
            contribution_fund_ids.add(fund.id)

        if import_errors:
            db.rollback()
            return FundMigrationImportResponse(
                success=False,
                mode=mode,
                fund_rows=validation.fund_rows,
                lp_rows=validation.lp_rows,
                contribution_rows=validation.contribution_rows,
                created_funds=0,
                updated_funds=0,
                created_lps=0,
                updated_lps=0,
                created_contributions=0,
                synced_address_books=0,
                warnings=validation.warnings,
                errors=import_errors,
                validation=validation,
            )

        db.flush()
        for fund_id in contribution_fund_ids:
            recalculate_fund_stats(db, fund_id)
        locked_at = datetime.utcnow()
        for fund in fund_map.values():
            fund.initial_import_completed_at = locked_at
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"마이그레이션 import 중 오류가 발생했습니다: {exc}") from exc

    return FundMigrationImportResponse(
        success=True,
        mode=mode,
        fund_rows=validation.fund_rows,
        lp_rows=validation.lp_rows,
        contribution_rows=validation.contribution_rows,
        created_funds=created_funds,
        updated_funds=updated_funds,
        created_lps=created_lps,
        updated_lps=updated_lps,
        created_contributions=created_contributions,
        synced_address_books=synced_address_books,
        warnings=validation.warnings,
        errors=[],
        validation=validation,
    )

@router.get("/api/funds/{fund_id}", response_model=FundResponse)
def get_fund(fund_id: int, db: Session = Depends(get_db)):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")
    return fund


@router.post(
    "/api/funds/{fund_id}/add-formation-workflow",
    response_model=FundFormationWorkflowAddResponse,
    status_code=201,
)
def add_formation_workflow(
    fund_id: int,
    data: FundFormationWorkflowAddRequest,
    db: Session = Depends(get_db),
):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")
    if not _is_forming_status(fund.status):
        raise HTTPException(status_code=400, detail="결성예정 상태 조합에서만 생성할 수 있습니다")

    formation_slot_name = _resolve_formation_template_name(data.template_category_or_name)
    if not formation_slot_name:
        raise HTTPException(status_code=400, detail="지원하지 않는 결성 워크플로 분류입니다")

    template = _find_formation_template_by_id(db, data.template_id)
    if template is None:
        template = _find_formation_template(db, formation_slot_name)
    if not template:
        raise HTTPException(status_code=404, detail="결성 워크플로 템플릿을 찾을 수 없습니다")
    if not template.steps:
        raise HTTPException(status_code=400, detail="단계가 없는 템플릿은 생성할 수 없습니다")
    if (
        _normalize_template_identifier(formation_slot_name)
        == _normalize_template_identifier("결성총회 개최")
        and not _template_has_formation_paid_in_step(template)
    ):
        raise HTTPException(
            status_code=400,
            detail="결성총회 템플릿에는 출자금 납입 확인 단계가 필요합니다",
        )

    existing_instance = _find_existing_instance_by_formation_slot(
        db,
        fund_id=fund_id,
        formation_slot_name=formation_slot_name,
    )
    if existing_instance:
        raise HTTPException(status_code=409, detail="이미 추가된 결성 워크플로입니다")

    trigger_date = data.trigger_date or fund.formation_date or date.today()
    formation_slot_token = _normalize_template_identifier(formation_slot_name)
    try:
        instance = instantiate_workflow(
            db=db,
            workflow=template,
            name=f"[{fund.name}] {template.name}",
            trigger_date=trigger_date,
            memo=(
                "manual_fund_formation_workflow "
                f"fund_id={fund.id} "
                f"template_id={template.id} "
                f"{FORMATION_SLOT_MEMO_PREFIX}{formation_slot_token}"
            ),
            fund_id=fund.id,
            auto_commit=False,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return FundFormationWorkflowAddResponse(
        instance_id=instance.id,
        workflow_id=template.id,
        workflow_name=template.name,
        formation_slot=formation_slot_name,
        instance_name=instance.name,
        status=instance.status,
        trigger_date=instance.trigger_date,
        fund_id=fund.id,
    )


@router.post("/api/funds", response_model=FundResponse, status_code=201)
def create_fund(data: FundCreate, db: Session = Depends(get_db)):
    payload = data.model_dump()
    if payload.get("gp_commitment") is None and payload.get("gp_commitment_amount") is not None:
        payload["gp_commitment"] = payload.get("gp_commitment_amount")
    payload.pop("gp_commitment_amount", None)
    gp_entity = _resolve_gp_entity(
        db,
        gp_entity_id=payload.get("gp_entity_id"),
        gp_name=payload.get("gp"),
    )
    if gp_entity is not None:
        payload["gp_entity_id"] = gp_entity.id
        payload["gp"] = gp_entity.name
    else:
        payload["gp_entity_id"] = None
        payload["gp"] = (payload.get("gp") or "").strip() or None
    fund = Fund(**payload)

    try:
        db.add(fund)
        db.flush()
        sync_fund_history(db, fund)
        _cleanup_fund_capital_calls(db, fund.id)
        _sync_fee_config_from_fund(db, fund)

        _ensure_gp_lp_record(
            db,
            fund=fund,
            gp_name=fund.gp,
            gp_commitment=fund.gp_commitment,
            gp_entity=gp_entity,
        )
        if backbone_enabled():
            subject = sync_fund_graph(db, fund)
            maybe_emit_mutation(
                db,
                subject=subject,
                event_type="fund.created",
                after=record_snapshot(fund),
                origin_model="fund",
                origin_id=fund.id,
            )
            if gp_entity is not None:
                sync_gp_entity_graph(db, gp_entity)

        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="조합 생성 중 중복 제약조건 오류가 발생했습니다") from exc
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"조합 생성 중 오류가 발생했습니다: {exc}") from exc

    db.refresh(fund)
    _run_compliance_rule_checks(
        db=db,
        fund_id=fund.id,
        trigger_source="fund_create",
        trigger_source_id=fund.id,
    )
    return fund


@router.put("/api/funds/{fund_id}", response_model=FundResponse)
def update_fund(fund_id: int, data: FundUpdate, db: Session = Depends(get_db)):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")

    before = record_snapshot(fund)
    update_payload = data.model_dump(exclude_unset=True)
    if update_payload.get("gp_commitment") is None and update_payload.get("gp_commitment_amount") is not None:
        update_payload["gp_commitment"] = update_payload.get("gp_commitment_amount")
    update_payload.pop("gp_commitment_amount", None)
    gp_entity = None
    if "gp_entity_id" in update_payload or "gp" in update_payload:
        next_gp_entity_id = update_payload.get("gp_entity_id", fund.gp_entity_id)
        next_gp_name = update_payload.get("gp", fund.gp)
        if "gp_entity_id" in update_payload and update_payload["gp_entity_id"] is None and "gp" not in update_payload:
            next_gp_name = None
        gp_entity = _resolve_gp_entity(
            db,
            gp_entity_id=next_gp_entity_id,
            gp_name=next_gp_name,
        )
        if gp_entity is not None:
            update_payload["gp_entity_id"] = gp_entity.id
            update_payload["gp"] = gp_entity.name
        elif "gp_entity_id" in update_payload:
            update_payload["gp_entity_id"] = None
            update_payload["gp"] = (update_payload.get("gp") or "").strip() or None
        elif "gp" in update_payload:
            update_payload["gp_entity_id"] = None
            update_payload["gp"] = (update_payload.get("gp") or "").strip() or None

    for key, val in update_payload.items():
        setattr(fund, key, val)

    sync_fund_history(db, fund)
    _sync_fee_config_from_fund(db, fund, set(update_payload.keys()))
    _ensure_gp_lp_record(
        db,
        fund=fund,
        gp_name=fund.gp,
        gp_commitment=fund.gp_commitment,
        gp_entity=gp_entity,
    )
    if backbone_enabled():
        db.flush()
        subject = sync_fund_graph(db, fund)
        maybe_emit_mutation(
            db,
            subject=subject,
            event_type="fund.updated",
            before=before,
            after=record_snapshot(fund),
            origin_model="fund",
            origin_id=fund.id,
        )
        if gp_entity is not None:
            sync_gp_entity_graph(db, gp_entity)
    db.commit()
    db.refresh(fund)
    _run_compliance_rule_checks(
        db=db,
        fund_id=fund.id,
        trigger_source="fund_update",
        trigger_source_id=fund.id,
    )
    return fund


def _is_completed_status(status: str | None) -> bool:
    return (status or "").strip().lower() == "completed"


def _cleanup_fund_capital_calls(db: Session, fund_id: int) -> None:
    call_ids = [
        int(row[0])
        for row in (
            db.query(CapitalCall.id)
            .filter(CapitalCall.fund_id == fund_id)
            .all()
        )
    ]
    if not call_ids:
        return

    (
        db.query(CapitalCallItem)
        .filter(CapitalCallItem.capital_call_id.in_(call_ids))
        .delete(synchronize_session=False)
    )
    (
        db.query(CapitalCall)
        .filter(CapitalCall.id.in_(call_ids))
        .delete(synchronize_session=False)
    )


def _cleanup_fund_tasks_and_workflows(db: Session, fund_id: int) -> None:
    completed_workflows = (
        db.query(WorkflowInstance)
        .filter(
            WorkflowInstance.fund_id == fund_id,
            func.lower(func.coalesce(WorkflowInstance.status, "")) == "completed",
        )
        .all()
    )
    for instance in completed_workflows:
        instance.fund_id = None

    non_completed_workflows = (
        db.query(WorkflowInstance)
        .filter(
            WorkflowInstance.fund_id == fund_id,
            func.lower(func.coalesce(WorkflowInstance.status, "")) != "completed",
        )
        .all()
    )
    deleted_workflow_ids: set[int] = set()
    handled_task_ids: set[int] = set()

    for instance in non_completed_workflows:
        deleted_workflow_ids.add(instance.id)
        for step_instance in instance.step_instances:
            if not step_instance.task_id:
                continue
            task = db.get(Task, step_instance.task_id)
            step_instance.task_id = None
            if not task:
                continue

            if _is_completed_status(task.status):
                task.fund_id = None
                task.workflow_instance_id = None
                task.workflow_step_order = None
            else:
                db.delete(task)
            handled_task_ids.add(task.id)

        db.delete(instance)

    tasks = db.query(Task).filter(Task.fund_id == fund_id).all()
    for task in tasks:
        if task.id in handled_task_ids:
            continue

        if _is_completed_status(task.status):
            task.fund_id = None
            if task.workflow_instance_id in deleted_workflow_ids:
                task.workflow_instance_id = None
                task.workflow_step_order = None
            continue

        (
            db.query(WorkflowStepInstance)
            .filter(WorkflowStepInstance.task_id == task.id)
            .update({WorkflowStepInstance.task_id: None}, synchronize_session=False)
        )
        db.delete(task)


@router.delete("/api/funds/{fund_id}", status_code=204)
def delete_fund(fund_id: int, db: Session = Depends(get_db)):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")
    before = record_snapshot(fund)
    try:
        _cleanup_fund_capital_calls(db, fund_id)
        _cleanup_fund_tasks_and_workflows(db, fund_id)
        db.flush()
        if backbone_enabled():
            subject = mark_subject_deleted(
                db,
                subject_type="fund",
                native_id=fund.id,
                payload=before,
            )
            maybe_emit_mutation(
                db,
                subject=subject,
                event_type="fund.deleted",
                before=before,
                origin_model="fund",
                origin_id=fund.id,
            )
        db.delete(fund)
        db.commit()
    except Exception:
        db.rollback()
        raise


@router.get("/api/funds/{fund_id}/lps", response_model=list[LPResponse])
def list_lps(fund_id: int, db: Session = Depends(get_db)):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")
    return db.query(LP).filter(LP.fund_id == fund_id).order_by(LP.id.desc()).all()


@router.post("/api/funds/{fund_id}/lps", response_model=LPResponse, status_code=201)
def create_lp(fund_id: int, data: LPCreate, db: Session = Depends(get_db)):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")

    payload = data.model_dump()
    address_book = _get_lp_address_book(db, payload.get("address_book_id"))

    if address_book:
        payload["business_number"] = _normalize_lp_optional(payload.get("business_number")) or address_book.business_number
        payload["name"] = _normalize_lp_text(payload.get("name")) or address_book.name
        payload["type"] = _normalize_lp_text(payload.get("type")) or address_book.type
        payload["contact"] = _normalize_lp_optional(payload.get("contact")) or address_book.contact
        payload["address"] = _normalize_lp_optional(payload.get("address")) or address_book.address

    payload["name"] = _normalize_lp_text(payload.get("name"))
    payload["type"] = coerce_lp_type(payload.get("type"))
    payload["business_number"] = _normalize_lp_optional(payload.get("business_number"))
    payload["contact"] = _normalize_lp_optional(payload.get("contact"))
    payload["address"] = _normalize_lp_optional(payload.get("address"))

    if not payload["name"] or not payload["type"]:
        raise HTTPException(status_code=400, detail="LP name and type are required")

    validate_lp_paid_in_pair(
        commitment=payload.get("commitment"),
        paid_in=payload.get("paid_in"),
    )

    _ensure_lp_unique_in_fund(
        db,
        fund_id,
        name=payload["name"],
        business_number=payload.get("business_number"),
        address_book_id=payload.get("address_book_id"),
    )

    lp = LP(fund_id=fund_id, **payload)
    db.add(lp)
    try:
        db.flush()
        if backbone_enabled():
            subject = sync_lp_graph(db, lp)
            maybe_emit_mutation(
                db,
                subject=subject,
                event_type="lp.created",
                after=record_snapshot(lp),
                origin_model="lp",
                origin_id=lp.id,
            )
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="Duplicate LP in the same fund is not allowed") from exc
    db.refresh(lp)

    try:
        ComplianceEngine(db).on_lp_changed(fund_id, "joined")
    except Exception as exc:  # noqa: BLE001 - hook failures must not break main flow
        db.rollback()
        logger.warning(
            "compliance hook failed on create_lp: fund_id=%s lp_id=%s error=%s",
            fund_id,
            lp.id,
            exc,
        )
    _run_compliance_rule_checks(
        db=db,
        fund_id=fund_id,
        trigger_source="lp_create",
        trigger_source_id=lp.id,
    )
    return lp


@router.put("/api/funds/{fund_id}/lps/{lp_id}", response_model=LPResponse)
def update_lp(fund_id: int, lp_id: int, data: LPUpdate, db: Session = Depends(get_db)):
    lp = db.get(LP, lp_id)
    if not lp or lp.fund_id != fund_id:
        raise HTTPException(status_code=404, detail="LP를 찾을 수 없습니다")

    before = record_snapshot(lp)
    payload = data.model_dump(exclude_unset=True)
    if "name" in payload:
        payload["name"] = _normalize_lp_text(payload.get("name"))
    if "type" in payload:
        payload["type"] = coerce_lp_type(payload.get("type"))
    if "business_number" in payload:
        payload["business_number"] = _normalize_lp_optional(payload.get("business_number"))
    if "contact" in payload:
        payload["contact"] = _normalize_lp_optional(payload.get("contact"))
    if "address" in payload:
        payload["address"] = _normalize_lp_optional(payload.get("address"))

    if "address_book_id" in payload:
        address_book = _get_lp_address_book(db, payload.get("address_book_id"))
        if address_book:
            payload.setdefault("name", address_book.name)
            payload.setdefault("type", address_book.type)
            payload.setdefault("business_number", address_book.business_number)
            payload.setdefault("contact", address_book.contact)
            payload.setdefault("address", address_book.address)

    next_name = _normalize_lp_text(payload.get("name", lp.name))
    next_type = coerce_lp_type(payload.get("type", lp.type))
    next_business_number = _normalize_lp_optional(payload.get("business_number", lp.business_number))
    next_address_book_id = payload.get("address_book_id", lp.address_book_id)

    if not next_name or not next_type:
        raise HTTPException(status_code=400, detail="LP name and type are required")

    _ensure_lp_unique_in_fund(
        db,
        fund_id,
        name=next_name,
        business_number=next_business_number,
        address_book_id=next_address_book_id,
        current_lp_id=lp.id,
    )

    has_call_items, paid_in_total = calculate_lp_paid_in_from_calls(db, fund_id, lp_id)
    if has_call_items:
        payload["paid_in"] = paid_in_total

    next_commitment = payload.get("commitment", lp.commitment)
    next_paid_in = payload.get("paid_in", lp.paid_in)
    validate_lp_paid_in_pair(
        commitment=next_commitment,
        paid_in=next_paid_in,
    )

    for key, val in payload.items():
        setattr(lp, key, val)

    try:
        if backbone_enabled():
            db.flush()
            subject = sync_lp_graph(db, lp)
            maybe_emit_mutation(
                db,
                subject=subject,
                event_type="lp.updated",
                before=before,
                after=record_snapshot(lp),
                origin_model="lp",
                origin_id=lp.id,
            )
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="Duplicate LP in the same fund is not allowed") from exc
    db.refresh(lp)
    _run_compliance_rule_checks(
        db=db,
        fund_id=fund_id,
        trigger_source="lp_update",
        trigger_source_id=lp.id,
    )
    return lp


@router.delete("/api/funds/{fund_id}/lps/{lp_id}", status_code=204)
def delete_lp(fund_id: int, lp_id: int, db: Session = Depends(get_db)):
    lp = db.get(LP, lp_id)
    if not lp or lp.fund_id != fund_id:
        raise HTTPException(status_code=404, detail="LP를 찾을 수 없습니다")

    before = record_snapshot(lp)
    if backbone_enabled():
        subject = mark_subject_deleted(
            db,
            subject_type="lp",
            native_id=lp.id,
            payload=before,
        )
        maybe_emit_mutation(
            db,
            subject=subject,
            event_type="lp.deleted",
            before=before,
            origin_model="lp",
            origin_id=lp.id,
        )
    db.delete(lp)
    db.commit()

    try:
        ComplianceEngine(db).on_lp_changed(fund_id, "withdrawn")
    except Exception as exc:  # noqa: BLE001 - hook failures must not break main flow
        db.rollback()
        logger.warning(
            "compliance hook failed on delete_lp: fund_id=%s lp_id=%s error=%s",
            fund_id,
            lp_id,
            exc,
        )
    _run_compliance_rule_checks(
        db=db,
        fund_id=fund_id,
        trigger_source="lp_delete",
        trigger_source_id=lp_id,
    )


@router.put("/api/funds/{fund_id}/notice-periods", response_model=list[FundNoticePeriodResponse])
def replace_notice_periods(
    fund_id: int,
    data: list[FundNoticePeriodCreate],
    db: Session = Depends(get_db),
):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")

    before = [{"id": row.id, "notice_type": row.notice_type, "label": row.label, "business_days": row.business_days, "day_basis": row.day_basis, "memo": row.memo} for row in fund.notice_periods]
    fund.notice_periods.clear()
    for item in data:
        fund.notice_periods.append(
            FundNoticePeriod(
                notice_type=item.notice_type.strip(),
                label=item.label.strip(),
                business_days=item.business_days,
                day_basis=(item.day_basis or "business").strip().lower(),
                memo=item.memo.strip() if item.memo else None,
            )
        )

    if backbone_enabled():
        db.flush()
        subject = sync_fund_graph(db, fund)
        maybe_emit_mutation(
            db,
            subject=subject,
            event_type="fund.notice_periods.updated",
            before={"notice_periods": before},
            after={"notice_periods": [{"notice_type": row.notice_type, "label": row.label, "business_days": row.business_days, "day_basis": row.day_basis, "memo": row.memo} for row in fund.notice_periods]},
            origin_model="fund_notice_period",
            origin_id=fund.id,
        )
    db.commit()
    db.refresh(fund)
    return sorted(fund.notice_periods, key=lambda row: row.id)


@router.put("/api/funds/{fund_id}/key-terms", response_model=list[FundKeyTermResponse])
def replace_key_terms(
    fund_id: int,
    data: list[FundKeyTermCreate],
    db: Session = Depends(get_db),
):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")

    before = [{"id": row.id, "category": row.category, "label": row.label, "value": row.value, "article_ref": row.article_ref} for row in fund.key_terms]
    fund.key_terms.clear()
    for item in data:
        fund.key_terms.append(
            FundKeyTerm(
                category=item.category.strip(),
                label=item.label.strip(),
                value=item.value.strip(),
                article_ref=item.article_ref.strip() if item.article_ref else None,
            )
        )

    if backbone_enabled():
        db.flush()
        subject = sync_fund_graph(db, fund)
        maybe_emit_mutation(
            db,
            subject=subject,
            event_type="fund.key_terms.updated",
            before={"key_terms": before},
            after={"key_terms": [{"category": row.category, "label": row.label, "value": row.value, "article_ref": row.article_ref} for row in fund.key_terms]},
            origin_model="fund_key_term",
            origin_id=fund.id,
        )
    db.commit()
    db.refresh(fund)
    return sorted(fund.key_terms, key=lambda row: row.id)


@router.get("/api/funds/{fund_id}/notice-periods/{notice_type}", response_model=FundNoticePeriodResponse)
def get_notice_period(fund_id: int, notice_type: str, db: Session = Depends(get_db)):
    period = (
        db.query(FundNoticePeriod)
        .filter(
            FundNoticePeriod.fund_id == fund_id,
            FundNoticePeriod.notice_type == notice_type,
        )
        .first()
    )
    if not period:
        raise HTTPException(status_code=404, detail="통지기간을 찾을 수 없습니다")
    return period


@router.get("/api/funds/{fund_id}/calculate-deadline")
def calculate_deadline(
    fund_id: int,
    target_date: date,
    notice_type: str,
    db: Session = Depends(get_db),
):
    fund = db.get(Fund, fund_id)
    if not fund:
        raise HTTPException(status_code=404, detail="조합을 찾을 수 없습니다")

    period = (
        db.query(FundNoticePeriod)
        .filter(
            FundNoticePeriod.fund_id == fund_id,
            FundNoticePeriod.notice_type == notice_type,
        )
        .first()
    )
    if not period:
        raise HTTPException(status_code=404, detail="통지기간을 찾을 수 없습니다")

    day_basis = (period.day_basis or "business").strip().lower()
    if day_basis == "calendar":
        deadline = target_date - timedelta(days=period.business_days)
    else:
        deadline = calculate_business_days_before(target_date, period.business_days)
    return {
        "target_date": target_date.isoformat(),
        "notice_type": notice_type,
        "business_days": period.business_days,
        "notice_days": period.business_days,
        "day_basis": day_basis,
        "deadline": deadline.isoformat(),
        "label": period.label,
    }



