from __future__ import annotations

import calendar
import json
import re
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from models.accounting import Account, JournalEntry
from models.provisional_fs import ProvisionalFS


class FSExcelExporter:
    """Export provisional SFP/IS and monthly journal lines into an excel workbook."""

    def export(self, provisional_fs: ProvisionalFS, fund, db: Session) -> str:
        year_month = provisional_fs.year_month
        year, month = self._parse_year_month(year_month)
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])

        sfp_data = self._load_json(provisional_fs.sfp_data)
        is_data = self._load_json(provisional_fs.is_data)

        workbook = self._load_template()

        journal_ws = self._ensure_sheet(workbook, "가결산전표")
        sfp_ws = self._ensure_sheet(workbook, "SFP")
        is_ws = self._ensure_sheet(workbook, "IS")

        self._write_journal_sheet(journal_ws, provisional_fs.fund_id, month_start, month_end, db)
        self._write_sfp_sheet(sfp_ws, sfp_data, month_end)
        self._write_is_sheet(is_ws, is_data, month_start, month_end)

        output_dir = Path(__file__).resolve().parents[1] / "generated" / "provisional_fs"
        output_dir.mkdir(parents=True, exist_ok=True)

        fund_name = getattr(fund, "name", "fund") if fund is not None else "fund"
        safe_name = re.sub(r"[^0-9A-Za-z가-힣._-]+", "_", fund_name).strip("_") or "fund"
        filename = f"{safe_name}_{year_month}_provisional_fs.xlsx"
        output_path = output_dir / filename
        workbook.save(output_path)
        return str(output_path)

    def _load_template(self):
        templates_dir = Path(__file__).resolve().parents[2] / "templates"
        candidates = sorted(templates_dir.glob("*가결산FS*.xlsx"))
        if candidates:
            return load_workbook(candidates[0])
        return Workbook()

    def _ensure_sheet(self, workbook, name: str):
        if name in workbook.sheetnames:
            return workbook[name]
        if len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet":
            workbook.active.title = name
            return workbook.active
        return workbook.create_sheet(name)

    def _write_journal_sheet(
        self,
        ws: Worksheet,
        fund_id: int,
        month_start: date,
        month_end: date,
        db: Session,
    ) -> None:
        entries = (
            db.query(JournalEntry)
            .filter(
                JournalEntry.fund_id == fund_id,
                JournalEntry.entry_date >= month_start,
                JournalEntry.entry_date <= month_end,
            )
            .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc())
            .all()
        )

        account_map = {row.id: row for row in db.query(Account).all()}

        # Clear input area in template.
        for row in range(7, 74):
            for col in range(2, 8):
                ws.cell(row=row, column=col).value = None

        current_row = 7
        for entry in entries:
            if current_row > 73:
                break

            debit_lines = [line for line in entry.lines if float(line.debit or 0) > 0]
            credit_lines = [line for line in entry.lines if float(line.credit or 0) > 0]

            debit_name = ""
            credit_name = ""
            amount = 0.0

            if debit_lines:
                top_debit = sorted(debit_lines, key=lambda row: float(row.debit or 0), reverse=True)[0]
                account = account_map.get(top_debit.account_id)
                debit_name = account.name if account else str(top_debit.account_id)
                amount = max(amount, float(top_debit.debit or 0))

            if credit_lines:
                top_credit = sorted(credit_lines, key=lambda row: float(row.credit or 0), reverse=True)[0]
                account = account_map.get(top_credit.account_id)
                credit_name = account.name if account else str(top_credit.account_id)
                amount = max(amount, float(top_credit.credit or 0))

            if amount == 0:
                total_debit = sum(float(line.debit or 0) for line in entry.lines)
                total_credit = sum(float(line.credit or 0) for line in entry.lines)
                amount = max(total_debit, total_credit)

            ws.cell(row=current_row, column=2).value = datetime.combine(entry.entry_date, time.min)
            ws.cell(row=current_row, column=3).value = debit_name
            ws.cell(row=current_row, column=4).value = amount
            ws.cell(row=current_row, column=5).value = credit_name
            ws.cell(row=current_row, column=6).value = amount
            ws.cell(row=current_row, column=7).value = entry.description or ""

            current_row += 1

    def _write_sfp_sheet(self, ws: Worksheet, sfp_data: dict, month_end: date) -> None:
        col = self._ensure_month_column(ws, month_end, income_statement=False)

        self._set_by_label(ws, "유동자산", col, sfp_data.get("current_assets", 0.0))
        self._set_by_label(ws, "창업투자자산", col, sfp_data.get("investment_assets", 0.0))
        self._set_by_label(ws, "경영지원자산", col, sfp_data.get("non_current_assets", 0.0))
        self._set_by_label(ws, "비유동자산", col, sfp_data.get("non_current_assets", 0.0))
        self._set_by_label(ws, "자산총계", col, sfp_data.get("total_assets", 0.0))

        self._set_by_label(ws, "유동부채", col, sfp_data.get("current_liabilities", 0.0))
        self._set_by_label(ws, "비유동부채", col, sfp_data.get("non_current_liabilities", 0.0))
        self._set_by_label(ws, "부채총계", col, sfp_data.get("total_liabilities", 0.0))

        self._set_by_label(ws, "출자금", col, sfp_data.get("capital", 0.0))
        self._set_by_label(ws, "자본총계", col, sfp_data.get("total_equity", 0.0))

        retained = float(sfp_data.get("retained_earnings", 0.0) or 0)
        if retained < 0:
            self._set_by_label(ws, "결손금", col, abs(retained))
        else:
            self._set_by_label(ws, "결손금", col, 0.0)
            self._set_by_label(ws, "이익잉여금", col, retained)

    def _write_is_sheet(self, ws: Worksheet, is_data: dict, month_start: date, month_end: date) -> None:
        col = self._ensure_month_column(ws, month_end, income_statement=True)

        # IS has two date rows in template: row4(start), row5(end)
        ws.cell(row=4, column=col).value = datetime.combine(month_start, time.min)
        ws.cell(row=5, column=col).value = datetime.combine(month_end, time.min)

        self._set_by_label(ws, "투자수익", col, is_data.get("investment_income", 0.0))
        self._set_by_label(ws, "운용투자수익", col, is_data.get("operating_invest_income", 0.0))
        self._set_by_label(ws, "기타의 영업수익", col, is_data.get("other_operating_income", 0.0))
        self._set_by_label(ws, "영업수익", col, is_data.get("operating_revenue", 0.0))

        self._set_by_label(ws, "관리보수", col, is_data.get("management_fee", 0.0))
        self._set_by_label(ws, "성과보수", col, is_data.get("performance_fee", 0.0))
        self._set_by_label(ws, "수탁보수", col, is_data.get("trustee_fee", 0.0))
        self._set_by_label(ws, "감사수수료", col, is_data.get("audit_fee", 0.0))
        self._set_by_label(ws, "투자비용", col, is_data.get("investment_expense", 0.0))
        self._set_by_label(ws, "기타영업비용", col, is_data.get("other_operating_expense", 0.0))
        self._set_by_label(ws, "영업비용", col, is_data.get("operating_expense", 0.0))

        self._set_by_label(ws, "영업이익", col, is_data.get("operating_income", 0.0))
        self._set_by_label(ws, "당기순이익", col, is_data.get("net_income", 0.0))

    def _ensure_month_column(self, ws: Worksheet, month_end: date, *, income_statement: bool) -> int:
        # Template monthly columns start at E(5).
        target_dt = datetime.combine(month_end, time.min)
        date_row = 5 if income_statement else 4

        last_filled_col = 5
        for col in range(5, max(ws.max_column, 5) + 1):
            value = ws.cell(row=date_row, column=col).value
            if value in (None, ""):
                continue
            last_filled_col = col
            if isinstance(value, datetime):
                if value.date() == month_end:
                    return col
            elif isinstance(value, date):
                if value == month_end:
                    return col

        next_col = last_filled_col + 1
        ws.cell(row=date_row, column=next_col).value = target_dt
        return next_col

    def _set_by_label(self, ws: Worksheet, label: str, col: int, value: float) -> None:
        row = self._find_row_by_label(ws, label)
        if not row:
            return
        ws.cell(row=row, column=col).value = float(value or 0)

    def _find_row_by_label(self, ws: Worksheet, label: str) -> int | None:
        for row in range(1, ws.max_row + 1):
            text = ws.cell(row=row, column=4).value
            if isinstance(text, str) and label in text:
                return row
        return None

    def _load_json(self, raw: str | None) -> dict:
        if not raw:
            return {}
        try:
            value = json.loads(raw)
            if isinstance(value, dict):
                return value
        except Exception:  # noqa: BLE001
            pass
        return {}

    def _parse_year_month(self, year_month: str) -> tuple[int, int]:
        year_s, month_s = year_month.split("-", 1)
        return int(year_s), int(month_s)
