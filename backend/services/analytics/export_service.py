from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from schemas.analytics import AnalyticsQueryRequest, AnalyticsQueryResponse


HEADER_FILL = PatternFill(fill_type="solid", fgColor="DCE8FF")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="F5F9FF")
BOLD_FONT = Font(bold=True)


def export_query_to_xlsx(query: AnalyticsQueryRequest, response: AnalyticsQueryResponse) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "분석결과"

    if response.meta.mode == "table":
        headers = [field.label for field in response.table_fields]
        keys = [field.key for field in response.table_fields]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = BOLD_FONT
            cell.fill = HEADER_FILL
        for row in response.rows:
            ws.append([row.get(key) for key in keys])
    else:
        dim_keys = [field.key for field in response.row_fields] + [field.key for field in response.column_fields]
        measure_keys = [field.key for field in response.value_fields]
        headers = [field.label for field in response.row_fields] + [field.label for field in response.column_fields] + [field.label for field in response.value_fields]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = BOLD_FONT
            cell.fill = HEADER_FILL
        for row in response.rows:
            ws.append([row.get(key) for key in dim_keys + measure_keys])
        if response.grand_totals:
            ws.append([])
            ws.append(["총계"] + [None] * (len(dim_keys) - 1) + [response.grand_totals.get(key) for key in measure_keys])
            total_row_idx = ws.max_row
            for cell in ws[total_row_idx]:
                cell.font = BOLD_FONT
                cell.fill = SUBHEADER_FILL

    desc = wb.create_sheet("조건설명")
    desc.append(["항목", "값"])
    desc[1][0].font = BOLD_FONT
    desc[1][1].font = BOLD_FONT
    desc[1][0].fill = HEADER_FILL
    desc[1][1].fill = HEADER_FILL
    desc.append(["Subject", query.subject_key])
    desc.append(["Mode", query.mode])
    desc.append(["Rows", ", ".join(query.rows)])
    desc.append(["Columns", ", ".join(query.columns)])
    desc.append(["Values", ", ".join(f"{item.key}:{item.aggregate or ''}" for item in query.values)])
    desc.append(["Selected Fields", ", ".join(query.selected_fields)])
    desc.append(["Filters", _stringify(query.filters)])
    desc.append(["Sorts", _stringify(query.sorts)])

    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                max_length = max(max_length, len(str(cell.value or "")))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _stringify(rows: list[Any]) -> str:
    parts = []
    for item in rows:
        if hasattr(item, "model_dump"):
            parts.append(str(item.model_dump()))
        else:
            parts.append(str(item))
    return " | ".join(parts)

