from __future__ import annotations

import csv
import json
from datetime import date, datetime
from io import BytesIO
from io import StringIO
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from models.fund import Fund, LP
from models.gp_entity import GPEntity
from models.investment import Investment, PortfolioCompany
from models.lp_contribution import LPContribution
from models.phase3 import CapitalCall, CapitalCallItem, Distribution, ExitTrade
from models.proposal_data import (
    FundHistory,
    ProposalApplication,
    ProposalApplicationFund,
    ProposalFieldOverride,
    ProposalRowOverride,
    ProposalSnapshot,
    FundManager,
    FundManagerHistory,
    FundManagerProfileHistory,
    FundSubscription,
    GPEntityHistory,
    GPFinancial,
    GPShareholder,
    ManagerAward,
    ManagerCareer,
    ManagerEducation,
    ManagerInvestment,
    PortfolioCompanyHistory,
    ProposalVersion,
)
from models.transaction import Transaction
from models.user import User
from models.valuation import Valuation
from services.lp_types import normalize_lp_type

_BASELINE_DATE = date(1900, 1, 1)
_HEADER_FILL = PatternFill(fill_type="solid", start_color="1F6FB8", end_color="1F6FB8")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_CELL_BORDER = Border(
    left=Side(style="thin", color="D9E2EF"),
    right=Side(style="thin", color="D9E2EF"),
    top=Side(style="thin", color="D9E2EF"),
    bottom=Side(style="thin", color="D9E2EF"),
)


def _json_default(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    return round(float(value), 2)


def _safe_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            return None
    return None


def _serialize_gp_entity(entity: GPEntity) -> dict[str, Any]:
    return {
        "id": entity.id,
        "name": entity.name,
        "entity_type": entity.entity_type,
        "business_number": entity.business_number,
        "registration_number": entity.registration_number,
        "representative": entity.representative,
        "address": entity.address,
        "phone": entity.phone,
        "email": entity.email,
        "founding_date": entity.founding_date,
        "license_date": entity.license_date,
        "capital": entity.capital,
        "total_employees": entity.total_employees,
        "fund_manager_count": entity.fund_manager_count,
        "paid_in_capital": entity.paid_in_capital,
        "notes": entity.notes,
        "is_primary": entity.is_primary,
    }


def _serialize_fund(fund: Fund) -> dict[str, Any]:
    return {
        "id": fund.id,
        "name": fund.name,
        "type": fund.type,
        "business_number": fund.business_number,
        "formation_date": fund.formation_date,
        "registration_number": fund.registration_number,
        "registration_date": fund.registration_date,
        "status": fund.status,
        "regulation_type": fund.regulation_type,
        "setup_type": fund.setup_type,
        "gp_entity_id": fund.gp_entity_id,
        "gp": fund.gp,
        "fund_manager": fund.fund_manager,
        "co_gp": fund.co_gp,
        "has_co_gp": bool(fund.has_co_gp),
        "trustee": fund.trustee,
        "commitment_total": fund.commitment_total,
        "gp_commitment": fund.gp_commitment,
        "contribution_type": fund.contribution_type,
        "aum": fund.aum,
        "investment_period_end": fund.investment_period_end,
        "maturity_date": fund.maturity_date,
        "dissolution_date": fund.dissolution_date,
        "mgmt_fee_rate": fund.mgmt_fee_rate,
        "performance_fee_rate": fund.performance_fee_rate,
        "hurdle_rate": fund.hurdle_rate,
        "account_number": fund.account_number,
    }


def _serialize_company(company: PortfolioCompany) -> dict[str, Any]:
    return {
        "id": company.id,
        "name": company.name,
        "business_number": company.business_number,
        "ceo": company.ceo,
        "address": company.address,
        "industry": company.industry,
        "industry_code": company.industry_code,
        "vics_registered": company.vics_registered,
        "corp_number": company.corp_number,
        "founded_date": company.founded_date,
        "analyst": company.analyst,
        "contact_name": company.contact_name,
        "contact_email": company.contact_email,
        "contact_phone": company.contact_phone,
        "memo": company.memo,
    }


def _serialize_manager(manager: FundManager) -> dict[str, Any]:
    return {
        "id": manager.id,
        "gp_entity_id": manager.gp_entity_id,
        "name": manager.name,
        "birth_date": manager.birth_date,
        "nationality": manager.nationality,
        "phone": manager.phone,
        "fax": manager.fax,
        "email": manager.email,
        "department": manager.department,
        "position": manager.position,
        "join_date": manager.join_date,
        "resign_date": manager.resign_date,
        "is_core": bool(manager.is_core),
        "is_representative": bool(manager.is_representative),
    }


def _dump_snapshot(snapshot: dict[str, Any]) -> str:
    return json.dumps(snapshot, ensure_ascii=False, default=_json_default, sort_keys=True)


def _sync_history(
    db: Session,
    *,
    history_model: Any,
    key_attr: str,
    source_id: int,
    snapshot: dict[str, Any],
    effective_date: date | None = None,
    baseline_date: date | None = None,
) -> None:
    field = getattr(history_model, key_attr)
    snapshot_json = _dump_snapshot(snapshot)
    rows = (
        db.query(history_model)
        .filter(field == source_id)
        .order_by(history_model.valid_from.asc(), history_model.id.asc())
        .all()
    )
    if not rows:
        db.add(
            history_model(
                **{
                    key_attr: source_id,
                    "valid_from": baseline_date or _BASELINE_DATE,
                    "valid_to": None,
                    "snapshot_json": snapshot_json,
                }
            )
        )
        return

    latest = rows[-1]
    if latest.snapshot_json == snapshot_json:
        return

    next_valid_from = effective_date or date.today()
    latest.valid_to = next_valid_from
    db.add(
        history_model(
            **{
                key_attr: source_id,
                "valid_from": next_valid_from,
                "valid_to": None,
                "snapshot_json": snapshot_json,
            }
        )
    )


def sync_gp_entity_history(db: Session, entity: GPEntity, effective_date: date | None = None) -> None:
    _sync_history(
        db,
        history_model=GPEntityHistory,
        key_attr="gp_entity_id",
        source_id=entity.id,
        snapshot=_serialize_gp_entity(entity),
        effective_date=effective_date,
        baseline_date=entity.founding_date or _BASELINE_DATE,
    )


def sync_fund_history(db: Session, fund: Fund, effective_date: date | None = None) -> None:
    _sync_history(
        db,
        history_model=FundHistory,
        key_attr="fund_id",
        source_id=fund.id,
        snapshot=_serialize_fund(fund),
        effective_date=effective_date,
        baseline_date=fund.formation_date or _BASELINE_DATE,
    )


def sync_company_history(db: Session, company: PortfolioCompany, effective_date: date | None = None) -> None:
    _sync_history(
        db,
        history_model=PortfolioCompanyHistory,
        key_attr="company_id",
        source_id=company.id,
        snapshot=_serialize_company(company),
        effective_date=effective_date,
        baseline_date=company.founded_date or _BASELINE_DATE,
    )


def sync_fund_manager_history(db: Session, manager: FundManager, effective_date: date | None = None) -> None:
    _sync_history(
        db,
        history_model=FundManagerProfileHistory,
        key_attr="fund_manager_id",
        source_id=manager.id,
        snapshot=_serialize_manager(manager),
        effective_date=effective_date,
        baseline_date=manager.join_date or _BASELINE_DATE,
    )


def _resolve_history_snapshot(
    db: Session,
    *,
    history_model: Any,
    key_attr: str,
    source_id: int,
    as_of_date: date,
    fallback_snapshot: dict[str, Any],
) -> dict[str, Any]:
    field = getattr(history_model, key_attr)
    row = (
        db.query(history_model)
        .filter(
            field == source_id,
            history_model.valid_from <= as_of_date,
            or_(history_model.valid_to.is_(None), history_model.valid_to > as_of_date),
        )
        .order_by(history_model.valid_from.desc(), history_model.id.desc())
        .first()
    )
    if not row:
        return fallback_snapshot
    try:
        parsed = json.loads(row.snapshot_json)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        return fallback_snapshot
    return fallback_snapshot


def _fund_matches_gp(fund: Fund, selected_gp: GPEntity | None) -> bool:
    if selected_gp is None:
        return True

    if fund.gp_entity_id is not None:
        return fund.gp_entity_id == selected_gp.id

    normalized_gp = (selected_gp.name or "").strip()
    if not normalized_gp:
        return True

    targets = [fund.gp or "", fund.co_gp or ""]
    return any(normalized_gp == target.strip() for target in targets if target)


def _calculate_paid_in_as_of(db: Session, fund_id: int, as_of_date: date) -> float:
    contribution_count = (
        db.query(func.count(LPContribution.id))
        .filter(LPContribution.fund_id == fund_id)
        .scalar()
    )
    if int(contribution_count or 0) > 0:
        paid_in_total = (
            db.query(func.coalesce(func.sum(LPContribution.amount), 0.0))
            .filter(
                LPContribution.fund_id == fund_id,
                LPContribution.actual_paid_date.isnot(None),
                LPContribution.actual_paid_date <= as_of_date,
            )
            .scalar()
        )
        return _to_float(paid_in_total)

    paid_from_calls = (
        db.query(func.coalesce(func.sum(CapitalCallItem.amount), 0.0))
        .join(CapitalCall, CapitalCall.id == CapitalCallItem.capital_call_id)
        .filter(
            CapitalCall.fund_id == fund_id,
            CapitalCallItem.paid == 1,
            CapitalCallItem.paid_date.isnot(None),
            CapitalCallItem.paid_date <= as_of_date,
        )
        .scalar()
    )
    if paid_from_calls:
        return _to_float(paid_from_calls)

    fallback = (
        db.query(func.coalesce(func.sum(LP.paid_in), 0.0))
        .filter(LP.fund_id == fund_id)
        .scalar()
    )
    return _to_float(fallback)


def _calculate_invested_as_of(db: Session, fund_id: int, as_of_date: date) -> float:
    total = (
        db.query(func.coalesce(func.sum(Investment.amount), 0.0))
        .filter(
            Investment.fund_id == fund_id,
            Investment.investment_date.isnot(None),
            Investment.investment_date <= as_of_date,
        )
        .scalar()
    )
    return _to_float(total)


def _calculate_exit_total_as_of(db: Session, fund_id: int, as_of_date: date) -> float:
    total = (
        db.query(func.coalesce(func.sum(ExitTrade.amount), 0.0))
        .filter(
            ExitTrade.fund_id == fund_id,
            ExitTrade.trade_date <= as_of_date,
        )
        .scalar()
    )
    return _to_float(total)


def _calculate_nav_as_of(db: Session, fund_id: int, as_of_date: date) -> float:
    valuation_rows = (
        db.query(Valuation)
        .filter(
            Valuation.fund_id == fund_id,
            Valuation.as_of_date <= as_of_date,
        )
        .order_by(Valuation.investment_id.asc(), Valuation.as_of_date.desc(), Valuation.id.desc())
        .all()
    )
    if not valuation_rows:
        return round(max(_calculate_invested_as_of(db, fund_id, as_of_date) - _calculate_exit_total_as_of(db, fund_id, as_of_date), 0.0), 2)

    latest_by_investment: dict[int, float] = {}
    for row in valuation_rows:
        if row.investment_id in latest_by_investment:
            continue
        latest_by_investment[row.investment_id] = _to_float(row.value)
    return _to_float(sum(latest_by_investment.values()))


def _build_fund_snapshot(db: Session, fund: Fund, as_of_date: date) -> dict[str, Any]:
    fund_snapshot = _resolve_history_snapshot(
        db,
        history_model=FundHistory,
        key_attr="fund_id",
        source_id=fund.id,
        as_of_date=as_of_date,
        fallback_snapshot=_serialize_fund(fund),
    )
    return {
        "id": fund.id,
        "name": fund_snapshot.get("name", fund.name),
        "type": fund_snapshot.get("type", fund.type),
        "status": fund_snapshot.get("status", fund.status),
        "gp_entity_id": fund_snapshot.get("gp_entity_id", fund.gp_entity_id),
        "gp": fund_snapshot.get("gp", fund.gp),
        "co_gp": fund_snapshot.get("co_gp", fund.co_gp),
        "has_co_gp": bool(fund_snapshot.get("has_co_gp", fund.has_co_gp or bool(fund.co_gp))),
        "fund_manager": fund_snapshot.get("fund_manager", fund.fund_manager),
        "formation_date": _safe_date(fund_snapshot.get("formation_date")) or fund.formation_date,
        "registration_date": _safe_date(fund_snapshot.get("registration_date")) or fund.registration_date,
        "investment_period_end": _safe_date(fund_snapshot.get("investment_period_end")) or fund.investment_period_end,
        "maturity_date": _safe_date(fund_snapshot.get("maturity_date")) or fund.maturity_date,
        "commitment_total": fund_snapshot.get("commitment_total", fund.commitment_total),
        "paid_in_total": _calculate_paid_in_as_of(db, fund.id, as_of_date),
        "invested_total": _calculate_invested_as_of(db, fund.id, as_of_date),
        "exit_total": _calculate_exit_total_as_of(db, fund.id, as_of_date),
        "nav_total": _calculate_nav_as_of(db, fund.id, as_of_date),
    }


def _build_readiness(
    *,
    template_type: str,
    as_of_date: date,
    selected_gp_entity: dict[str, Any] | None,
    funds: list[dict[str, Any]],
    gp_financials: list[GPFinancial],
    gp_shareholders: list[GPShareholder],
    fund_managers: list[FundManager],
    subscriptions: list[FundSubscription],
) -> dict[str, Any]:
    missing_items: list[str] = []
    warnings: list[str] = []

    if selected_gp_entity is None:
        missing_items.append("GP를 선택해야 합니다.")
    if not funds:
        missing_items.append("대상 조합을 1개 이상 선택해야 합니다.")
    if selected_gp_entity is not None:
        if not selected_gp_entity.get("business_number"):
            missing_items.append("GP 사업자번호가 없습니다.")
        if not selected_gp_entity.get("address"):
            missing_items.append("GP 주소가 없습니다.")
        if selected_gp_entity.get("total_employees") in (None, 0):
            warnings.append("GP 임직원 수가 비어 있습니다.")
        if selected_gp_entity.get("fund_manager_count") in (None, 0):
            warnings.append("GP 운용인력 수가 비어 있습니다.")

    if not gp_financials:
        warnings.append("기준일 이하 GP 재무현황이 없습니다.")
    if not gp_shareholders:
        warnings.append("기준일 이하 GP 주주현황이 없습니다.")
    if not fund_managers:
        warnings.append("GP 소속 운용인력이 없습니다.")
    if template_type in {"motae-5", "nong-motae"} and not subscriptions:
        warnings.append("기준일 이하 청약이력이 없습니다.")

    return {
        "as_of_date": as_of_date,
        "template_type": template_type,
        "is_ready": len(missing_items) == 0,
        "missing_items": missing_items,
        "warnings": warnings,
    }


def _build_metrics(summary: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {"label": "대상 조합", "value": f"{summary['selected_fund_count']}개", "hint": "현재 컨텍스트 기준"},
        {"label": "운용인력", "value": f"{summary['selected_manager_count']}명", "hint": "선택 GP 소속"},
        {"label": "총 약정", "value": f"₩{summary['total_commitment']:,.0f}", "hint": "선택 조합 합계"},
        {"label": "누적 납입", "value": f"₩{summary['total_paid_in']:,.0f}", "hint": "기준일 이하"},
        {"label": "투자 집행", "value": f"₩{summary['total_invested']:,.0f}", "hint": "기준일 이하"},
        {"label": "회수 금액", "value": f"₩{summary['total_exit_amount']:,.0f}", "hint": "기준일 이하"},
    ]


def version_to_response(version: ProposalVersion) -> dict[str, Any]:
    try:
        fund_ids = json.loads(version.fund_ids_json or "[]")
    except json.JSONDecodeError:
        fund_ids = []
    if not isinstance(fund_ids, list):
        fund_ids = []
    return {
        "id": version.id,
        "template_type": version.template_type,
        "gp_entity_id": version.gp_entity_id,
        "fund_ids": [int(value) for value in fund_ids if isinstance(value, int)],
        "as_of_date": version.as_of_date,
        "status": version.status,
        "render_snapshot_json": version.render_snapshot_json,
        "generated_filename": version.generated_filename,
        "created_by": version.created_by,
        "created_at": version.created_at,
        "updated_at": version.updated_at,
    }


def resolve_proposal_workspace(
    db: Session,
    *,
    template_type: str,
    as_of_date: date,
    gp_entity_id: int | None = None,
    fund_ids: list[int] | None = None,
    version: ProposalVersion | None = None,
    auto_select_default_gp: bool = True,
    auto_select_all_funds: bool = True,
) -> dict[str, Any]:
    gp_entities = db.query(GPEntity).order_by(GPEntity.is_primary.desc(), GPEntity.id.asc()).all()
    selected_gp = db.get(GPEntity, gp_entity_id) if gp_entity_id is not None else (gp_entities[0] if auto_select_default_gp and gp_entities else None)

    selected_gp_snapshot = None
    if selected_gp is not None:
        selected_gp_snapshot = _resolve_history_snapshot(
            db,
            history_model=GPEntityHistory,
            key_attr="gp_entity_id",
            source_id=selected_gp.id,
            as_of_date=as_of_date,
            fallback_snapshot=_serialize_gp_entity(selected_gp),
        )

    all_funds = db.query(Fund).order_by(Fund.name.asc(), Fund.id.asc()).all()
    candidate_funds = [fund for fund in all_funds if _fund_matches_gp(fund, selected_gp)]
    selected_fund_ids = [int(value) for value in (fund_ids or [])]
    if not selected_fund_ids and candidate_funds and auto_select_all_funds:
        selected_fund_ids = [fund.id for fund in candidate_funds]
    selected_funds = [fund for fund in candidate_funds if fund.id in set(selected_fund_ids)]
    candidate_fund_snapshots = [_build_fund_snapshot(db, fund, as_of_date) for fund in candidate_funds]
    fund_snapshots = [_build_fund_snapshot(db, fund, as_of_date) for fund in selected_funds]

    gp_financials = []
    gp_shareholders = []
    managers = []
    if selected_gp is not None:
        gp_financials = db.query(GPFinancial).filter(GPFinancial.gp_entity_id == selected_gp.id, GPFinancial.fiscal_year_end <= as_of_date).order_by(GPFinancial.fiscal_year_end.desc(), GPFinancial.id.desc()).all()
        gp_shareholders = db.query(GPShareholder).filter(GPShareholder.gp_entity_id == selected_gp.id, GPShareholder.snapshot_date <= as_of_date).order_by(GPShareholder.snapshot_date.desc(), GPShareholder.id.desc()).all()
        managers = db.query(FundManager).filter(FundManager.gp_entity_id == selected_gp.id).order_by(FundManager.is_representative.desc(), FundManager.is_core.desc(), FundManager.name.asc()).all()

    manager_ids = [manager.id for manager in managers]
    manager_careers = db.query(ManagerCareer).filter(ManagerCareer.fund_manager_id.in_(manager_ids)).filter(or_(ManagerCareer.start_date.is_(None), ManagerCareer.start_date <= as_of_date)).order_by(ManagerCareer.start_date.desc().nullslast(), ManagerCareer.id.desc()).all() if manager_ids else []
    manager_educations = db.query(ManagerEducation).filter(ManagerEducation.fund_manager_id.in_(manager_ids)).order_by(ManagerEducation.graduation_date.desc().nullslast(), ManagerEducation.id.desc()).all() if manager_ids else []
    manager_awards = db.query(ManagerAward).filter(ManagerAward.fund_manager_id.in_(manager_ids)).filter(or_(ManagerAward.award_date.is_(None), ManagerAward.award_date <= as_of_date)).order_by(ManagerAward.award_date.desc().nullslast(), ManagerAward.id.desc()).all() if manager_ids else []
    manager_investments = db.query(ManagerInvestment).filter(ManagerInvestment.fund_manager_id.in_(manager_ids)).filter(or_(ManagerInvestment.investment_date.is_(None), ManagerInvestment.investment_date <= as_of_date)).order_by(ManagerInvestment.investment_date.desc().nullslast(), ManagerInvestment.id.desc()).all() if manager_ids else []
    manager_histories = db.query(FundManagerHistory).filter(FundManagerHistory.fund_manager_id.in_(manager_ids), FundManagerHistory.change_date <= as_of_date).order_by(FundManagerHistory.change_date.desc(), FundManagerHistory.id.desc()).all() if manager_ids else []
    subscriptions = db.query(FundSubscription).filter(FundSubscription.fund_id.in_([fund.id for fund in selected_funds]), FundSubscription.subscription_date <= as_of_date).order_by(FundSubscription.subscription_date.desc(), FundSubscription.id.desc()).all() if selected_funds else []

    summary = {
        "selected_fund_count": len(fund_snapshots),
        "selected_manager_count": len(managers),
        "total_commitment": round(sum(_to_float(fund.get("commitment_total")) for fund in fund_snapshots), 2),
        "total_paid_in": round(sum(_to_float(fund.get("paid_in_total")) for fund in fund_snapshots), 2),
        "total_invested": round(sum(_to_float(fund.get("invested_total")) for fund in fund_snapshots), 2),
        "total_exit_amount": round(sum(_to_float(fund.get("exit_total")) for fund in fund_snapshots), 2),
    }
    readiness = _build_readiness(
        template_type=template_type,
        as_of_date=as_of_date,
        selected_gp_entity=selected_gp_snapshot,
        funds=fund_snapshots,
        gp_financials=gp_financials,
        gp_shareholders=gp_shareholders,
        fund_managers=managers,
        subscriptions=subscriptions,
    )

    render_snapshot = None
    if version is not None and version.render_snapshot_json:
        try:
            render_snapshot = json.loads(version.render_snapshot_json)
        except json.JSONDecodeError:
            render_snapshot = None

    return {
        "template_type": template_type,
        "as_of_date": as_of_date,
        "gp_entities": [_serialize_gp_entity(row) for row in gp_entities],
        "available_funds": candidate_fund_snapshots,
        "funds": fund_snapshots,
        "selected_gp_entity": selected_gp_snapshot,
        "selected_fund_ids": [fund["id"] for fund in fund_snapshots],
        "summary": summary,
        "metrics": _build_metrics(summary),
        "gp_financials": gp_financials,
        "gp_shareholders": gp_shareholders,
        "fund_managers": managers,
        "manager_careers": manager_careers,
        "manager_educations": manager_educations,
        "manager_awards": manager_awards,
        "manager_investments": manager_investments,
        "fund_manager_histories": manager_histories,
        "fund_subscriptions": subscriptions,
        "readiness": readiness,
        "version": version_to_response(version) if version is not None else None,
        "render_snapshot": render_snapshot,
    }


def create_proposal_version_snapshot(
    db: Session,
    *,
    template_type: str,
    as_of_date: date,
    gp_entity_id: int | None,
    fund_ids: list[int],
    created_by_user: User | None,
) -> ProposalVersion:
    version = ProposalVersion(
        template_type=template_type,
        gp_entity_id=gp_entity_id,
        fund_ids_json=json.dumps(fund_ids, ensure_ascii=False),
        as_of_date=as_of_date,
        status="draft",
        created_by=(created_by_user.id if created_by_user else None),
    )
    db.add(version)
    db.flush()
    return version


def freeze_proposal_version(
    db: Session,
    *,
    version: ProposalVersion,
) -> ProposalVersion:
    workspace = resolve_proposal_workspace(
        db,
        template_type=version.template_type,
        as_of_date=version.as_of_date,
        gp_entity_id=version.gp_entity_id,
        fund_ids=json.loads(version.fund_ids_json or "[]"),
        version=version,
    )
    version.render_snapshot_json = json.dumps(workspace, ensure_ascii=False, default=_json_default)
    version.status = "frozen"
    version.updated_at = datetime.utcnow()
    return version


def _apply_header(ws: Any, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_fit(ws: Any) -> None:
    for col in ws.columns:
        length = 10
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value) + 2, 48))
            cell.border = _CELL_BORDER
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[col_letter].width = length


def _write_sheet(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    _apply_header(ws, headers)
    for row in rows:
        ws.append(row)
    _auto_fit(ws)


def _legacy_growth_finance_sheets(workspace: dict[str, Any], db: Session) -> list[tuple[str, list[str], list[list[Any]]]]:
    gp = workspace.get("selected_gp_entity") or {}
    funds = workspace.get("funds", [])
    gp_financials: list[GPFinancial] = workspace.get("gp_financials", [])
    managers: list[FundManager] = workspace.get("fund_managers", [])
    careers: list[ManagerCareer] = workspace.get("manager_careers", [])
    educations: list[ManagerEducation] = workspace.get("manager_educations", [])
    manager_investments: list[ManagerInvestment] = workspace.get("manager_investments", [])
    manager_histories: list[FundManagerHistory] = workspace.get("fund_manager_histories", [])
    as_of_date: date = workspace["as_of_date"]
    fund_ids = [fund["id"] for fund in funds]
    lps = db.query(LP).filter(LP.fund_id.in_(fund_ids)).all() if fund_ids else []
    transactions = (
        db.query(Transaction)
        .filter(Transaction.fund_id.in_(fund_ids), Transaction.transaction_date <= as_of_date)
        .all()
        if fund_ids else []
    )
    core_manager_ids = {manager.id for manager in managers if manager.is_core}
    general_manager_ids = {manager.id for manager in managers if manager.id not in core_manager_ids}

    return [
        ("작성안내", ["항목", "값"], [["기준일", as_of_date.isoformat()], ["안내", "기준일 컨텍스트로 생성된 제안서 데이터입니다."]]),
        ("1.1.운용펀드", ["FundName", "FormationYMD", "MaturityYMD", "CommitmentTotal", "PaidInTotal", "FundTypeDesc", "FundStatusDesc"], [
            [fund["name"], _json_default(fund["formation_date"]), _json_default(fund["maturity_date"]), fund["commitment_total"], fund["paid_in_total"], fund["type"], fund["status"]]
            for fund in funds
        ]),
        ("1.2.투자연혁내역", ["FundMemberType", "FundMemberName", "FundingAmt"], [
            [normalize_lp_type(lp.type) or lp.type, lp.name, lp.commitment] for lp in lps
        ]),
        ("2.1.운용사 개요", ["GPNAME", "CompanyNo", "CompanySSN", "CEO", "FoundationYMD", "Address", "TotalEmployees", "FundMgrCnt", "PaidInCapital"], [[
            gp.get("name"),
            gp.get("registration_number"),
            gp.get("business_number"),
            gp.get("representative"),
            _json_default(gp.get("founding_date")),
            gp.get("address"),
            gp.get("total_employees"),
            gp.get("fund_manager_count"),
            gp.get("paid_in_capital"),
        ]]),
        ("2.2.운용사 재무현황", ["FYearEndYMD", "Assets", "CurrentAssets", "Liabilities", "CurrentLiabilities", "Equity", "Revenue", "OperatingIncome", "NetIncome"], [
            [row.fiscal_year_end.isoformat(), row.total_assets, row.current_assets, row.total_liabilities, row.current_liabilities, row.total_equity, row.revenue, row.operating_income, row.net_income]
            for row in gp_financials
        ]),
        ("2.3.운용사 펀드현황", ["FundName", "FundStatusDesc", "FundTypeDesc", "FormationYMD", "MaturityYMD", "CommitmentTotal", "PaidInTotal"], [
            [fund["name"], fund["status"], fund["type"], _json_default(fund["formation_date"]), _json_default(fund["maturity_date"]), fund["commitment_total"], fund["paid_in_total"]]
            for fund in funds
        ]),
        ("2.4.운용사 인력현황", ["FundName", "InvtComName", "InvestYMD", "InvestmentType"], [
            [row.fund_name, row.company_name, _json_default(row.investment_date), row.instrument]
            for row in manager_investments if row.is_current_company
        ]),
        ("3.1.운용인력 개요", ["FundMgrID", "FundMgrName", "BirthYMDS", "NationalityDesc", "BuName", "Position", "JoinYMD", "IsCoreManager"], [
            [row.id, row.name, _json_default(row.birth_date), row.nationality, row.department, row.position, _json_default(row.join_date), row.is_core]
            for row in managers
        ]),
        ("3.2.1.핵심운용인력 근무현황", ["CompanyName", "CompanyType", "DepartmentName", "PositionName", "WorkBegYMD", "WorkEndYMD", "MainTask", "IsInvestmentExp"], [
            [row.company_name, row.company_type, row.department, row.position, _json_default(row.start_date), _json_default(row.end_date), row.main_task, row.is_investment_exp]
            for row in careers if row.fund_manager_id in core_manager_ids
        ]),
        ("3.2.2.일반운용인력 근무현황", ["CompanyName", "CompanyType", "DepartmentName", "PositionName", "WorkBegYMD", "WorkEndYMD", "MainTask", "IsInvestmentExp"], [
            [row.company_name, row.company_type, row.department, row.position, _json_default(row.start_date), _json_default(row.end_date), row.main_task, row.is_investment_exp]
            for row in careers if row.fund_manager_id in general_manager_ids
        ]),
        ("3.3.운용인력 분류", ["GraduationYMD", "SchoolName", "MajorName", "DegreeType", "Country"], [
            [_json_default(row.graduation_date), row.school_name, row.major, row.degree, row.country]
            for row in educations
        ]),
        ("3.4.1.핵심운용인력 경력", ["CompanyName", "FundName", "InvtComName", "InvestYMD", "ContribRate"], [
            [row.source_company_name, row.fund_name, row.company_name, _json_default(row.investment_date), row.contrib_rate]
            for row in manager_investments if row.fund_manager_id in core_manager_ids
        ]),
        ("3.4.2.일반운용인력 경력", ["CompanyName", "FundName", "InvtComName", "InvestYMD", "ContribRate"], [
            [row.source_company_name, row.fund_name, row.company_name, _json_default(row.investment_date), row.contrib_rate]
            for row in manager_investments if row.fund_manager_id in general_manager_ids
        ]),
        ("4.1.펀드 실적", ["FundName", "TransactionType", "Amount", "TransactionDate"], [
            [next((fund["name"] for fund in funds if fund["id"] == row.fund_id), row.fund_id), row.type, row.amount, row.transaction_date.isoformat()]
            for row in transactions
        ]),
        ("4.2.펀드 인력별 이력", ["FundName", "ChangeCategory", "ManagerName", "ChangeDate", "Role"], [
            [next((fund["name"] for fund in funds if fund["id"] == row.fund_id), row.fund_id), row.change_type, next((manager.name for manager in managers if manager.id == row.fund_manager_id), row.fund_manager_id), row.change_date.isoformat(), row.role_after or row.role_before]
            for row in manager_histories
        ]),
    ]


def _legacy_motae5_sheets(workspace: dict[str, Any]) -> list[tuple[str, list[str], list[list[Any]]]]:
    funds = workspace.get("funds", [])
    shareholders: list[GPShareholder] = workspace.get("gp_shareholders", [])
    subscriptions: list[FundSubscription] = workspace.get("fund_subscriptions", [])
    latest_subscription_by_fund: dict[int, FundSubscription] = {}
    for row in subscriptions:
        if row.fund_id not in latest_subscription_by_fund:
            latest_subscription_by_fund[row.fund_id] = row

    return [
        ("청약이력 관련 총괄(전체)", ["조합명", "결성일", "청약일", "운용펀드수", "Co-GP여부"], [
            [fund["name"], _json_default(fund["formation_date"]), _json_default(latest_subscription_by_fund.get(fund["id"]).subscription_date) if fund["id"] in latest_subscription_by_fund else None, len(funds), False]
            for fund in funds
        ]),
        ("운용펀드 관련 총괄", ["조합명", "운용(투자)기간", "해산 예정일", "약정총액", "납입총액", "투자금액", "잔여자산"], [
            [fund["name"], _json_default(fund["investment_period_end"]), _json_default(fund["maturity_date"]), fund["commitment_total"], fund["paid_in_total"], fund["invested_total"], fund["nav_total"]]
            for fund in funds
        ]),
        ("대표 및 운용인력 개요", ["주주명", "주식수", "취득가액", "지분율", "비고"], [
            [row.name, row.shares, row.acquisition_amount, row.ownership_pct, row.memo]
            for row in shareholders
        ]),
    ]


def _legacy_motae6_sheets(workspace: dict[str, Any]) -> list[tuple[str, list[str], list[list[Any]]]]:
    managers: list[FundManager] = workspace.get("fund_managers", [])
    investments: list[ManagerInvestment] = workspace.get("manager_investments", [])
    core_ids = {row.id for row in managers if row.is_core or row.is_representative}
    exit_rows = [row for row in investments if row.exit_date or row.exit_amount]

    def _investment_row(row: ManagerInvestment) -> list[Any]:
        manager_name = next((manager.name for manager in managers if manager.id == row.fund_manager_id), row.fund_manager_id)
        return [manager_name, row.source_company_name, row.fund_name, None, row.is_current_company, row.company_name, None, row.instrument, row.discovery_contrib, row.review_contrib, _json_default(row.investment_date)]

    return [
        ("대표펀드매니저별 투자회수실적", ["대표펀드매니저", "근무회사명", "펀드명", "운용펀드여부", "투자기업명", "증권종류", "발굴기여율", "심사기여율", "투자실행일"], [_investment_row(row) for row in investments if row.fund_manager_id in core_ids]),
        ("운용인력별 투자회수실적", ["운용인력", "근무회사명", "펀드명", "운용펀드여부", "투자기업명", "증권종류", "발굴기여율", "심사기여율", "투자실행일"], [_investment_row(row) for row in investments if row.fund_manager_id not in core_ids]),
        ("최근 개인의 투자회수 현황", ["투자기업명", "증권종류", "투자실행일", "회수일", "투자금액", "회수금액", "잔여금액"], [[row.company_name, row.instrument, _json_default(row.investment_date), _json_default(row.exit_date), row.amount, row.exit_amount, (_to_float(row.amount) - _to_float(row.exit_amount))] for row in exit_rows]),
        ("최근 펀드의 투자회수 현황", ["펀드명", "투자기업명", "투자실행일", "회수일", "투자금액", "회수금액"], [[row.fund_name, row.company_name, _json_default(row.investment_date), _json_default(row.exit_date), row.amount, row.exit_amount] for row in exit_rows]),
    ]


def _legacy_motae7_sheets(workspace: dict[str, Any]) -> list[tuple[str, list[str], list[list[Any]]]]:
    managers: list[FundManager] = workspace.get("fund_managers", [])
    careers: list[ManagerCareer] = workspace.get("manager_careers", [])
    educations: list[ManagerEducation] = workspace.get("manager_educations", [])
    awards: list[ManagerAward] = workspace.get("manager_awards", [])
    investments: list[ManagerInvestment] = workspace.get("manager_investments", [])
    core_ids = {row.id for row in managers if row.is_core or row.is_representative}

    def _profile_rows(target_ids: set[int]) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for manager in managers:
            if manager.id not in target_ids:
                continue
            manager_education = next((row for row in educations if row.fund_manager_id == manager.id), None)
            manager_career = next((row for row in careers if row.fund_manager_id == manager.id), None)
            manager_award = next((row for row in awards if row.fund_manager_id == manager.id), None)
            rows.append([
                manager.name, manager.position, _json_default(manager.birth_date), manager.phone, manager.fax, manager.email,
                manager_award.award_name if manager_award else None,
                f"{_json_default(manager_education.admission_date)} ~ {_json_default(manager_education.graduation_date)}" if manager_education else None,
                manager_education.school_name if manager_education else None,
                manager_education.major if manager_education else None,
                manager_education.degree if manager_education else None,
                f"{_json_default(manager_career.start_date)} ~ {_json_default(manager_career.end_date)}" if manager_career else None,
                manager_career.company_name if manager_career else None,
                manager_career.company_type if manager_career else None,
                manager_career.department if manager_career else None,
                manager_career.position if manager_career else None,
                manager_career.main_task if manager_career else None,
                manager_career.employment_type if manager_career else None,
            ])
        return rows

    return [
        ("운용인력(대표자) 이력", ["성명", "직위", "생년월일", "전화", "팩스", "E-Mail", "수상내역", "학력기간", "학교명", "학과", "학위", "경력기간", "회사명", "회사유형", "부서명", "직위", "업무", "경력구분"], _profile_rows(core_ids)),
        ("운용인력(핵심) 이력", ["성명", "직위", "생년월일", "전화", "팩스", "E-Mail", "수상내역", "학력기간", "학교명", "학과", "학위", "경력기간", "회사명", "회사유형", "부서명", "직위", "업무", "경력구분"], _profile_rows({row.id for row in managers if row.id not in core_ids})),
        ("핵심운용인력 참여펀드 현황", ["이름", "펀드명", "펀드역할", "투자금액", "잔여금액"], [[next((manager.name for manager in managers if manager.id == row.fund_manager_id), row.fund_manager_id), row.fund_name, row.role, row.amount, max(_to_float(row.amount) - _to_float(row.exit_amount), 0)] for row in investments if row.fund_manager_id in core_ids]),
        ("핵심운용인력 이직내역", ["이름", "이전회사", "이전펀드명", "역할", "이직일"], [[next((manager.name for manager in managers if manager.id == row.fund_manager_id), row.fund_manager_id), row.source_company_name, row.fund_name, row.role, _json_default(row.exit_date)] for row in investments if row.fund_manager_id in core_ids and not row.is_current_company]),
    ]


def _legacy_nong_motae_sheets(workspace: dict[str, Any]) -> list[tuple[str, list[str], list[list[Any]]]]:
    gp = workspace.get("selected_gp_entity") or {}
    funds = workspace.get("funds", [])
    gp_financials: list[GPFinancial] = workspace.get("gp_financials", [])
    managers: list[FundManager] = workspace.get("fund_managers", [])
    investments: list[ManagerInvestment] = workspace.get("manager_investments", [])
    subscriptions: list[FundSubscription] = workspace.get("fund_subscriptions", [])
    shareholders: list[GPShareholder] = workspace.get("gp_shareholders", [])
    latest_financial = gp_financials[0] if gp_financials else None

    return [
        ("(별첨1)투자신청서", ["회사명", "설립일자", "주소", "운용인력수", "납입자본금", "자산총계", "자본총계", "담당자", "조합명(예정)", "약정금액", "대표펀드매니저"], [[gp.get("name"), _json_default(gp.get("founding_date")), gp.get("address"), gp.get("fund_manager_count"), latest_financial.paid_in_capital if latest_financial else gp.get("paid_in_capital"), latest_financial.total_assets if latest_financial else None, latest_financial.total_equity if latest_financial else None, gp.get("representative"), funds[0]["name"] if funds else None, funds[0]["commitment_total"] if funds else None, managers[0].name if managers else None]]),
        ("(별첨6-1)운용펀드 현황 총괄", ["운용사명", "펀드명", "투자기업명", "투자일", "투자유형"], [[gp.get("name"), row.fund_name, row.company_name, _json_default(row.investment_date), row.instrument] for row in investments if row.is_current_company]),
        ("(별첨6-2)펀드별 투자 현황", ["조합명", "조합유형", "대표펀드매니저", "결성일", "만기일", "약정총액", "납입총액", "투자금액", "잔여가치"], [[fund["name"], fund["type"], fund["fund_manager"], _json_default(fund["formation_date"]), _json_default(fund["maturity_date"]), fund["commitment_total"], fund["paid_in_total"], fund["invested_total"], fund["nav_total"]] for fund in funds]),
        ("(별첨6-3)청약이력투자 현황", ["조합명", "청약시기", "약정총액", "납입총액", "투자금액", "목표수익률"], [[next((fund["name"] for fund in funds if fund["id"] == row.fund_id), row.fund_id), row.subscription_date.isoformat(), next((fund["commitment_total"] for fund in funds if fund["id"] == row.fund_id), None), next((fund["paid_in_total"] for fund in funds if fund["id"] == row.fund_id), None), next((fund["invested_total"] for fund in funds if fund["id"] == row.fund_id), None), row.target_irr] for row in subscriptions]),
        ("(별첨7-1)대표자원 개요", ["성명", "직위", "소속", "입사일", "핵심여부"], [[row.name, row.position, gp.get("name"), _json_default(row.join_date), row.is_core] for row in managers]),
        ("(별첨7-2)운용인력 주요회사경력", ["운용인력", "근무회사명", "펀드명", "투자기업명", "증권종류", "투자실행일", "회수일", "투자금액", "회수금액"], [[next((manager.name for manager in managers if manager.id == row.fund_manager_id), row.fund_manager_id), row.source_company_name, row.fund_name, row.company_name, row.instrument, _json_default(row.investment_date), _json_default(row.exit_date), row.amount, row.exit_amount] for row in investments]),
        ("(별첨7-3)운용인력 청약참여내역", ["운용인력", "근무회사명", "조합명", "약정총액", "투자금액", "매니저구분"], [[next((manager.name for manager in managers if manager.id == row.fund_manager_id), row.fund_manager_id), row.source_company_name, row.fund_name, next((fund["commitment_total"] for fund in funds if fund["name"] == row.fund_name), None), row.amount, row.role] for row in investments]),
        ("(별첨8)운용사 주요매체 및 주요 출자자명단", ["주주명", "주식수", "취득가액", "지분율", "비고"], [[row.name, row.shares, row.acquisition_amount, row.ownership_pct, row.memo] for row in shareholders]),
    ]


def build_proposal_workbook(workspace: dict[str, Any], db: Session) -> tuple[bytes, str]:
    template_type = workspace["template_type"]
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    if template_type == "growth-finance":
        sheet_defs = _growth_finance_sheets(workspace, db)
        filename = f"growth_finance_{workspace['as_of_date'].isoformat()}.xlsx"
    elif template_type == "motae-5":
        sheet_defs = _motae5_sheets(workspace, db)
        filename = f"motae5_{workspace['as_of_date'].isoformat()}.xlsx"
    elif template_type == "motae-6":
        sheet_defs = _motae6_sheets(workspace, db)
        filename = f"motae6_{workspace['as_of_date'].isoformat()}.xlsx"
    elif template_type == "motae-7":
        sheet_defs = _motae7_sheets(workspace, db)
        filename = f"motae7_{workspace['as_of_date'].isoformat()}.xlsx"
    else:
        sheet_defs = _nong_motae_sheets(workspace, db)
        filename = f"nong_motae_{workspace['as_of_date'].isoformat()}.xlsx"

    for title, headers, rows in sheet_defs:
        ws = wb.create_sheet(title[:31])
        _write_sheet(ws, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), filename


_TEMPLATE_LABELS = {
    "growth-finance": "성장금융",
    "motae-5": "모태 별첨5",
    "motae-6": "모태 별첨6",
    "motae-7": "모태 별첨7",
    "nong-motae": "농모태",
}


_WORKBENCH_SHEET_CATALOG: dict[str, list[dict[str, Any]]] = {
    "growth-finance": [],
    "motae-5": [],
    "motae-6": [],
    "motae-7": [],
    "nong-motae": [],
}

_WORKBENCH_SHEET_CATALOG["growth-finance"] = [
    {"code": "guide", "title": "작성요령", "kind": "info", "description": "작성 기준일과 안내 문구를 확인하는 시트입니다.", "columns": [("item", "항목"), ("value", "값")]},
    {"code": "proposal-fund", "title": "1.1.제안펀드", "kind": "table", "description": "제안 대상 조합의 기본 정보를 복사하거나 내려받습니다.", "columns": [("fund_name", "펀드명"), ("formation_date", "결성일"), ("maturity_date", "만기일"), ("commitment_total", "약정총액"), ("paid_in_total", "납입총액"), ("fund_type", "투자기구"), ("fund_status", "펀드상태")]},
    {"code": "investment-history", "title": "1.2.출자예상내역", "kind": "table", "description": "출자자 및 약정 정보를 정리한 시트입니다.", "columns": [("lp_type", "출자자유형"), ("lp_name", "출자자명"), ("commitment", "출자약정예상액")]},
    {"code": "gp-overview", "title": "2.1.제안사 개요", "kind": "scalar", "description": "GP 기본 정보를 초안용 최종값으로 정리합니다.", "columns": [("gp_name", "제안사명"), ("corp_number", "법인등록번호"), ("business_number", "사업자등록번호"), ("ceo", "대표자명"), ("founded_date", "설립일"), ("address", "주소"), ("total_employees", "전체 인원수"), ("fund_manager_count", "전문인력수"), ("paid_in_capital", "납입자본금")]},
    {"code": "gp-financials", "title": "2.2.제안사 재무현황", "kind": "table", "description": "기준일 이하 최신 재무 데이터를 확인합니다.", "columns": [("fiscal_year_end", "(가)결산일자"), ("total_assets", "총자산"), ("current_assets", "유동자산"), ("total_liabilities", "총부채"), ("current_liabilities", "유동부채"), ("total_equity", "자기자본(자본총계)"), ("revenue", "영업수익"), ("operating_income", "영업이익"), ("net_income", "당기순이익")]},
    {"code": "gp-funds", "title": "2.3.제안사 펀드현황", "kind": "table", "description": "GP가 운용한 조합 현황을 정리합니다.", "columns": [("fund_name", "펀드명"), ("fund_status", "펀드 상태"), ("fund_type", "펀드법적유형"), ("formation_date", "결성일"), ("maturity_date", "만기일"), ("commitment_total", "약정총액"), ("paid_in_total", "납입총액")]},
    {"code": "gp-investments", "title": "2.4.제안사 투자현황", "kind": "table", "description": "현 회사 기준 투자 이력을 정리합니다.", "columns": [("fund_name", "펀드명"), ("company_name", "피투자기업명"), ("investment_date", "투자일"), ("instrument", "투자형태")]},
    {"code": "manager-overview", "title": "3.1.참여인력 개요", "kind": "table", "description": "제안서에 포함될 운용인력 기본 정보를 검토합니다.", "columns": [("manager_id", "운용인력ID"), ("manager_name", "성명"), ("birth_date", "생년월일"), ("nationality", "국적"), ("department", "부서"), ("position", "직위"), ("join_date", "입사일"), ("is_core", "핵심운용인력 여부")]},
    {"code": "manager-careers-core", "title": "3.4.1.핵심운용인력 경력", "kind": "table", "description": "핵심운용인력 근무 이력을 정리합니다.", "columns": [("company_name", "회사명"), ("company_type", "회사유형"), ("department", "부서명"), ("position", "직위"), ("start_date", "시작일"), ("end_date", "종료일"), ("main_task", "담당업무"), ("is_investment_exp", "투자관련 경력여부")]},
    {"code": "manager-careers-general", "title": "3.4.2.일반운용인력 경력", "kind": "table", "description": "일반운용인력 근무 이력을 정리합니다.", "columns": [("company_name", "회사명"), ("company_type", "회사유형"), ("department", "부서명"), ("position", "직위"), ("start_date", "시작일"), ("end_date", "종료일"), ("main_task", "담당업무"), ("is_investment_exp", "투자관련 경력여부")]},
    {"code": "manager-educations", "title": "3.3.참여인력 학력", "kind": "table", "description": "운용인력 학력 이력을 정리합니다.", "columns": [("graduation_date", "졸업일/수료일"), ("school_name", "졸업/수료 학교명"), ("major", "학과(전공)명"), ("degree", "학위명"), ("country", "취득국명")]},
    {"code": "manager-investments-core", "title": "3.2.1.핵심운용인력 투자현황", "kind": "table", "description": "핵심운용인력 투자 경력을 정리합니다.", "columns": [("company_name", "재직회사명"), ("fund_name", "펀드명"), ("portfolio_name", "투자기업명"), ("investment_date", "투자일"), ("contrib_rate", "기여율")]},
    {"code": "manager-investments-general", "title": "3.2.2.일반운용인력 투자현황", "kind": "table", "description": "일반운용인력 투자 경력을 정리합니다.", "columns": [("company_name", "재직회사명"), ("fund_name", "펀드명"), ("portfolio_name", "투자기업명"), ("investment_date", "투자일"), ("contrib_rate", "기여율")]},
    {"code": "fund-performance", "title": "4.1.펀드 거래", "kind": "table", "description": "거래 내역을 기준일 기준으로 정리합니다.", "columns": [("fund_name", "펀드명"), ("transaction_type", "거래구분"), ("amount", "거래금액 (원)"), ("transaction_date", "거래일")]},
    {"code": "fund-manager-changes", "title": "4.2.펀드 인력변경 이력", "kind": "table", "description": "펀드별 인력 선임 및 퇴임 이력을 관리합니다.", "columns": [("fund_name", "펀드명"), ("change_type", "기재대상핵심인력 변경구분"), ("manager_name", "이름"), ("change_date", "변경일자"), ("role", "역할")]},
]

_WORKBENCH_SHEET_CATALOG["motae-5"] = [
    {"code": "subscription-summary", "title": "청약이력 관련 총괄(전체)", "kind": "table", "description": "과거 청약이력과 조합 기본 정보를 정리합니다.", "columns": [("fund_name", "조합명"), ("formation_date", "결성일"), ("subscription_date", "청약일"), ("fund_count", "운용펀드 수"), ("has_co_gp", "공동 운용사(Co-GP) 여부")]},
    {"code": "fund-summary", "title": "2. 운용중인 조합 총괄", "kind": "table", "description": "운용 조합의 잔액 및 실적을 정리합니다.", "columns": [("fund_name", "조합명"), ("investment_period_end", "운용(투자)기간"), ("maturity_date", "해산 예정일"), ("commitment_total", "약정총액"), ("paid_in_total", "납입총액"), ("invested_total", "투자금액"), ("nav_total", "잔여자산")]},
    {"code": "shareholder-summary", "title": "3. 주주 및 출자자 명부", "kind": "table", "description": "주요 주주 정보를 정리합니다.", "columns": [("name", "주주명"), ("shares", "주식수"), ("acquisition_amount", "납입액"), ("ownership_pct", "지분율"), ("memo", "비고")]},
]

_WORKBENCH_SHEET_CATALOG["motae-6"] = [
    {"code": "representative-recovery", "title": "1. 대표펀드매니저 투자회수실적", "kind": "table", "description": "대표펀드매니저 기준 투자회수실적을 정리합니다.", "columns": [("manager_name", "대표펀드매니저"), ("company_name", "재직회사명"), ("fund_name", "펀드명"), ("fund_type", "펀드 법적형태"), ("is_current_company", "현재 회사 여부"), ("portfolio_name", "투자기업명"), ("instrument", "투자형태"), ("discovery_contrib", "발굴 기여율(%)"), ("review_contrib", "심사 기여율(%)"), ("investment_date", "투자일")]},
    {"code": "manager-recovery", "title": "2. 참여인력 투자회수실적", "kind": "table", "description": "핵심운용인력 기준 투자회수실적을 정리합니다.", "columns": [("manager_name", "참여인력"), ("company_name", "재직회사명"), ("fund_name", "펀드명"), ("fund_type", "펀드 법적형태"), ("is_current_company", "현재 회사 여부"), ("portfolio_name", "투자기업명"), ("instrument", "투자형태"), ("discovery_contrib", "발굴 기여율(%)"), ("review_contrib", "심사 기여율(%)"), ("investment_date", "투자일")]},
    {"code": "recent-personal-recovery", "title": "3. 최근 개인 투자회수 현황", "kind": "table", "description": "개인 기준 최근 투자회수 내역을 정리합니다.", "columns": [("portfolio_name", "투자기업명"), ("instrument", "투자형태"), ("investment_date", "투자일"), ("exit_date", "회수일"), ("amount", "투자금액"), ("exit_amount", "회수금액"), ("remaining_amount", "잔여금액")]},
    {"code": "recent-fund-recovery", "title": "4. 최근 펀드 투자회수 현황", "kind": "table", "description": "펀드 기준 최근 투자회수 내역을 정리합니다.", "columns": [("fund_name", "펀드명"), ("portfolio_name", "투자기업명"), ("investment_date", "투자일"), ("exit_date", "회수일"), ("amount", "투자금액"), ("exit_amount", "회수금액")]},
]

_WORKBENCH_SHEET_CATALOG["motae-7"] = [
    {"code": "representative-profile", "title": "2. 참여인력(대표) 이력", "kind": "table", "description": "대표자 이력을 정리합니다.", "columns": [("name", "성명"), ("position", "직위"), ("birth_date", "생년월일"), ("phone", "전화"), ("fax", "팩스"), ("email", "E-Mail"), ("award_name", "수상내역"), ("education_period", "학력기간"), ("school_name", "학교명"), ("major", "학과(전공)명"), ("degree", "학위명"), ("career_period", "경력기간"), ("company_name", "직장명(운용사 구분)"), ("company_type", "회사유형"), ("department", "부서명"), ("career_position", "직위"), ("main_task", "담당업무"), ("employment_type", "경력구분")]},
    {"code": "core-profile", "title": "2. 참여인력(핵심) 이력", "kind": "table", "description": "핵심운용인력 이력을 정리합니다.", "columns": [("name", "성명"), ("position", "직위"), ("birth_date", "생년월일"), ("phone", "전화"), ("fax", "팩스"), ("email", "E-Mail"), ("award_name", "수상내역"), ("education_period", "학력기간"), ("school_name", "학교명"), ("major", "학과(전공)명"), ("degree", "학위명"), ("career_period", "경력기간"), ("company_name", "직장명(운용사 구분)"), ("company_type", "회사유형"), ("department", "부서명"), ("career_position", "직위"), ("main_task", "담당업무"), ("employment_type", "경력구분")]},
    {"code": "core-fund-participation", "title": "3. 핵심운용인력 참여펀드 현황", "kind": "table", "description": "핵심운용인력의 참여펀드 현황을 정리합니다.", "columns": [("manager_name", "성명"), ("fund_name", "참여펀드명"), ("role", "참여펀드 역할"), ("amount", "투자금액"), ("remaining_amount", "잔여금액")]},
    {"code": "core-job-change", "title": "4. 핵심운용인력 이직내역", "kind": "table", "description": "핵심운용인력의 과거 소속 및 이직 이력을 정리합니다.", "columns": [("manager_name", "성명"), ("previous_company", "이전 회사"), ("fund_name", "이전 펀드명"), ("role", "역할"), ("change_date", "이직일")]},
]

_WORKBENCH_SHEET_CATALOG["nong-motae"] = [
    {"code": "application-form", "title": "(별첨1) 투자신청서", "kind": "scalar", "description": "농모태 투자신청서 기본 항목을 정리합니다.", "columns": [("gp_name", "회사명"), ("founded_date", "설립연월일"), ("address", "주소"), ("fund_manager_count", "운용인력수"), ("paid_in_capital", "납입자본금"), ("total_assets", "자산총계"), ("total_equity", "자본총계"), ("representative", "담당자"), ("fund_name", "조합명(예정)"), ("commitment_total", "약정금액"), ("manager_name", "대표펀드매니저")]},
    {"code": "fund-overview", "title": "(별첨6-1) 운용펀드 현황 총괄", "kind": "table", "description": "운용중 펀드의 투자 현황을 정리합니다.", "columns": [("gp_name", "운용사명"), ("fund_name", "펀드명"), ("portfolio_name", "투자기업명"), ("investment_date", "투자일"), ("instrument", "투자방식")]},
    {"code": "fund-investments", "title": "(별첨6-2) 펀드별 투자 현황", "kind": "table", "description": "펀드별 약정 및 투자 집행 현황을 정리합니다.", "columns": [("fund_name", "조합명"), ("fund_type", "조합 구분"), ("fund_manager", "대표/펀드매니저"), ("formation_date", "등록(성립)일"), ("maturity_date", "청산시기(예정)"), ("commitment_total", "약정총액"), ("paid_in_total", "납입총액"), ("invested_total", "투자총액"), ("nav_total", "조합가치")]},
    {"code": "subscription-history", "title": "(별첨6-3) 청약이력 투자 현황", "kind": "table", "description": "청약이력 및 목표수익률을 정리합니다.", "columns": [("fund_name", "조합명"), ("subscription_date", "청약시기"), ("commitment_total", "약정총액"), ("paid_in_total", "납입총액"), ("invested_total", "투자총액"), ("target_irr", "목표수익률")]},
    {"code": "representative-overview", "title": "(별첨7-1) 인적자원 관련", "kind": "table", "description": "대표 및 핵심 인력 개요를 정리합니다.", "columns": [("name", "성명"), ("position", "직위"), ("company_name", "소속"), ("join_date", "입사일"), ("is_core", "핵심여부")]},
    {"code": "manager-major-careers", "title": "(별첨7-2) 참여인력 투자회수실적", "kind": "table", "description": "운용인력 주요 회사 경력과 투자회수 실적을 정리합니다.", "columns": [("manager_name", "참여인력"), ("company_name", "재직회사명"), ("fund_name", "펀드명"), ("portfolio_name", "투자기업명"), ("instrument", "투자형태"), ("investment_date", "투자시작일"), ("exit_date", "회수종료일"), ("amount", "투자금액"), ("exit_amount", "회수금액")]},
    {"code": "manager-subscription-history", "title": "(별첨7-3) 참여인력 조합청산실적", "kind": "table", "description": "운용인력의 조합 참여이력을 정리합니다.", "columns": [("manager_name", "참여인력"), ("company_name", "재직회사명"), ("fund_name", "조합명"), ("commitment_total", "약정총액"), ("amount", "투자금액"), ("role", "매니저구분")]},
    {"code": "key-shareholders", "title": "(별첨8) 제안사 주주명부 및 조합 출자자명부", "kind": "table", "description": "주요 주주 정보를 정리합니다.", "columns": [("name", "주주명"), ("shares", "주식수"), ("acquisition_amount", "납입액"), ("ownership_pct", "지분율"), ("memo", "비고")]},
]


def _first_row_by_manager(rows: list[Any]) -> dict[int, Any]:
    result: dict[int, Any] = {}
    for row in rows:
        manager_id = getattr(row, "fund_manager_id", None)
        if manager_id is None or manager_id in result:
            continue
        result[manager_id] = row
    return result


def _format_period(start_value: Any, end_value: Any) -> str | None:
    start_text = _stringify_cell(start_value)
    end_text = _stringify_cell(end_value)
    if not start_text and not end_text:
        return None
    if start_text and end_text:
        return f"{start_text} ~ {end_text}"
    return start_text or end_text


def _select_representative_manager(managers: list[FundManager]) -> FundManager | None:
    for manager in managers:
        if manager.is_representative:
            return manager
    for manager in managers:
        if manager.is_core:
            return manager
    return None


def _proposal_views(workspace: dict[str, Any], db: Session) -> dict[str, Any]:
    cached = workspace.get("_proposal_views")
    if cached is not None:
        return cached

    as_of_date: date = workspace["as_of_date"]
    selected_funds: list[dict[str, Any]] = list(workspace.get("funds", []))
    gp_financials: list[GPFinancial] = workspace.get("gp_financials", [])
    shareholders: list[GPShareholder] = workspace.get("gp_shareholders", [])
    managers: list[FundManager] = workspace.get("fund_managers", [])
    careers: list[ManagerCareer] = workspace.get("manager_careers", [])
    educations: list[ManagerEducation] = workspace.get("manager_educations", [])
    awards: list[ManagerAward] = workspace.get("manager_awards", [])
    investments: list[ManagerInvestment] = workspace.get("manager_investments", [])
    manager_histories: list[FundManagerHistory] = workspace.get("fund_manager_histories", [])
    subscriptions: list[FundSubscription] = workspace.get("fund_subscriptions", [])

    selected_funds_by_id = {
        int(fund["id"]): fund
        for fund in selected_funds
        if fund.get("id") is not None
    }
    related_fund_ids = {
        row.fund_id
        for row in investments
        if row.fund_id is not None
    }
    related_fund_ids.update(row.fund_id for row in manager_histories if row.fund_id is not None)
    related_fund_ids.update(row.fund_id for row in subscriptions if row.fund_id is not None)

    all_funds_by_id = dict(selected_funds_by_id)
    missing_fund_ids = [fund_id for fund_id in related_fund_ids if fund_id not in all_funds_by_id]
    if missing_fund_ids:
        extra_funds = (
            db.query(Fund)
            .filter(Fund.id.in_(missing_fund_ids))
            .order_by(Fund.id.asc())
            .all()
        )
        for fund in extra_funds:
            all_funds_by_id[fund.id] = _build_fund_snapshot(db, fund, as_of_date)

    managers_by_id = {manager.id: manager for manager in managers}
    representative_ids = {manager.id for manager in managers if manager.is_representative}
    core_ids = {manager.id for manager in managers if manager.is_core}
    key_manager_ids = {manager.id for manager in managers if manager.is_core or manager.is_representative}

    latest_education_by_manager = _first_row_by_manager(educations)
    latest_career_by_manager = _first_row_by_manager(careers)
    latest_award_by_manager = _first_row_by_manager(awards)

    proposal_managers: list[dict[str, Any]] = []
    for manager in managers:
        proposal_managers.append(
            {
                "id": manager.id,
                "name": manager.name,
                "birth_date": manager.birth_date,
                "nationality": manager.nationality,
                "phone": manager.phone,
                "fax": manager.fax,
                "email": manager.email,
                "department": manager.department,
                "position": manager.position,
                "join_date": manager.join_date,
                "is_core": bool(manager.is_core),
                "is_representative": bool(manager.is_representative),
                "is_key_person": bool(manager.is_core or manager.is_representative),
            }
        )

    proposal_manager_careers: list[dict[str, Any]] = []
    for row in careers:
        manager = managers_by_id.get(row.fund_manager_id)
        proposal_manager_careers.append(
            {
                "fund_manager_id": row.fund_manager_id,
                "manager_name": manager.name if manager else row.fund_manager_id,
                "company_name": row.company_name,
                "company_type": row.company_type,
                "department": row.department,
                "position": row.position,
                "start_date": row.start_date,
                "end_date": row.end_date,
                "main_task": row.main_task,
                "is_investment_exp": row.is_investment_exp,
                "employment_type": row.employment_type,
                "is_core": bool(manager.is_core) if manager else False,
                "is_representative": bool(manager.is_representative) if manager else False,
                "is_key_person": bool(manager and (manager.is_core or manager.is_representative)),
            }
        )

    proposal_manager_educations: list[dict[str, Any]] = []
    for row in educations:
        manager = managers_by_id.get(row.fund_manager_id)
        proposal_manager_educations.append(
            {
                "fund_manager_id": row.fund_manager_id,
                "manager_name": manager.name if manager else row.fund_manager_id,
                "school_name": row.school_name,
                "major": row.major,
                "degree": row.degree,
                "admission_date": row.admission_date,
                "graduation_date": row.graduation_date,
                "country": row.country,
            }
        )

    proposal_manager_investments: list[dict[str, Any]] = []
    for row in investments:
        manager = managers_by_id.get(row.fund_manager_id)
        fund = all_funds_by_id.get(row.fund_id) if row.fund_id is not None else None
        proposal_manager_investments.append(
            {
                "id": row.id,
                "fund_manager_id": row.fund_manager_id,
                "manager_name": manager.name if manager else row.fund_manager_id,
                "is_core": bool(manager.is_core) if manager else False,
                "is_representative": bool(manager.is_representative) if manager else False,
                "is_key_person": bool(manager and (manager.is_core or manager.is_representative)),
                "fund_id": row.fund_id,
                "fund_name": row.fund_name or (fund.get("name") if fund else None),
                "fund_type": fund.get("type") if fund else None,
                "fund_status": fund.get("status") if fund else None,
                "commitment_total": fund.get("commitment_total") if fund else None,
                "source_company_name": row.source_company_name,
                "company_name": row.company_name,
                "investment_date": row.investment_date,
                "instrument": row.instrument,
                "amount": row.amount,
                "exit_date": row.exit_date,
                "exit_amount": row.exit_amount,
                "remaining_amount": max(_to_float(row.amount) - _to_float(row.exit_amount), 0),
                "role": row.role,
                "discovery_contrib": row.discovery_contrib,
                "review_contrib": row.review_contrib,
                "contrib_rate": row.contrib_rate,
                "is_current_company": bool(row.is_current_company),
                "has_exit": bool(row.exit_date or row.exit_amount),
            }
        )

    proposal_manager_changes: list[dict[str, Any]] = []
    for row in manager_histories:
        manager = managers_by_id.get(row.fund_manager_id)
        fund = all_funds_by_id.get(row.fund_id)
        proposal_manager_changes.append(
            {
                "fund_id": row.fund_id,
                "fund_name": fund.get("name") if fund else row.fund_id,
                "fund_manager_id": row.fund_manager_id,
                "manager_name": manager.name if manager else row.fund_manager_id,
                "change_date": row.change_date,
                "change_type": row.change_type,
                "role": row.role_after or row.role_before,
            }
        )

    latest_shareholder_snapshot_date = shareholders[0].snapshot_date if shareholders else None
    proposal_shareholders = [
        row
        for row in shareholders
        if latest_shareholder_snapshot_date is not None and row.snapshot_date == latest_shareholder_snapshot_date
    ]

    proposal_subscriptions: list[dict[str, Any]] = []
    latest_subscription_by_fund: dict[int, dict[str, Any]] = {}
    for row in subscriptions:
        fund = all_funds_by_id.get(row.fund_id)
        subscription_row = {
            "fund_id": row.fund_id,
            "fund_name": fund.get("name") if fund else row.fund_id,
            "subscription_date": row.subscription_date,
            "target_irr": row.target_irr,
            "target_commitment": row.target_commitment,
            "actual_commitment": row.actual_commitment,
            "commitment_total": fund.get("commitment_total") if fund else None,
            "paid_in_total": fund.get("paid_in_total") if fund else None,
            "invested_total": fund.get("invested_total") if fund else None,
        }
        proposal_subscriptions.append(subscription_row)
        latest_subscription_by_fund.setdefault(row.fund_id, subscription_row)

    proposal_manager_profiles: list[dict[str, Any]] = []
    for manager in managers:
        education = latest_education_by_manager.get(manager.id)
        career = latest_career_by_manager.get(manager.id)
        award = latest_award_by_manager.get(manager.id)
        proposal_manager_profiles.append(
            {
                "manager_id": manager.id,
                "name": manager.name,
                "position": manager.position,
                "birth_date": manager.birth_date,
                "phone": manager.phone,
                "fax": manager.fax,
                "email": manager.email,
                "award_name": award.award_name if award else None,
                "education_period": _format_period(
                    education.admission_date if education else None,
                    education.graduation_date if education else None,
                ),
                "school_name": education.school_name if education else None,
                "major": education.major if education else None,
                "degree": education.degree if education else None,
                "career_period": _format_period(
                    career.start_date if career else None,
                    career.end_date if career else None,
                ),
                "company_name": career.company_name if career else None,
                "company_type": career.company_type if career else None,
                "department": career.department if career else None,
                "career_position": career.position if career else None,
                "main_task": career.main_task if career else None,
                "employment_type": career.employment_type if career else None,
                "is_core": bool(manager.is_core),
                "is_representative": bool(manager.is_representative),
            }
        )

    representative_manager = _select_representative_manager(managers)
    representative_manager_view = None
    if representative_manager is not None:
        representative_manager_view = next(
            (row for row in proposal_managers if row["id"] == representative_manager.id),
            None,
        )

    views = {
        "proposal_funds": selected_funds,
        "all_funds_by_id": all_funds_by_id,
        "proposal_managers": proposal_managers,
        "proposal_manager_careers": proposal_manager_careers,
        "proposal_manager_educations": proposal_manager_educations,
        "proposal_manager_investments": proposal_manager_investments,
        "proposal_manager_changes": proposal_manager_changes,
        "proposal_manager_profiles": proposal_manager_profiles,
        "proposal_subscriptions": proposal_subscriptions,
        "latest_subscription_by_fund": latest_subscription_by_fund,
        "proposal_shareholders": proposal_shareholders,
        "representative_manager_ids": representative_ids,
        "core_manager_ids": core_ids,
        "key_manager_ids": key_manager_ids,
        "single_selected_fund": selected_funds[0] if len(selected_funds) == 1 else None,
        "representative_manager": representative_manager_view,
        "latest_financial": gp_financials[0] if gp_financials else None,
    }
    workspace["_proposal_views"] = views
    return views


def _growth_finance_sheets(workspace: dict[str, Any], db: Session) -> list[tuple[str, list[str], list[list[Any]]]]:
    gp = workspace.get("selected_gp_entity") or {}
    gp_financials: list[GPFinancial] = workspace.get("gp_financials", [])
    as_of_date: date = workspace["as_of_date"]
    views = _proposal_views(workspace, db)
    funds = views["proposal_funds"]
    managers = views["proposal_managers"]
    careers = views["proposal_manager_careers"]
    educations = views["proposal_manager_educations"]
    investments = views["proposal_manager_investments"]
    manager_changes = views["proposal_manager_changes"]

    fund_ids = [fund["id"] for fund in funds]
    lps = (
        db.query(LP)
        .filter(LP.fund_id.in_(fund_ids))
        .order_by(LP.fund_id.asc(), LP.id.asc())
        .all()
        if fund_ids else []
    )
    transactions = (
        db.query(Transaction)
        .filter(Transaction.fund_id.in_(fund_ids), Transaction.transaction_date <= as_of_date)
        .order_by(Transaction.transaction_date.desc(), Transaction.id.desc())
        .all()
        if fund_ids else []
    )
    growth_core_manager_ids = {row["id"] for row in managers if row["is_core"]}
    growth_general_manager_ids = {row["id"] for row in managers if row["id"] not in growth_core_manager_ids}
    selected_funds_by_id = {fund["id"]: fund for fund in funds}

    return [
        ("작성요령", ["항목", "값"], [["기준일", as_of_date.isoformat()], ["안내", "기준일 기준으로 ERP 데이터에서 자동 구성한 제안서 데이터입니다."]]),
        ("1.1.제안펀드", ["펀드명", "결성일", "만기일", "약정총액", "납입총액", "투자기구", "펀드상태"], [
            [fund["name"], _json_default(fund.get("formation_date")), _json_default(fund.get("maturity_date")), fund.get("commitment_total"), fund.get("paid_in_total"), fund.get("type"), fund.get("status")]
            for fund in funds
        ]),
        ("1.2.출자예상내역", ["출자자유형", "출자자명", "출자약정예상액"], [
            [normalize_lp_type(lp.type) or lp.type, lp.name, lp.commitment] for lp in lps
        ]),
        ("2.1.제안사 개요", ["제안사명", "법인등록번호", "사업자등록번호", "대표자명", "설립일", "주소", "전체 인원수", "전문인력수", "납입자본금"], [[
            gp.get("name"),
            gp.get("registration_number"),
            gp.get("business_number"),
            gp.get("representative"),
            _json_default(gp.get("founding_date")),
            gp.get("address"),
            gp.get("total_employees"),
            gp.get("fund_manager_count"),
            gp.get("paid_in_capital"),
        ]]),
        ("2.2.제안사 재무현황", ["(가)결산일자", "총자산", "유동자산", "총부채", "유동부채", "자기자본(자본총계)", "영업수익", "영업이익", "당기순이익"], [
            [_json_default(row.fiscal_year_end), row.total_assets, row.current_assets, row.total_liabilities, row.current_liabilities, row.total_equity, row.revenue, row.operating_income, row.net_income]
            for row in gp_financials
        ]),
        ("2.3.제안사 펀드현황", ["펀드명", "펀드 상태", "펀드법적유형", "결성일", "만기일", "약정총액", "납입총액"], [
            [fund["name"], fund.get("status"), fund.get("type"), _json_default(fund.get("formation_date")), _json_default(fund.get("maturity_date")), fund.get("commitment_total"), fund.get("paid_in_total")]
            for fund in funds
        ]),
        ("2.4.제안사 투자현황", ["펀드명", "피투자기업명", "투자일", "투자형태"], [
            [row["fund_name"], row["company_name"], _json_default(row["investment_date"]), row["instrument"]]
            for row in investments if row["is_current_company"]
        ]),
        ("3.1.참여인력 개요", ["운용인력ID", "성명", "생년월일", "국적", "부서", "직위", "입사일", "핵심운용인력 여부"], [
            [row["id"], row["name"], _json_default(row["birth_date"]), row["nationality"], row["department"], row["position"], _json_default(row["join_date"]), row["is_core"]]
            for row in managers
        ]),
        ("3.4.1.핵심운용인력 경력", ["회사명", "회사유형", "부서명", "직위", "시작일", "종료일", "담당업무", "투자관련 경력여부"], [
            [row["company_name"], row["company_type"], row["department"], row["position"], _json_default(row["start_date"]), _json_default(row["end_date"]), row["main_task"], row["is_investment_exp"]]
            for row in careers if row["fund_manager_id"] in growth_core_manager_ids
        ]),
        ("3.4.2.일반운용인력 경력", ["회사명", "회사유형", "부서명", "직위", "시작일", "종료일", "담당업무", "투자관련 경력여부"], [
            [row["company_name"], row["company_type"], row["department"], row["position"], _json_default(row["start_date"]), _json_default(row["end_date"]), row["main_task"], row["is_investment_exp"]]
            for row in careers if row["fund_manager_id"] in growth_general_manager_ids
        ]),
        ("3.3.참여인력 학력", ["졸업일/수료일", "졸업/수료 학교명", "학과(전공)명", "학위명", "취득국명"], [
            [_json_default(row["graduation_date"]), row["school_name"], row["major"], row["degree"], row["country"]]
            for row in educations
        ]),
        ("3.2.1.핵심운용인력 투자현황", ["재직회사명", "펀드명", "투자기업명", "투자일", "기여율"], [
            [row["source_company_name"], row["fund_name"], row["company_name"], _json_default(row["investment_date"]), row["contrib_rate"]]
            for row in investments if row["fund_manager_id"] in growth_core_manager_ids
        ]),
        ("3.2.2.일반운용인력 투자현황", ["재직회사명", "펀드명", "투자기업명", "투자일", "기여율"], [
            [row["source_company_name"], row["fund_name"], row["company_name"], _json_default(row["investment_date"]), row["contrib_rate"]]
            for row in investments if row["fund_manager_id"] in growth_general_manager_ids
        ]),
        ("4.1.펀드 거래", ["펀드명", "거래구분", "거래금액 (원)", "거래일"], [
            [selected_funds_by_id.get(row.fund_id, {}).get("name", row.fund_id), row.type, row.amount, _json_default(row.transaction_date)]
            for row in transactions
        ]),
        ("4.2.펀드 인력변경 이력", ["펀드명", "기재대상핵심인력 변경구분", "이름", "변경일자", "역할"], [
            [row["fund_name"], row["change_type"], row["manager_name"], _json_default(row["change_date"]), row["role"]]
            for row in manager_changes
        ]),
    ]


def _motae5_sheets(workspace: dict[str, Any], db: Session) -> list[tuple[str, list[str], list[list[Any]]]]:
    views = _proposal_views(workspace, db)
    funds = views["proposal_funds"]
    shareholders: list[GPShareholder] = views["proposal_shareholders"]
    latest_subscription_by_fund: dict[int, dict[str, Any]] = views["latest_subscription_by_fund"]

    return [
        ("청약이력 관련 총괄(전체)", ["조합명", "결성일", "청약일", "운용펀드 수", "공동 운용사(Co-GP) 여부"], [
            [fund["name"], _json_default(fund.get("formation_date")), _json_default(latest_subscription_by_fund.get(fund["id"], {}).get("subscription_date")), len(funds), bool(fund.get("has_co_gp"))]
            for fund in funds
        ]),
        ("2. 운용중인 조합 총괄", ["조합명", "운용(투자)기간", "해산 예정일", "약정총액", "납입총액", "투자금액", "잔여자산"], [
            [fund["name"], _json_default(fund.get("investment_period_end")), _json_default(fund.get("maturity_date")), fund.get("commitment_total"), fund.get("paid_in_total"), fund.get("invested_total"), fund.get("nav_total")]
            for fund in funds
        ]),
        ("3. 주주 및 출자자 명부", ["주주명", "주식수", "납입액", "지분율", "비고"], [
            [row.name, row.shares, row.acquisition_amount, row.ownership_pct, row.memo]
            for row in shareholders
        ]),
    ]


def _motae6_sheets(workspace: dict[str, Any], db: Session) -> list[tuple[str, list[str], list[list[Any]]]]:
    views = _proposal_views(workspace, db)
    investments = views["proposal_manager_investments"]
    representative_ids: set[int] = views["representative_manager_ids"]
    representative_rows = [row for row in investments if row["fund_manager_id"] in representative_ids]
    participant_rows = [row for row in investments if row["fund_manager_id"] not in representative_ids]
    exit_rows = [row for row in investments if row["has_exit"]]

    def _recovery_row(row: dict[str, Any]) -> list[Any]:
        return [
            row["manager_name"],
            row["source_company_name"],
            row["fund_name"],
            row["fund_type"],
            row["is_current_company"],
            row["company_name"],
            row["instrument"],
            row["discovery_contrib"],
            row["review_contrib"],
            _json_default(row["investment_date"]),
        ]

    return [
        ("1. 대표펀드매니저 투자회수실적", ["대표펀드매니저", "재직회사명", "펀드명", "펀드 법적형태", "현재 회사 여부", "투자기업명", "투자형태", "발굴 기여율(%)", "심사 기여율(%)", "투자일"], [
            _recovery_row(row) for row in representative_rows
        ]),
        ("2. 참여인력 투자회수실적", ["참여인력", "재직회사명", "펀드명", "펀드 법적형태", "현재 회사 여부", "투자기업명", "투자형태", "발굴 기여율(%)", "심사 기여율(%)", "투자일"], [
            _recovery_row(row) for row in participant_rows
        ]),
        ("3. 최근 개인 투자회수 현황", ["투자기업명", "투자형태", "투자일", "회수일", "투자금액", "회수금액", "잔여금액"], [
            [row["company_name"], row["instrument"], _json_default(row["investment_date"]), _json_default(row["exit_date"]), row["amount"], row["exit_amount"], row["remaining_amount"]]
            for row in exit_rows
        ]),
        ("4. 최근 펀드 투자회수 현황", ["펀드명", "투자기업명", "투자일", "회수일", "투자금액", "회수금액"], [
            [row["fund_name"], row["company_name"], _json_default(row["investment_date"]), _json_default(row["exit_date"]), row["amount"], row["exit_amount"]]
            for row in exit_rows
        ]),
    ]


def _motae7_sheets(workspace: dict[str, Any], db: Session) -> list[tuple[str, list[str], list[list[Any]]]]:
    views = _proposal_views(workspace, db)
    profiles = views["proposal_manager_profiles"]
    representative_ids: set[int] = views["representative_manager_ids"]
    key_manager_ids: set[int] = views["key_manager_ids"]
    core_profile_ids = {
        row["id"]
        for row in views["proposal_managers"]
        if row["is_core"] and row["id"] not in representative_ids
    }
    investments = views["proposal_manager_investments"]

    def _profile_row(row: dict[str, Any]) -> list[Any]:
        return [
            row["name"],
            row["position"],
            _json_default(row["birth_date"]),
            row["phone"],
            row["fax"],
            row["email"],
            row["award_name"],
            row["education_period"],
            row["school_name"],
            row["major"],
            row["degree"],
            row["career_period"],
            row["company_name"],
            row["company_type"],
            row["department"],
            row["career_position"],
            row["main_task"],
            row["employment_type"],
        ]

    return [
        ("2. 참여인력(대표) 이력", ["성명", "직위", "생년월일", "전화", "팩스", "E-Mail", "수상내역", "학력기간", "학교명", "학과(전공)명", "학위명", "경력기간", "직장명(운용사 구분)", "회사유형", "부서명", "직위", "담당업무", "경력구분"], [
            _profile_row(row) for row in profiles if row["manager_id"] in representative_ids
        ]),
        ("2. 참여인력(핵심) 이력", ["성명", "직위", "생년월일", "전화", "팩스", "E-Mail", "수상내역", "학력기간", "학교명", "학과(전공)명", "학위명", "경력기간", "직장명(운용사 구분)", "회사유형", "부서명", "직위", "담당업무", "경력구분"], [
            _profile_row(row) for row in profiles if row["manager_id"] in core_profile_ids
        ]),
        ("3. 핵심운용인력 참여펀드 현황", ["성명", "참여펀드명", "참여펀드 역할", "투자금액", "잔여금액"], [
            [row["manager_name"], row["fund_name"], row["role"], row["amount"], row["remaining_amount"]]
            for row in investments if row["fund_manager_id"] in key_manager_ids
        ]),
        ("4. 핵심운용인력 이직내역", ["성명", "이전 회사", "이전 펀드명", "역할", "이직일"], [
            [row["manager_name"], row["source_company_name"], row["fund_name"], row["role"], _json_default(row["exit_date"])]
            for row in investments if row["fund_manager_id"] in key_manager_ids and not row["is_current_company"]
        ]),
    ]


def _nong_motae_sheets(workspace: dict[str, Any], db: Session) -> list[tuple[str, list[str], list[list[Any]]]]:
    gp = workspace.get("selected_gp_entity") or {}
    views = _proposal_views(workspace, db)
    latest_financial: GPFinancial | None = views["latest_financial"]
    funds = views["proposal_funds"]
    managers = views["proposal_managers"]
    investments = views["proposal_manager_investments"]
    subscriptions = views["proposal_subscriptions"]
    shareholders: list[GPShareholder] = views["proposal_shareholders"]
    single_selected_fund = views["single_selected_fund"]
    representative_manager = views["representative_manager"]

    return [
        ("(별첨1) 투자신청서", ["회사명", "설립연월일", "주소", "운용인력수", "납입자본금", "자산총계", "자본총계", "담당자", "조합명(예정)", "약정금액", "대표펀드매니저"], [[
            gp.get("name"),
            _json_default(gp.get("founding_date")),
            gp.get("address"),
            gp.get("fund_manager_count"),
            latest_financial.paid_in_capital if latest_financial else gp.get("paid_in_capital"),
            latest_financial.total_assets if latest_financial else None,
            latest_financial.total_equity if latest_financial else None,
            gp.get("representative"),
            single_selected_fund.get("name") if single_selected_fund else None,
            single_selected_fund.get("commitment_total") if single_selected_fund else None,
            representative_manager.get("name") if representative_manager else None,
        ]]),
        ("(별첨6-1) 운용펀드 현황 총괄", ["운용사명", "펀드명", "투자기업명", "투자일", "투자방식"], [
            [gp.get("name"), row["fund_name"], row["company_name"], _json_default(row["investment_date"]), row["instrument"]]
            for row in investments if row["is_current_company"]
        ]),
        ("(별첨6-2) 펀드별 투자 현황", ["조합명", "조합 구분", "대표/펀드매니저", "등록(성립)일", "청산시기(예정)", "약정총액", "납입총액", "투자총액", "조합가치"], [
            [fund["name"], fund.get("type"), fund.get("fund_manager"), _json_default(fund.get("registration_date") or fund.get("formation_date")), _json_default(fund.get("maturity_date")), fund.get("commitment_total"), fund.get("paid_in_total"), fund.get("invested_total"), fund.get("nav_total")]
            for fund in funds
        ]),
        ("(별첨6-3) 청약이력 투자 현황", ["조합명", "청약시기", "약정총액", "납입총액", "투자총액", "목표수익률"], [
            [row["fund_name"], _json_default(row["subscription_date"]), row["commitment_total"], row["paid_in_total"], row["invested_total"], row["target_irr"]]
            for row in subscriptions
        ]),
        ("(별첨7-1) 인적자원 관련", ["성명", "직위", "소속", "입사일", "핵심여부"], [
            [row["name"], row["position"], gp.get("name"), _json_default(row["join_date"]), row["is_core"]]
            for row in managers
        ]),
        ("(별첨7-2) 참여인력 투자회수실적", ["참여인력", "재직회사명", "펀드명", "투자기업명", "투자형태", "투자시작일", "회수종료일", "투자금액", "회수금액"], [
            [row["manager_name"], row["source_company_name"], row["fund_name"], row["company_name"], row["instrument"], _json_default(row["investment_date"]), _json_default(row["exit_date"]), row["amount"], row["exit_amount"]]
            for row in investments if row["has_exit"]
        ]),
        ("(별첨7-3) 참여인력 조합청산실적", ["참여인력", "재직회사명", "조합명", "약정총액", "투자금액", "매니저구분"], [
            [row["manager_name"], row["source_company_name"], row["fund_name"], row["commitment_total"], row["amount"], row["role"]]
            for row in investments
        ]),
        ("(별첨8) 제안사 주주명부 및 조합 출자자명부", ["주주명", "주식수", "납입액", "지분율", "비고"], [
            [row.name, row.shares, row.acquisition_amount, row.ownership_pct, row.memo]
            for row in shareholders
        ]),
    ]


def _template_label(template_type: str) -> str:
    return _TEMPLATE_LABELS.get(template_type, template_type)


def _json_load(raw: str | None, fallback: Any = None) -> Any:
    if raw is None:
        return fallback
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return fallback


def _serialize_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=_json_default)


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _build_row_cells(columns: list[dict[str, str]], values: list[Any]) -> dict[str, Any]:
    cells: dict[str, Any] = {}
    for index, column in enumerate(columns):
        cells[column["key"]] = values[index] if index < len(values) else None
    return cells


def _catalog_for_template(template_type: str) -> list[dict[str, Any]]:
    catalog = _WORKBENCH_SHEET_CATALOG.get(template_type)
    if catalog is None:
        raise ValueError(f"Unsupported template type: {template_type}")
    return catalog


def _base_sheet_defs_for_workbench(workspace: dict[str, Any], db: Session) -> list[dict[str, Any]]:
    template_type = workspace["template_type"]
    if template_type == "growth-finance":
        raw_defs = _growth_finance_sheets(workspace, db)
    elif template_type == "motae-5":
        raw_defs = _motae5_sheets(workspace, db)
    elif template_type == "motae-6":
        raw_defs = _motae6_sheets(workspace, db)
    elif template_type == "motae-7":
        raw_defs = _motae7_sheets(workspace, db)
    else:
        raw_defs = _nong_motae_sheets(workspace, db)

    catalog = _catalog_for_template(template_type)
    sheet_defs: list[dict[str, Any]] = []
    for index, meta in enumerate(catalog):
        raw_rows: list[list[Any]] = []
        if index < len(raw_defs):
            _, _, raw_rows = raw_defs[index]
        columns = [{"key": key, "label": label} for key, label in meta["columns"]]
        sheet_defs.append(
            {
                "code": meta["code"],
                "title": meta["title"],
                "kind": meta["kind"],
                "description": meta.get("description"),
                "columns": columns,
                "rows": raw_rows,
            }
        )
    return sheet_defs


def _load_application_fund_ids(db: Session, application_id: int) -> list[int]:
    rows = (
        db.query(ProposalApplicationFund)
        .filter(ProposalApplicationFund.application_id == application_id)
        .order_by(ProposalApplicationFund.display_order.asc(), ProposalApplicationFund.id.asc())
        .all()
    )
    return [row.fund_id for row in rows]


def proposal_application_to_response(db: Session, application: ProposalApplication) -> dict[str, Any]:
    gp_name = None
    if application.gp_entity_id is not None:
        gp = db.get(GPEntity, application.gp_entity_id)
        gp_name = gp.name if gp else None
    fund_ids = _load_application_fund_ids(db, application.id)
    return {
        "id": application.id,
        "title": application.title,
        "template_type": application.template_type,
        "institution_type": application.institution_type,
        "gp_entity_id": application.gp_entity_id,
        "gp_entity_name": gp_name,
        "as_of_date": application.as_of_date,
        "status": application.status,
        "submitted_at": application.submitted_at,
        "created_by": application.created_by,
        "created_at": application.created_at,
        "updated_at": application.updated_at,
        "fund_ids": fund_ids,
        "fund_count": len(fund_ids),
    }


def _replace_application_funds(db: Session, application: ProposalApplication, fund_ids: list[int]) -> None:
    (
        db.query(ProposalApplicationFund)
        .filter(ProposalApplicationFund.application_id == application.id)
        .delete(synchronize_session=False)
    )
    for index, fund_id in enumerate(fund_ids):
        db.add(
            ProposalApplicationFund(
                application_id=application.id,
                fund_id=fund_id,
                display_order=index,
            )
        )


def create_proposal_application(
    db: Session,
    *,
    title: str,
    template_type: str,
    institution_type: str | None,
    gp_entity_id: int | None,
    as_of_date: date,
    fund_ids: list[int],
    created_by_user: User | None,
) -> ProposalApplication:
    application = ProposalApplication(
        title=title,
        template_type=template_type,
        institution_type=institution_type,
        gp_entity_id=gp_entity_id,
        as_of_date=as_of_date,
        status="draft",
        created_by=(created_by_user.id if created_by_user else None),
    )
    db.add(application)
    db.flush()
    _replace_application_funds(db, application, fund_ids)
    db.flush()
    return application


def update_proposal_application(
    db: Session,
    *,
    application: ProposalApplication,
    title: str | None = None,
    template_type: str | None = None,
    institution_type: str | None = None,
    has_gp_entity: bool = False,
    gp_entity_id: int | None = None,
    as_of_date: date | None = None,
    status: str | None = None,
    fund_ids: list[int] | None = None,
) -> ProposalApplication:
    if title is not None:
        application.title = title
    if template_type is not None:
        application.template_type = template_type
    if institution_type is not None:
        application.institution_type = institution_type
    if has_gp_entity:
        application.gp_entity_id = gp_entity_id
    if as_of_date is not None:
        application.as_of_date = as_of_date
    if status is not None:
        application.status = status
    application.updated_at = datetime.utcnow()
    if fund_ids is not None:
        _replace_application_funds(db, application, fund_ids)
    db.flush()
    return application


def _load_field_overrides_by_sheet(db: Session, application_id: int) -> dict[str, dict[str, ProposalFieldOverride]]:
    rows = (
        db.query(ProposalFieldOverride)
        .filter(ProposalFieldOverride.application_id == application_id)
        .all()
    )
    result: dict[str, dict[str, ProposalFieldOverride]] = {}
    for row in rows:
        result.setdefault(row.sheet_code, {})[row.field_key] = row
    return result


def _load_row_overrides_by_sheet(db: Session, application_id: int) -> dict[str, dict[str, ProposalRowOverride]]:
    rows = (
        db.query(ProposalRowOverride)
        .filter(ProposalRowOverride.application_id == application_id)
        .all()
    )
    result: dict[str, dict[str, ProposalRowOverride]] = {}
    for row in rows:
        result.setdefault(row.sheet_code, {})[row.row_key] = row
    return result


def replace_field_overrides(
    db: Session,
    *,
    application: ProposalApplication,
    sheet_code: str,
    overrides: list[dict[str, Any]],
) -> None:
    (
        db.query(ProposalFieldOverride)
        .filter(
            ProposalFieldOverride.application_id == application.id,
            ProposalFieldOverride.sheet_code == sheet_code,
        )
        .delete(synchronize_session=False)
    )
    for override in overrides:
        db.add(
            ProposalFieldOverride(
                application_id=application.id,
                sheet_code=sheet_code,
                field_key=str(override["field_key"]),
                value_json=_serialize_json(override.get("value")),
                source_note=override.get("source_note"),
            )
        )
    application.updated_at = datetime.utcnow()
    db.flush()


def replace_row_overrides(
    db: Session,
    *,
    application: ProposalApplication,
    sheet_code: str,
    overrides: list[dict[str, Any]],
) -> None:
    (
        db.query(ProposalRowOverride)
        .filter(
            ProposalRowOverride.application_id == application.id,
            ProposalRowOverride.sheet_code == sheet_code,
        )
        .delete(synchronize_session=False)
    )
    for override in overrides:
        db.add(
            ProposalRowOverride(
                application_id=application.id,
                sheet_code=sheet_code,
                row_key=str(override["row_key"]),
                row_mode=str(override.get("row_mode") or "override"),
                row_payload_json=_serialize_json(override.get("row_payload") or {}),
                source_note=override.get("source_note"),
            )
        )
    application.updated_at = datetime.utcnow()
    db.flush()


def _build_sheet_copy_text(
    *,
    kind: str,
    columns: list[dict[str, str]],
    fields: list[dict[str, Any]],
    rows: list[dict[str, Any]],
) -> str:
    if kind == "scalar":
        header = "\t".join(field["label"] for field in fields)
        values = "\t".join(_stringify_cell(field["final_value"]) for field in fields)
        return f"{header}\n{values}" if fields else ""

    lines = ["\t".join(column["label"] for column in columns)]
    for row in rows:
        lines.append("\t".join(_stringify_cell(row["final_cells"].get(column["key"])) for column in columns))
    return "\n".join(lines)


def _resolve_sheet_view(
    *,
    application: ProposalApplication,
    sheet_def: dict[str, Any],
    field_overrides: dict[str, ProposalFieldOverride],
    row_overrides: dict[str, ProposalRowOverride],
) -> dict[str, Any]:
    columns: list[dict[str, str]] = sheet_def["columns"]
    raw_rows: list[list[Any]] = sheet_def["rows"]
    kind = sheet_def["kind"]

    rows: list[dict[str, Any]] = []
    fields: list[dict[str, Any]] = []

    if kind == "scalar":
        default_cells = _build_row_cells(columns, raw_rows[0] if raw_rows else [])
        for column in columns:
            override = field_overrides.get(column["key"])
            final_value = _json_load(override.value_json) if override is not None else default_cells.get(column["key"])
            fields.append(
                {
                    "key": column["key"],
                    "label": column["label"],
                    "default_value": default_cells.get(column["key"]),
                    "final_value": final_value,
                    "source": "초안 오버라이드" if override is not None else "자동 불러오기",
                    "is_overridden": override is not None,
                }
            )
    else:
        for index, raw_row in enumerate(raw_rows):
            row_key = f"base:{index + 1}"
            default_cells = _build_row_cells(columns, raw_row)
            override = row_overrides.get(row_key)
            if override is not None and override.row_mode == "hide":
                continue
            payload = _json_load(override.row_payload_json, fallback={}) if override is not None else {}
            final_cells = {**default_cells, **(payload if isinstance(payload, dict) else {})}
            rows.append(
                {
                    "row_key": row_key,
                    "default_cells": default_cells,
                    "final_cells": final_cells,
                    "source": "초안 오버라이드" if override is not None else "자동 불러오기",
                    "is_manual": False,
                    "is_overridden": override is not None,
                }
            )

        for row_key, override in row_overrides.items():
            if row_key.startswith("base:") or override.row_mode == "hide":
                continue
            payload = _json_load(override.row_payload_json, fallback={})
            final_cells = payload if isinstance(payload, dict) else {}
            rows.append(
                {
                    "row_key": row_key,
                    "default_cells": {},
                    "final_cells": {column["key"]: final_cells.get(column["key"]) for column in columns},
                    "source": "수기 추가",
                    "is_manual": True,
                    "is_overridden": True,
                }
            )

    return {
        "application_id": application.id,
        "sheet_code": sheet_def["code"],
        "title": sheet_def["title"],
        "kind": kind,
        "description": sheet_def.get("description"),
        "columns": columns,
        "fields": fields,
        "rows": rows,
        "copy_text": _build_sheet_copy_text(kind=kind, columns=columns, fields=fields, rows=rows),
        "download_filename": f"{application.id}_{sheet_def['code']}.xlsx",
        "is_frozen": application.status == "frozen",
    }


def _build_sheet_descriptor(sheet_view: dict[str, Any]) -> dict[str, Any]:
    empty_value_count = 0
    if sheet_view["kind"] == "scalar":
        empty_value_count = sum(1 for field in sheet_view["fields"] if field["final_value"] in (None, ""))
    else:
        for row in sheet_view["rows"]:
            empty_value_count += sum(1 for value in row["final_cells"].values() if value in (None, ""))
    has_overrides = any(field["is_overridden"] for field in sheet_view["fields"]) or any(row["is_overridden"] for row in sheet_view["rows"])
    return {
        "code": sheet_view["sheet_code"],
        "title": sheet_view["title"],
        "kind": sheet_view["kind"],
        "description": sheet_view.get("description"),
        "row_count": len(sheet_view["rows"]),
        "field_count": len(sheet_view["fields"]),
        "has_overrides": has_overrides,
        "empty_value_count": empty_value_count,
    }


def _resolve_live_application_data(db: Session, application: ProposalApplication) -> dict[str, Any]:
    fund_ids = _load_application_fund_ids(db, application.id)
    workspace = resolve_proposal_workspace(
        db,
        template_type=application.template_type,
        as_of_date=application.as_of_date,
        gp_entity_id=application.gp_entity_id,
        fund_ids=fund_ids,
        auto_select_default_gp=False,
        auto_select_all_funds=False,
    )
    field_overrides = _load_field_overrides_by_sheet(db, application.id)
    row_overrides = _load_row_overrides_by_sheet(db, application.id)
    sheet_views = []
    for sheet_def in _base_sheet_defs_for_workbench(workspace, db):
        sheet_views.append(
            _resolve_sheet_view(
                application=application,
                sheet_def=sheet_def,
                field_overrides=field_overrides.get(sheet_def["code"], {}),
                row_overrides=row_overrides.get(sheet_def["code"], {}),
            )
        )
    return {
        "application": proposal_application_to_response(db, application),
        "readiness": workspace["readiness"],
        "descriptors": [_build_sheet_descriptor(sheet_view) for sheet_view in sheet_views],
        "sheet_views": sheet_views,
    }


def _get_latest_application_snapshot(db: Session, application_id: int) -> ProposalSnapshot | None:
    return (
        db.query(ProposalSnapshot)
        .filter(ProposalSnapshot.application_id == application_id)
        .order_by(ProposalSnapshot.created_at.desc(), ProposalSnapshot.id.desc())
        .first()
    )


def _load_snapshot_payload(db: Session, application: ProposalApplication) -> dict[str, Any] | None:
    snapshot = _get_latest_application_snapshot(db, application.id)
    if snapshot is None:
        return None
    parsed = _json_load(snapshot.payload_json, fallback=None)
    return parsed if isinstance(parsed, dict) else None


def _resolve_application_payload(db: Session, application: ProposalApplication) -> dict[str, Any]:
    if application.status == "frozen":
        payload = _load_snapshot_payload(db, application)
        if payload is not None:
            return payload
    return _resolve_live_application_data(db, application)


def list_proposal_applications(db: Session) -> list[dict[str, Any]]:
    rows = (
        db.query(ProposalApplication)
        .order_by(ProposalApplication.updated_at.desc(), ProposalApplication.id.desc())
        .all()
    )
    return [proposal_application_to_response(db, row) for row in rows]


def get_proposal_application_detail(db: Session, application: ProposalApplication) -> dict[str, Any]:
    payload = _resolve_application_payload(db, application)
    detail = dict(payload["application"])
    detail["readiness"] = payload["readiness"]
    return detail


def list_proposal_application_sheets(db: Session, application: ProposalApplication) -> list[dict[str, Any]]:
    return _resolve_application_payload(db, application)["descriptors"]


def get_proposal_application_sheet(db: Session, application: ProposalApplication, sheet_code: str) -> dict[str, Any]:
    for sheet_view in _resolve_application_payload(db, application)["sheet_views"]:
        if sheet_view["sheet_code"] == sheet_code:
            return sheet_view
    raise KeyError(sheet_code)


def freeze_proposal_application(db: Session, *, application: ProposalApplication) -> ProposalApplication:
    payload = _resolve_live_application_data(db, application)
    (
        db.query(ProposalSnapshot)
        .filter(ProposalSnapshot.application_id == application.id)
        .delete(synchronize_session=False)
    )
    db.add(
        ProposalSnapshot(
            application_id=application.id,
            snapshot_type="resolved",
            payload_json=_serialize_json(payload),
        )
    )
    application.status = "frozen"
    application.submitted_at = application.submitted_at or datetime.utcnow()
    application.updated_at = datetime.utcnow()
    db.flush()
    return application


def _sheet_rows_for_export(sheet_view: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    if sheet_view["kind"] == "scalar":
        return (
            ["항목", "값", "출처"],
            [[field["label"], field["final_value"], field["source"]] for field in sheet_view["fields"]],
        )
    return (
        [column["label"] for column in sheet_view["columns"]],
        [[row["final_cells"].get(column["key"]) for column in sheet_view["columns"]] for row in sheet_view["rows"]],
    )


def _export_single_sheet_xlsx(application: ProposalApplication, sheet_view: dict[str, Any]) -> tuple[bytes, str, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_view["title"][:31]
    headers, rows = _sheet_rows_for_export(sheet_view)
    _write_sheet(ws, headers, rows)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), f"{application.id}_{sheet_view['sheet_code']}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _export_all_sheets_xlsx(application: ProposalApplication, sheet_views: list[dict[str, Any]]) -> tuple[bytes, str, str]:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_view in sheet_views:
        ws = wb.create_sheet(sheet_view["title"][:31])
        headers, rows = _sheet_rows_for_export(sheet_view)
        _write_sheet(ws, headers, rows)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), f"{application.id}_{application.template_type}_draft.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _export_single_sheet_csv(application: ProposalApplication, sheet_view: dict[str, Any]) -> tuple[bytes, str, str]:
    headers, rows = _sheet_rows_for_export(sheet_view)
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig"), f"{application.id}_{sheet_view['sheet_code']}.csv", "text/csv; charset=utf-8"


def export_proposal_application(
    db: Session,
    *,
    application: ProposalApplication,
    scope: str = "sheet",
    format_type: str = "xlsx",
    sheet_code: str | None = None,
) -> tuple[bytes, str, str]:
    payload = _resolve_application_payload(db, application)
    sheet_views = payload["sheet_views"]
    if scope == "sheet":
        if not sheet_code:
            raise ValueError("sheet_code is required for sheet export")
        target = next((sheet for sheet in sheet_views if sheet["sheet_code"] == sheet_code), None)
        if target is None:
            raise KeyError(sheet_code)
        if format_type == "csv":
            return _export_single_sheet_csv(application, target)
        return _export_single_sheet_xlsx(application, target)
    return _export_all_sheets_xlsx(application, sheet_views)
