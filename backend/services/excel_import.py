from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.fund import LP
from models.investment import Investment
from models.phase3 import Distribution
from models.transaction import Transaction
from models.valuation import Valuation

_IMPORT_REQUIRED_COLUMNS: dict[str, list[str]] = {
    "investments": ["fund_id", "company_id", "amount"],
    "lps": ["fund_id", "name", "type", "commitment"],
    "transactions": ["investment_id", "fund_id", "company_id", "transaction_date", "type", "amount"],
    "valuations": ["investment_id", "fund_id", "company_id", "as_of_date", "value"],
}


def _normalize_header(value: str | None) -> str:
    return (value or "").strip().lower()


def _coerce_number(value) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def _parse_date(value) -> date:
    if isinstance(value, date):
        return value
    if value is None:
        raise ValueError("date is required")
    return date.fromisoformat(str(value))


def _sheet_rows(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
    headers = [_normalize_header(str(value)) for value in raw_headers]

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        row_dict: dict[str, object] = {}
        has_value = False
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            if value not in (None, ""):
                has_value = True
            row_dict[header] = value
        if has_value:
            rows.append(row_dict)
    return headers, rows


async def parse_excel_preview(file_bytes: bytes, import_type: str) -> dict:
    normalized_type = _normalize_header(import_type)
    if normalized_type not in _IMPORT_REQUIRED_COLUMNS:
        raise ValueError("unsupported import_type")

    headers, rows = _sheet_rows(file_bytes)
    required = _IMPORT_REQUIRED_COLUMNS[normalized_type]

    errors: list[dict] = []
    valid_rows = 0
    for index, row in enumerate(rows, start=2):
        row_errors: list[str] = []
        for col in required:
            if row.get(col) in (None, ""):
                row_errors.append(f"{col} 누락")
        if row_errors:
            errors.append({"row": index, "errors": row_errors})
        else:
            valid_rows += 1

    preview_rows = rows[:10]
    return {
        "headers": headers,
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "error_rows": errors,
        "preview": preview_rows,
    }


async def confirm_excel_import(
    db: Session,
    file_bytes: bytes,
    import_type: str,
    options: dict,
) -> dict:
    normalized_type = _normalize_header(import_type)
    if normalized_type not in _IMPORT_REQUIRED_COLUMNS:
        raise ValueError("unsupported import_type")

    _, rows = _sheet_rows(file_bytes)
    required = _IMPORT_REQUIRED_COLUMNS[normalized_type]

    imported_count = 0
    skipped_count = 0
    errors: list[str] = []

    for index, row in enumerate(rows, start=2):
        if any(row.get(col) in (None, "") for col in required):
            skipped_count += 1
            errors.append(f"row {index}: required columns missing")
            continue

        try:
            if normalized_type == "investments":
                db.add(
                    Investment(
                        fund_id=int(row["fund_id"]),
                        company_id=int(row["company_id"]),
                        amount=_coerce_number(row.get("amount")),
                        investment_date=_parse_date(row.get("investment_date")) if row.get("investment_date") not in (None, "") else None,
                        instrument=str(row.get("instrument") or "").strip() or None,
                        status=str(row.get("status") or "active").strip() or "active",
                    )
                )
            elif normalized_type == "lps":
                db.add(
                    LP(
                        fund_id=int(row["fund_id"]),
                        name=str(row.get("name") or "").strip(),
                        type=str(row.get("type") or "법인").strip(),
                        commitment=int(round(_coerce_number(row.get("commitment")))),
                        paid_in=int(round(_coerce_number(row.get("paid_in")))),
                        contact=str(row.get("contact") or "").strip() or None,
                        business_number=str(row.get("business_number") or "").strip() or None,
                        address=str(row.get("address") or "").strip() or None,
                    )
                )
            elif normalized_type == "transactions":
                db.add(
                    Transaction(
                        investment_id=int(row["investment_id"]),
                        fund_id=int(row["fund_id"]),
                        company_id=int(row["company_id"]),
                        transaction_date=_parse_date(row.get("transaction_date")),
                        type=str(row.get("type") or "").strip(),
                        amount=_coerce_number(row.get("amount")),
                        transaction_subtype=str(row.get("transaction_subtype") or "").strip() or None,
                        counterparty=str(row.get("counterparty") or "").strip() or None,
                        memo=str(row.get("memo") or "").strip() or None,
                    )
                )
            elif normalized_type == "valuations":
                db.add(
                    Valuation(
                        investment_id=int(row["investment_id"]),
                        fund_id=int(row["fund_id"]),
                        company_id=int(row["company_id"]),
                        as_of_date=_parse_date(row.get("as_of_date")),
                        value=_coerce_number(row.get("value")),
                        method=str(row.get("method") or "import").strip() or "import",
                        evaluator=str(row.get("evaluator") or "excel_import").strip() or "excel_import",
                    )
                )
            imported_count += 1
        except Exception as exc:  # noqa: BLE001
            skipped_count += 1
            errors.append(f"row {index}: {exc}")

    try:
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        raise ValueError(f"import failed: {exc}") from exc

    return {
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "errors": errors,
    }
