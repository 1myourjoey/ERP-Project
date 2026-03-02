from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from models.compliance import ComplianceObligation, ComplianceRule
from models.fund import Fund
from models.investment import Investment, PortfolioCompany
from models.phase3 import CapitalCall, Distribution
from models.task import Task
from models.transaction import Transaction
from models.worklog import WorkLog

_HEADER_FILL = PatternFill(fill_type="solid", start_color="1F6FB8", end_color="1F6FB8")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_CELL_BORDER = Border(
    left=Side(style="thin", color="D9E2EF"),
    right=Side(style="thin", color="D9E2EF"),
    top=Side(style="thin", color="D9E2EF"),
    bottom=Side(style="thin", color="D9E2EF"),
)


def _auto_fit(ws) -> None:
    for col in ws.columns:
        length = 10
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value) + 2, 48))
            cell.border = _CELL_BORDER
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[col_letter].width = length


def _apply_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_styled_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    return wb


async def export_fund_summary(db: Session, fund_id: int) -> bytes:
    fund = db.get(Fund, fund_id)
    if not fund:
        raise ValueError("fund not found")

    wb = create_styled_workbook()

    ws_info = wb.active
    ws_info.title = "Fund"
    _apply_header(ws_info, ["항목", "값"])
    ws_info.append(["조합명", fund.name])
    ws_info.append(["유형", fund.type])
    ws_info.append(["상태", fund.status])
    ws_info.append(["총약정", float(fund.commitment_total or 0)])
    ws_info.append(["AUM", float(fund.aum or 0)])
    _auto_fit(ws_info)

    ws_lp = wb.create_sheet("LP")
    _apply_header(ws_lp, ["LP명", "유형", "약정", "납입", "연락처", "사업자번호"])
    for lp in fund.lps:
        ws_lp.append([
            lp.name,
            lp.type,
            float(lp.commitment or 0),
            float(lp.paid_in or 0),
            lp.contact,
            lp.business_number,
        ])
    _auto_fit(ws_lp)

    ws_inv = wb.create_sheet("Investments")
    _apply_header(ws_inv, ["ID", "기업", "투자일", "금액", "유형", "상태"])
    investments = db.query(Investment).filter(Investment.fund_id == fund_id).order_by(Investment.id.asc()).all()
    for row in investments:
        company = db.get(PortfolioCompany, row.company_id)
        ws_inv.append([
            row.id,
            company.name if company else row.company_id,
            row.investment_date.isoformat() if row.investment_date else None,
            float(row.amount or 0),
            row.instrument,
            row.status,
        ])
    _auto_fit(ws_inv)

    ws_calls = wb.create_sheet("CapitalCalls")
    _apply_header(ws_calls, ["ID", "요청일", "유형", "금액", "메모"])
    calls = db.query(CapitalCall).filter(CapitalCall.fund_id == fund_id).order_by(CapitalCall.call_date.asc()).all()
    for row in calls:
        ws_calls.append([row.id, row.call_date.isoformat(), row.call_type, float(row.total_amount or 0), row.memo])
    _auto_fit(ws_calls)

    ws_dist = wb.create_sheet("Distributions")
    _apply_header(ws_dist, ["ID", "배분일", "유형", "원금", "수익", "메모"])
    distributions = db.query(Distribution).filter(Distribution.fund_id == fund_id).order_by(Distribution.dist_date.asc()).all()
    for row in distributions:
        ws_dist.append([
            row.id,
            row.dist_date.isoformat(),
            row.dist_type,
            float(row.principal_total or 0),
            float(row.profit_total or 0),
            row.memo,
        ])
    _auto_fit(ws_dist)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def export_investments(db: Session, filters: dict) -> bytes:
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Investments"
    _apply_header(ws, ["ID", "Fund", "Company", "Date", "Amount", "Instrument", "Status"])

    query = db.query(Investment)
    fund_id = filters.get("fund_id")
    if fund_id:
        query = query.filter(Investment.fund_id == int(fund_id))
    rows = query.order_by(Investment.id.desc()).all()
    for row in rows:
        fund = db.get(Fund, row.fund_id)
        company = db.get(PortfolioCompany, row.company_id)
        ws.append([
            row.id,
            fund.name if fund else row.fund_id,
            company.name if company else row.company_id,
            row.investment_date.isoformat() if row.investment_date else None,
            float(row.amount or 0),
            row.instrument,
            row.status,
        ])

    _auto_fit(ws)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def export_transactions(db: Session, filters: dict) -> bytes:
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Transactions"
    _apply_header(ws, ["ID", "Fund", "Company", "Date", "Type", "Amount", "Counterparty", "Memo"])

    query = db.query(Transaction)
    fund_id = filters.get("fund_id")
    if fund_id:
        query = query.filter(Transaction.fund_id == int(fund_id))
    rows = query.order_by(Transaction.transaction_date.desc(), Transaction.id.desc()).all()
    for row in rows:
        fund = db.get(Fund, row.fund_id)
        company = db.get(PortfolioCompany, row.company_id)
        ws.append([
            row.id,
            fund.name if fund else row.fund_id,
            company.name if company else row.company_id,
            row.transaction_date.isoformat() if row.transaction_date else None,
            row.type,
            float(row.amount or 0),
            row.counterparty,
            row.memo,
        ])

    _auto_fit(ws)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def export_compliance_report(db: Session, fund_id: int, year: int, month: int) -> bytes:
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Compliance"
    _apply_header(ws, ["의무ID", "룰코드", "제목", "마감일", "상태", "증빙노트"])

    start = date(year, month, 1)
    if month == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, month + 1, 1) - date.resolution

    rows = (
        db.query(ComplianceObligation)
        .filter(
            ComplianceObligation.fund_id == fund_id,
            ComplianceObligation.due_date >= start,
            ComplianceObligation.due_date <= end,
        )
        .order_by(ComplianceObligation.due_date.asc(), ComplianceObligation.id.asc())
        .all()
    )

    for row in rows:
        rule = db.get(ComplianceRule, row.rule_id)
        ws.append([
            row.id,
            rule.rule_code if rule else None,
            rule.title if rule else None,
            row.due_date.isoformat() if row.due_date else None,
            row.status,
            row.evidence_note,
        ])

    _auto_fit(ws)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def export_worklogs(db: Session, filters: dict) -> bytes:
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "WorkLogs"
    _apply_header(ws, ["ID", "일자", "카테고리", "제목", "상태", "예상시간", "실제시간"])

    query = db.query(WorkLog)
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")
    if date_from:
        query = query.filter(WorkLog.date >= date.fromisoformat(str(date_from)))
    if date_to:
        query = query.filter(WorkLog.date <= date.fromisoformat(str(date_to)))

    rows = query.order_by(WorkLog.date.desc(), WorkLog.id.desc()).all()
    for row in rows:
        ws.append([
            row.id,
            row.date.isoformat() if row.date else None,
            row.category,
            row.title,
            row.status,
            row.estimated_time,
            row.actual_time,
        ])

    _auto_fit(ws)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
