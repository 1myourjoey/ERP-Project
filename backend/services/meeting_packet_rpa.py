from __future__ import annotations

import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

import olefile
import openpyxl
import pdfplumber
from sqlalchemy.orm import Session

from models.biz_report import BizReport
from models.document_template import DocumentTemplate
from models.fund import Fund
from models.investment import Investment
from models.regular_report import RegularReport
from schemas.meeting_packet import (
    MeetingPacketAnalyzeResponse,
    MeetingPacketFileItem,
    MeetingPacketGenerationPlanResponse,
    MeetingPacketGenerationSlot,
    MeetingPacketPackageItem,
)
from services.lp_types import is_special_lp_type


SLOT_LABELS = {
    "official_notice": "공문",
    "agenda_explanation": "의안설명서",
    "audit_report": "감사보고서",
    "financial_statement_certificate": "재무제표 증명원",
    "business_report": "영업보고서",
    "bylaw_amendment_draft": "개정 규약안",
    "bylaw_redline": "규약 신구대조표",
    "written_resolution": "서면의결서",
    "proxy_vote_notice": "의결권행사통보서",
    "minutes": "의사록",
}

GENERATION_MODE = {
    "official_notice": "erp_rpa",
    "agenda_explanation": "erp_rpa",
    "audit_report": "external_receive",
    "financial_statement_certificate": "external_receive",
    "business_report": "erp_rpa",
    "bylaw_amendment_draft": "erp_rpa_assisted",
    "bylaw_redline": "erp_rpa",
    "written_resolution": "erp_rpa",
    "proxy_vote_notice": "erp_rpa",
    "minutes": "erp_rpa_assisted",
}

PACKET_LABELS = {
    "fund_lp_regular_meeting_pex": "조합원총회 + 온기보고회(PEX형)",
    "fund_lp_regular_meeting_project": "조합원총회(프로젝트형)",
    "fund_lp_regular_meeting_project_with_bylaw_amendment": "조합원총회(규약변경 포함)",
    "gp_shareholders_meeting": "GP 사원총회",
    "unknown": "미분류 패키지",
}

PACKET_REQUIRED_SLOTS = {
    "fund_lp_regular_meeting_pex": [
        "official_notice",
        "agenda_explanation",
        "audit_report",
        "business_report",
        "proxy_vote_notice",
    ],
    "fund_lp_regular_meeting_project": [
        "official_notice",
        "agenda_explanation",
        "audit_report",
        "business_report",
        "written_resolution",
        "minutes",
    ],
    "fund_lp_regular_meeting_project_with_bylaw_amendment": [
        "official_notice",
        "agenda_explanation",
        "audit_report",
        "business_report",
        "bylaw_amendment_draft",
        "bylaw_redline",
        "written_resolution",
        "minutes",
    ],
    "gp_shareholders_meeting": [
        "official_notice",
        "agenda_explanation",
        "financial_statement_certificate",
        "written_resolution",
        "minutes",
    ],
}

LAYOUT_RECOMMENDATIONS = {
    "official_notice": "one_page",
    "agenda_explanation": "compact_report",
    "audit_report": "external_attachment",
    "financial_statement_certificate": "external_attachment",
    "business_report": "full_report",
    "bylaw_amendment_draft": "assisted_draft",
    "bylaw_redline": "compact_table",
    "written_resolution": "one_page",
    "proxy_vote_notice": "one_page",
    "minutes": "assisted_minutes",
}

SLOT_TEMPLATE_HINTS = {
    "official_notice": "공문_결성총회_출자이행통지",
    "agenda_explanation": "첨부1_결성총회_소집통지서",
    "written_resolution": "별첨6_서면결의서",
    "proxy_vote_notice": "별첨6_서면결의서",
}


class MeetingPacketRPAService:
    def recommend_packet_type(
        self,
        *,
        fund: Fund,
        include_bylaw_amendment: bool = False,
    ) -> tuple[str, list[str]]:
        reasons: list[str] = []
        normalized_name = (fund.name or "").strip().lower()
        has_special_partner = any(is_special_lp_type(lp.type) for lp in list(fund.lps or []))

        if "pex" in normalized_name:
            reasons.append("조합명에 PEX가 포함되어 PEX형 총회 패키지를 우선 추천합니다.")
        if has_special_partner:
            reasons.append("특별조합원이 포함되어 보고/통지 문서 구성이 중요합니다.")
        if include_bylaw_amendment:
            reasons.append("규약 변경 안건이 포함되어 변경안과 신구대조표 세트를 포함합니다.")

        if "pex" in normalized_name:
            return "fund_lp_regular_meeting_pex", reasons or ["PEX형 총회 패키지입니다."]
        if include_bylaw_amendment:
            return "fund_lp_regular_meeting_project_with_bylaw_amendment", reasons or ["규약 변경 안건 포함 패키지입니다."]
        return "fund_lp_regular_meeting_project", reasons or ["프로젝트 조합 기본 총회 패키지입니다."]

    def analyze_root(self, root_path: str) -> MeetingPacketAnalyzeResponse:
        root = Path(root_path)
        if not root.exists() or not root.is_dir():
            raise ValueError("root_path must be an existing directory")

        packages: list[MeetingPacketPackageItem] = []
        slot_counter: Counter[str] = Counter()

        for folder in sorted([item for item in root.iterdir() if item.is_dir()], key=lambda item: item.name):
            package = self._analyze_folder(folder)
            packages.append(package)
            slot_counter.update(item.slot for item in package.files)

        package_types = Counter(package.packet_type for package in packages)
        common_slots = sorted([slot for slot, count in slot_counter.items() if packages and count == len(packages)])
        varying_slots = sorted([slot for slot, count in slot_counter.items() if count != len(packages)])

        return MeetingPacketAnalyzeResponse(
            root_path=str(root),
            package_count=len(packages),
            packet_types={key: int(value) for key, value in sorted(package_types.items())},
            common_slots=common_slots,
            varying_slots=varying_slots,
            packages=packages,
        )

    def build_generation_plan(
        self,
        *,
        db: Session,
        fund_id: int,
        packet_type: str,
        meeting_date: date | None = None,
        meeting_time: str | None = None,
        meeting_method: str | None = None,
        report_year: int | None = None,
        include_bylaw_amendment: bool = False,
    ) -> MeetingPacketGenerationPlanResponse:
        fund = db.get(Fund, fund_id)
        if not fund:
            raise ValueError("fund not found")

        recommended_type, packet_reasoning = self.recommend_packet_type(
            fund=fund,
            include_bylaw_amendment=include_bylaw_amendment,
        )
        normalized_type = packet_type if packet_type in PACKET_LABELS else recommended_type
        slots = list(PACKET_REQUIRED_SLOTS.get(normalized_type, []))
        if include_bylaw_amendment and "bylaw_amendment_draft" not in slots:
            slots.extend(["bylaw_amendment_draft", "bylaw_redline", "written_resolution"])

        latest_biz_report = (
            db.query(BizReport)
            .filter(BizReport.fund_id == fund_id)
            .order_by(BizReport.report_year.desc(), BizReport.id.desc())
            .first()
        )
        latest_regular_report = (
            db.query(RegularReport)
            .filter(RegularReport.fund_id == fund_id)
            .order_by(RegularReport.created_at.desc(), RegularReport.id.desc())
            .first()
        )
        investments = (
            db.query(Investment)
            .filter(Investment.fund_id == fund_id)
            .order_by(Investment.id.asc())
            .all()
        )

        recipients_preview = [lp.name for lp in list(fund.lps or [])[:10]]
        agenda_preview = self._build_agenda_preview(
            packet_type=normalized_type,
            include_bylaw_amendment=include_bylaw_amendment,
            latest_biz_report=latest_biz_report,
        )

        warnings: list[str] = []
        if latest_biz_report is None and "business_report" in slots:
            warnings.append("최신 영업보고 데이터가 없어 ERP 원천 데이터 기반으로 초안을 계산합니다.")
        if latest_regular_report is None:
            warnings.append("정기보고 기록이 없어 일부 기한/보고 정보는 직접 확인이 필요할 수 있습니다.")
        if report_year is None and latest_biz_report is None:
            warnings.append("보고 연도를 지정하지 않아 가장 최근 회차 기준으로 패키지를 준비합니다.")

        slot_rows = [
            self._build_slot_plan(
                db=db,
                slot=slot,
                fund=fund,
                packet_type=normalized_type,
                latest_biz_report=latest_biz_report,
                latest_regular_report=latest_regular_report,
                investments=investments,
                include_bylaw_amendment=include_bylaw_amendment,
            )
            for slot in slots
        ]

        return MeetingPacketGenerationPlanResponse(
            fund_id=fund.id,
            fund_name=fund.name,
            packet_type=normalized_type,  # type: ignore[arg-type]
            packet_label=PACKET_LABELS.get(normalized_type, normalized_type),
            recommended_packet_type=recommended_type,  # type: ignore[arg-type]
            recommended_packet_label=PACKET_LABELS.get(recommended_type, recommended_type),
            packet_reasoning=packet_reasoning,
            meeting_date=meeting_date.isoformat() if meeting_date else None,
            meeting_time=meeting_time,
            meeting_method=meeting_method or "서면결의",
            slots=slot_rows,
            recipients_preview=recipients_preview,
            agenda_preview=agenda_preview,
            warnings=warnings,
        )

    def _analyze_folder(self, folder: Path) -> MeetingPacketPackageItem:
        file_paths = sorted([item for item in folder.iterdir() if item.is_file()], key=lambda item: item.name)
        analyzed_files: list[MeetingPacketFileItem] = []
        extracted: dict[str, Any] = {
            "fund_name": None,
            "meeting_date": None,
            "meeting_method": None,
            "document_number": None,
            "recipients": [],
            "agendas": [],
            "attachments": [],
        }

        for file_path in file_paths:
            slot = self._classify_slot(file_path.name)
            preview_text = self._extract_preview_text(file_path)
            analyzed_files.append(
                MeetingPacketFileItem(
                    slot=slot,
                    slot_label=SLOT_LABELS.get(slot, slot),
                    filename=file_path.name,
                    extension=file_path.suffix.lower(),
                    full_path=str(file_path),
                    generation_mode=GENERATION_MODE.get(slot, "unknown"),
                    preview_text=preview_text[:1200] if preview_text else None,
                )
            )
            self._merge_extracted_metadata(extracted, file_path.name, preview_text)

        packet_type = self._classify_packet_type(folder.name, analyzed_files, extracted)
        required_slots = PACKET_REQUIRED_SLOTS.get(packet_type, [])
        present_slots = {item.slot for item in analyzed_files}
        missing_slots = [slot for slot in required_slots if slot not in present_slots]
        extra_files = [item.filename for item in analyzed_files if item.slot not in required_slots and item.slot != "unknown"]

        package_notes: list[str] = []
        if packet_type == "fund_lp_regular_meeting_pex":
            package_notes.append("PEX/농모태형 영업보고서 패키지로 보이며 투자보고회 문구가 함께 포함됩니다.")
        if packet_type == "fund_lp_regular_meeting_project_with_bylaw_amendment":
            package_notes.append("규약 변경 안건이 포함되어 개정 규약안과 신구대조표가 함께 필요합니다.")
        if packet_type == "gp_shareholders_meeting":
            package_notes.append("조합 총회가 아니라 GP 법인 사원총회 패키지입니다.")

        return MeetingPacketPackageItem(
            package_name=folder.name,
            folder_path=str(folder),
            packet_type=packet_type,  # type: ignore[arg-type]
            packet_label=PACKET_LABELS.get(packet_type, packet_type),
            fund_name=extracted["fund_name"],
            meeting_date=extracted["meeting_date"],
            meeting_method=extracted["meeting_method"],
            document_number=extracted["document_number"],
            recipients=extracted["recipients"],
            agendas=extracted["agendas"],
            attachments=extracted["attachments"],
            files=analyzed_files,
            missing_slots=missing_slots,
            extra_files=extra_files,
            package_notes=package_notes,
            erp_dependencies=self._packet_dependencies(packet_type),
        )

    def _packet_dependencies(self, packet_type: str) -> list[str]:
        common = ["Fund", "LP/GPEntity", "FundNoticePeriod", "FundKeyTerm"]
        if packet_type == "fund_lp_regular_meeting_pex":
            return [*common, "BizReport", "ComplianceReviewRun", "Investment/Valuation"]
        if packet_type in {"fund_lp_regular_meeting_project", "fund_lp_regular_meeting_project_with_bylaw_amendment"}:
            return [*common, "BizReport", "Investment/Valuation", "RegularReport", "ComplianceReviewRun"]
        if packet_type == "gp_shareholders_meeting":
            return ["GPEntity", "GPProfile", "Financial statements metadata"]
        return common

    def _classify_packet_type(
        self,
        folder_name: str,
        files: list[MeetingPacketFileItem],
        extracted: dict[str, Any],
    ) -> str:
        lowered_name = folder_name.lower()
        file_names = " ".join(item.filename.lower() for item in files)
        preview_joined = " ".join(item.preview_text or "" for item in files[:3]).lower()
        has_bylaw_amendment = any(item.slot in {"bylaw_amendment_draft", "bylaw_redline"} for item in files)

        if "사원총회" in folder_name:
            return "gp_shareholders_meeting"
        if "pex" in lowered_name or "투자보고회" in preview_joined:
            return "fund_lp_regular_meeting_pex"
        if has_bylaw_amendment:
            return "fund_lp_regular_meeting_project_with_bylaw_amendment"
        if "조합원총회" in file_names or extracted.get("fund_name"):
            return "fund_lp_regular_meeting_project"
        return "unknown"

    def _classify_slot(self, filename: str) -> str:
        normalized = filename.replace(" ", "").lower()
        if "공문" in filename:
            return "official_notice"
        if "의안설명서" in filename:
            return "agenda_explanation"
        if "감사보고서" in filename:
            return "audit_report"
        if "재무제표증명원" in filename:
            return "financial_statement_certificate"
        if "영업보고서" in filename:
            return "business_report"
        if "신구대조표" in filename:
            return "bylaw_redline"
        if "규약" in filename and ("개정" in filename or "변경" in filename):
            return "bylaw_amendment_draft"
        if "의결권행사통보서" in filename:
            return "proxy_vote_notice"
        if "서면의결서" in filename:
            return "written_resolution"
        if "의사록" in filename:
            return "minutes"
        if normalized.endswith(".doc") and "의사록" in filename:
            return "minutes"
        return "unknown"

    def _extract_preview_text(self, path: Path) -> str:
        ext = path.suffix.lower()
        try:
            if ext == ".pdf":
                with pdfplumber.open(path) as pdf:
                    return "\n".join((page.extract_text() or "") for page in pdf.pages[:2]).strip()
            if ext == ".xlsx":
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows: list[str] = []
                for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                    values = [str(cell).strip() for cell in row if cell not in (None, "")]
                    if values:
                        rows.append(" | ".join(values))
                return "\n".join(rows)
            if ext == ".hwp":
                ole = olefile.OleFileIO(str(path))
                if ole.exists("PrvText"):
                    return ole.openstream("PrvText").read().decode("utf-16le", errors="ignore").strip()
                return ""
            if ext == ".doc":
                ole = olefile.OleFileIO(str(path))
                meta = ole.get_metadata()
                return " ".join(
                    value.decode("cp949", errors="ignore") if isinstance(value, bytes) else str(value or "")
                    for value in [meta.title, meta.subject, meta.author]
                    if value
                ).strip()
        except Exception:
            return ""
        return ""

    def _merge_extracted_metadata(self, extracted: dict[str, Any], filename: str, preview_text: str) -> None:
        text = f"{filename}\n{preview_text}"
        if not extracted["document_number"]:
            matched = re.search(r"문서번호\s*:\s*([^\n]+)", text)
            if matched:
                extracted["document_number"] = matched.group(1).strip()

        if not extracted["meeting_date"]:
            matched = re.search(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일", text)
            if matched:
                extracted["meeting_date"] = f"{matched.group(1)}-{int(matched.group(2)):02d}-{int(matched.group(3)):02d}"

        if not extracted["meeting_method"] and "서면결의" in text:
            extracted["meeting_method"] = "서면결의"

        if not extracted["fund_name"]:
            for pattern in [
                r"<([^<>]+조합)[^<>]*영업보고서",
                r"([^\n]+조합)\s+규약",
                r"내\s*용\s*:\s*([^\n]+총회[^ \n]*개최의 건)",
            ]:
                matched = re.search(pattern, text)
                if not matched:
                    continue
                value = matched.group(1).strip()
                if "개최의 건" in value:
                    value = (
                        value.replace("정기조합원총회 개최의 건", "")
                        .replace("제1기 ", "")
                        .replace("제2기 ", "")
                        .strip()
                    )
                extracted["fund_name"] = value
                break

        agenda_matches = re.findall(r"제\s*\d+\s*호\s*의안\s*[:)]\s*([^\n]+)", text)
        for agenda in agenda_matches:
            normalized = agenda.strip()
            if normalized and normalized not in extracted["agendas"]:
                extracted["agendas"].append(normalized)

        attachment_matches = re.findall(r"별첨\d+\.\s*([^\n]+)", text)
        for item in attachment_matches:
            normalized = item.strip()
            if normalized and normalized not in extracted["attachments"]:
                extracted["attachments"].append(normalized)

        recipient_match = re.search(r"수\s*신\s*:\s*([^\n]+)", text)
        if recipient_match:
            raw = recipient_match.group(1).strip()
            recipients = [part.strip() for part in re.split(r",|·", raw) if part.strip()]
            for item in recipients:
                if item not in extracted["recipients"]:
                    extracted["recipients"].append(item)

    def _build_agenda_preview(
        self,
        *,
        packet_type: str,
        include_bylaw_amendment: bool,
        latest_biz_report: BizReport | None,
    ) -> list[str]:
        agendas = ["재무제표 승인의 건"]
        if latest_biz_report is not None:
            agendas.insert(0, "감사보고 및 영업보고")
        if packet_type == "gp_shareholders_meeting":
            agendas = ["재무제표 승인의 건", "준법감시인 변경의 건"]
        if include_bylaw_amendment or packet_type == "fund_lp_regular_meeting_project_with_bylaw_amendment":
            agendas.extend(["신규 참여인력 추가의 건", "규약 변경의 건"])
        return agendas

    def _build_slot_plan(
        self,
        *,
        db: Session,
        slot: str,
        fund: Fund,
        packet_type: str,
        latest_biz_report: BizReport | None,
        latest_regular_report: RegularReport | None,
        investments: list[Investment],
        include_bylaw_amendment: bool,
    ) -> MeetingPacketGenerationSlot:
        data_points: list[str] = []
        source_systems: list[str] = []
        notes: list[str] = []
        preflight_warnings: list[str] = []
        required_external_documents: list[str] = []
        status = "ready"
        recommended_layout = LAYOUT_RECOMMENDATIONS.get(slot, "standard")
        builder_candidate: str | None = None
        template_candidate = self._template_candidate_for_slot(db, slot)

        if slot == "official_notice":
            data_points = ["문서번호", "수신자", "규약 근거 조항", "회의일시", "개최방식", "안건 목록", "첨부 목록"]
            source_systems = ["Fund", "LP", "FundNoticePeriod", "FundKeyTerm"]
            builder_candidate = "template"
        elif slot == "agenda_explanation":
            data_points = ["회의 기본정보", "보고사항", "부의안건", "요약 재무현황", "안건별 설명"]
            source_systems = ["Fund", "BizReport", "FundKeyTerm"]
            builder_candidate = "template"
        elif slot == "audit_report":
            data_points = ["외부 회계법인 감사보고서 수령본"]
            source_systems = ["ComplianceDocument"]
            status = "external_required"
            required_external_documents = ["회계법인 최종 감사보고서"]
        elif slot == "financial_statement_certificate":
            data_points = ["외부 발급 재무제표 증명원"]
            source_systems = ["ComplianceDocument"]
            status = "external_required"
            required_external_documents = ["재무제표 증명원 최신본"]
        elif slot == "business_report":
            data_points = ["조합 개요", "투자 현황", "회수 현황", "재무현황", "규약/법령 준수 여부", "특기사항"]
            source_systems = ["Fund", "BizReport", "Investment", "Valuation", "ComplianceReviewRun"]
            builder_candidate = "erp_biz_report_draft"
            if latest_biz_report is None:
                status = "partial"
                preflight_warnings.append("최신 BizReport가 없어 ERP 원천 데이터 기반 초안으로 생성됩니다.")
            if not investments:
                preflight_warnings.append("투자 데이터가 없어 영업보고서 투자 현황 섹션이 비어 있을 수 있습니다.")
            if packet_type == "fund_lp_regular_meeting_pex":
                notes.append("PEX/농모태형 영업보고서 목차를 사용합니다.")
            else:
                notes.append("프로젝트 조합 내부 준표준 영업보고서 목차를 사용합니다.")
        elif slot == "bylaw_amendment_draft":
            data_points = ["변경 대상 조항", "변경 후 문안", "운용인력 변경 내용"]
            source_systems = ["FundKeyTerm", "ComplianceReviewRun", "manual input"]
            status = "assisted"
            builder_candidate = "manual_assisted"
            notes.append("개정 규약안 본문은 ERP 기준값과 수동 문안 검토가 함께 필요합니다.")
        elif slot == "bylaw_redline":
            data_points = ["변경 조항", "변경 전", "변경 후", "비고"]
            source_systems = ["FundKeyTerm", "manual input"]
            status = "assisted" if include_bylaw_amendment or packet_type == "fund_lp_regular_meeting_project_with_bylaw_amendment" else "optional"
            builder_candidate = "erp_redline_table"
        elif slot in {"written_resolution", "proxy_vote_notice"}:
            data_points = ["회의일", "안건 목록", "좌수/약정좌수", "조합원명", "사업자번호/생년월일"]
            source_systems = ["Fund", "LP"]
            builder_candidate = "template"
        elif slot == "minutes":
            data_points = ["성원 보고", "안건별 의결 결과", "첨부 문서", "종료 시각", "서명자"]
            source_systems = ["Fund", "LP", "meeting results input"]
            status = "assisted"
            builder_candidate = "manual_assisted"
            notes.append("총회 종료 후 의결 결과와 출석 정보 입력이 필요합니다.")
            preflight_warnings.append("의결 결과와 출석 정보가 확정된 뒤 최종본 작성이 가능합니다.")

        if slot in {"official_notice", "written_resolution", "proxy_vote_notice"}:
            notes.append("한 페이지에 끝나는 압축형 레이아웃을 우선 사용합니다.")
        if slot == "agenda_explanation":
            notes.append("요약 재무표는 첫 페이지 우선 배치, 상세 설명은 다음 페이지로 넘길 수 있습니다.")

        return MeetingPacketGenerationSlot(
            slot=slot,
            slot_label=SLOT_LABELS.get(slot, slot),
            generation_mode=GENERATION_MODE.get(slot, "unknown"),
            status=status,
            recommended_layout=recommended_layout,
            builder_candidate=builder_candidate,
            template_candidate=template_candidate,
            source_systems=source_systems,
            data_points=data_points,
            preflight_warnings=preflight_warnings,
            required_external_documents=required_external_documents,
            notes=notes,
        )

    def _template_candidate_for_slot(self, db: Session, slot: str) -> str | None:
        hint = SLOT_TEMPLATE_HINTS.get(slot)
        if not hint:
            return None
        row = db.query(DocumentTemplate).filter(DocumentTemplate.name == hint).first()
        return row.name if row else hint
