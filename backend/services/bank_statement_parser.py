from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_AMOUNT_RE = re.compile(r"[^0-9.\-]")


class BankStatementParser:
    """Parse clipboard text or excel bank statements into normalized dict rows."""

    def parse_clipboard_text(self, text: str, fund_id: int) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        if not text or not text.strip():
            return rows

        for raw_line in text.splitlines():
            line = (raw_line or "").strip()
            if not line:
                continue
            if "거래일" in line and ("출금" in line or "입금" in line):
                continue

            parts = self._split_line(line)
            if len(parts) < 5:
                continue

            mapped = self._map_columns(parts)
            transaction_date = self._parse_datetime(mapped.get("transaction_date"))
            if not transaction_date:
                continue

            withdrawal = self._parse_amount(mapped.get("withdrawal"))
            deposit = self._parse_amount(mapped.get("deposit"))
            balance_after = self._parse_amount(mapped.get("balance_after"))

            # Ignore non-financial rows.
            if withdrawal == 0 and deposit == 0 and balance_after == 0:
                continue

            rows.append(
                {
                    "fund_id": fund_id,
                    "transaction_date": transaction_date,
                    "withdrawal": withdrawal,
                    "deposit": deposit,
                    "balance_after": balance_after if balance_after != 0 else None,
                    "description": self._clean_text(mapped.get("description")),
                    "counterparty": self._clean_text(mapped.get("counterparty")),
                    "bank_branch": self._clean_text(mapped.get("bank_branch")),
                    "account_number": self._clean_text(mapped.get("account_number")),
                    "year_month": transaction_date.strftime("%Y-%m"),
                }
            )

        return rows

    def parse_excel(self, file_path: str, fund_id: int) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        workbook = load_workbook(file_path, data_only=True)

        for sheet in workbook.worksheets:
            header = self._find_header(sheet)
            if not header:
                continue
            header_row, index_map = header

            empty_streak = 0
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                raw_date = sheet.cell(row=row_idx, column=index_map["transaction_date"]).value
                transaction_date = self._parse_datetime(raw_date)

                raw_withdrawal = sheet.cell(row=row_idx, column=index_map["withdrawal"]).value
                raw_deposit = sheet.cell(row=row_idx, column=index_map["deposit"]).value
                raw_balance = (
                    sheet.cell(row=row_idx, column=index_map["balance_after"]).value
                    if "balance_after" in index_map
                    else None
                )

                if transaction_date is None and (raw_withdrawal in (None, "")) and (raw_deposit in (None, "")):
                    empty_streak += 1
                    if empty_streak >= 3:
                        break
                    continue
                empty_streak = 0

                if transaction_date is None:
                    continue

                withdrawal = self._parse_amount(raw_withdrawal)
                deposit = self._parse_amount(raw_deposit)
                balance_after = self._parse_amount(raw_balance)

                if withdrawal == 0 and deposit == 0 and balance_after == 0:
                    continue

                description = self._cell_value(sheet, row_idx, index_map.get("description"))
                counterparty = self._cell_value(sheet, row_idx, index_map.get("counterparty"))
                bank_branch = self._cell_value(sheet, row_idx, index_map.get("bank_branch"))
                account_number = self._extract_account_number(sheet.title)

                rows.append(
                    {
                        "fund_id": fund_id,
                        "transaction_date": transaction_date,
                        "withdrawal": withdrawal,
                        "deposit": deposit,
                        "balance_after": balance_after if balance_after != 0 else None,
                        "description": self._clean_text(description),
                        "counterparty": self._clean_text(counterparty),
                        "bank_branch": self._clean_text(bank_branch),
                        "account_number": account_number,
                        "year_month": transaction_date.strftime("%Y-%m"),
                    }
                )

        return rows

    def _split_line(self, line: str) -> list[str]:
        if "\t" in line:
            return [self._clean_text(v) for v in line.split("\t")]
        if "|" in line:
            return [self._clean_text(v) for v in line.split("|")]

        # Fallback: split by 2+ spaces.
        return [self._clean_text(v) for v in re.split(r"\s{2,}", line)]

    def _map_columns(self, parts: list[str]) -> dict[str, str | None]:
        values = [self._clean_text(part) for part in parts]

        # Some exports start with row number.
        if values and values[0].isdigit() and len(values) >= 8:
            values = values[1:]

        while len(values) < 8:
            values.append(None)

        return {
            "transaction_date": values[0],
            "withdrawal": values[1],
            "deposit": values[2],
            "balance_after": values[3],
            "description": values[4],
            "counterparty": values[5],
            "bank_branch": values[6],
            "account_number": values[7],
        }

    def _find_header(self, sheet) -> tuple[int, dict[str, int]] | None:
        for row_idx in range(1, min(sheet.max_row, 40) + 1):
            text_by_col: dict[int, str] = {}
            for col_idx in range(1, min(sheet.max_column, 30) + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                text_by_col[col_idx] = text

            if not text_by_col:
                continue

            index_map: dict[str, int] = {}
            for col_idx, text in text_by_col.items():
                normalized = text.replace(" ", "")
                if "거래일" in normalized:
                    index_map["transaction_date"] = col_idx
                elif "출금" in normalized:
                    index_map["withdrawal"] = col_idx
                elif "입금" in normalized:
                    index_map["deposit"] = col_idx
                elif "잔액" in normalized:
                    index_map["balance_after"] = col_idx
                elif "거래내용" in normalized:
                    index_map["description"] = col_idx
                elif "거래기록" in normalized or "거래처" in normalized:
                    index_map["counterparty"] = col_idx
                elif "거래점" in normalized or "지점" in normalized:
                    index_map["bank_branch"] = col_idx

            required = {"transaction_date", "withdrawal", "deposit"}
            if required.issubset(index_map.keys()):
                return row_idx, index_map

        return None

    def _cell_value(self, sheet, row_idx: int, col_idx: int | None) -> str | None:
        if not col_idx:
            return None
        value = sheet.cell(row=row_idx, column=col_idx).value
        if value is None:
            return None
        return str(value)

    def _parse_datetime(self, value: Any) -> dt.datetime | None:
        if value is None:
            return None
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return dt.datetime.combine(value, dt.time.min)

        text = str(value).strip()
        if not text:
            return None

        fmts = [
            "%Y/%m/%d %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y.%m.%d %H:%M:%S",
            "%Y/%m/%d",
            "%Y-%m-%d",
            "%Y.%m.%d",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y",
        ]
        for fmt in fmts:
            try:
                parsed = dt.datetime.strptime(text, fmt)
                return parsed
            except ValueError:
                continue

        # Excel serial date fallback.
        if text.isdigit():
            serial = int(text)
            try:
                return dt.datetime(1899, 12, 30) + dt.timedelta(days=serial)
            except ValueError:
                return None

        return None

    def _parse_amount(self, value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()
        if not text:
            return 0.0

        multiplier = 1.0
        if text[-1:].upper() == "K":
            multiplier = 1_000.0
            text = text[:-1]
        elif text[-1:].upper() == "M":
            multiplier = 1_000_000.0
            text = text[:-1]
        elif text[-1:].upper() == "B":
            multiplier = 1_000_000_000.0
            text = text[:-1]

        cleaned = _AMOUNT_RE.sub("", text)
        if cleaned in {"", "-", "."}:
            return 0.0

        try:
            return float(cleaned) * multiplier
        except ValueError:
            return 0.0

    def _extract_account_number(self, sheet_name: str) -> str | None:
        text = Path(sheet_name).name
        match = re.search(r"(\d{3,})", text)
        if not match:
            return None
        return match.group(1)

    def _clean_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        return text
