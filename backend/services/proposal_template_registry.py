from __future__ import annotations

import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.proposal_template import (
    ProposalTemplate,
    ProposalTemplateFieldMapping,
    ProposalTemplateSheet,
    ProposalTemplateTableMapping,
    ProposalTemplateValidationRule,
    ProposalTemplateVersion,
)
from schemas.proposal_template import (
    ProposalTemplateCreate,
    ProposalTemplateFieldMappingInput,
    ProposalTemplateSheetInput,
    ProposalTemplateTableMappingInput,
    ProposalTemplateUpdate,
    ProposalTemplateValidationRuleInput,
    ProposalTemplateVersionCloneRequest,
    ProposalTemplateVersionCreate,
    ProposalTemplateVersionHeaderUpdate,
    ProposalTemplateVersionRegistryUpdate,
)


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _serialize_json(value: Any) -> str | None:
    if value is None:
        return None
    return json.dumps(value, ensure_ascii=False)


def _json_load(raw: str | None, fallback: Any) -> Any:
    if not raw:
        return fallback
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return fallback


def _sheet_code_from_name(sheet_name: str, index: int, used_codes: set[str]) -> str:
    normalized = re.sub(r"[^0-9A-Za-z]+", "-", sheet_name.strip().lower()).strip("-")
    if not normalized:
        normalized = f"sheet-{index}"
    candidate = normalized
    suffix = 2
    while candidate in used_codes:
        candidate = f"{normalized}-{suffix}"
        suffix += 1
    used_codes.add(candidate)
    return candidate


def _workbook_sheet_inputs(source_path: str) -> list[ProposalTemplateSheetInput]:
    workbook_path = Path(source_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {source_path}")
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=False)
    try:
        sheet_inputs: list[ProposalTemplateSheetInput] = []
        used_codes: set[str] = set()
        for index, sheet_name in enumerate(workbook.sheetnames, start=1):
            sheet_inputs.append(
                ProposalTemplateSheetInput(
                    sheet_code=_sheet_code_from_name(sheet_name, index, used_codes),
                    sheet_name=sheet_name,
                    sheet_kind="table",
                    display_order=index - 1,
                    is_required=True,
                )
            )
        return sheet_inputs
    finally:
        workbook.close()


def _version_summary(db: Session, version: ProposalTemplateVersion) -> dict[str, Any]:
    sheet_count = db.query(ProposalTemplateSheet).filter(ProposalTemplateSheet.template_version_id == version.id).count()
    field_mapping_count = (
        db.query(ProposalTemplateFieldMapping)
        .filter(ProposalTemplateFieldMapping.template_version_id == version.id)
        .count()
    )
    table_mapping_count = (
        db.query(ProposalTemplateTableMapping)
        .filter(ProposalTemplateTableMapping.template_version_id == version.id)
        .count()
    )
    validation_rule_count = (
        db.query(ProposalTemplateValidationRule)
        .filter(ProposalTemplateValidationRule.template_version_id == version.id)
        .count()
    )
    return {
        "id": version.id,
        "template_id": version.template_id,
        "version_label": version.version_label,
        "status": version.status,
        "source_path": version.source_path,
        "source_filename": version.source_filename,
        "effective_from": version.effective_from,
        "effective_to": version.effective_to,
        "notes": version.notes,
        "sheet_count": sheet_count,
        "field_mapping_count": field_mapping_count,
        "table_mapping_count": table_mapping_count,
        "validation_rule_count": validation_rule_count,
        "created_at": version.created_at,
        "updated_at": version.updated_at,
    }


def _template_summary(db: Session, template: ProposalTemplate) -> dict[str, Any]:
    versions = (
        db.query(ProposalTemplateVersion)
        .filter(ProposalTemplateVersion.template_id == template.id)
        .order_by(ProposalTemplateVersion.id.desc())
        .all()
    )
    active_version = next((row for row in versions if row.status == "active"), None)
    return {
        "id": template.id,
        "code": template.code,
        "name": template.name,
        "institution_type": template.institution_type,
        "legacy_template_type": template.legacy_template_type,
        "description": template.description,
        "output_format": template.output_format,
        "source_family": template.source_family,
        "is_active": template.is_active,
        "version_count": len(versions),
        "active_version_id": active_version.id if active_version is not None else None,
        "active_version_label": active_version.version_label if active_version is not None else None,
        "created_at": template.created_at,
        "updated_at": template.updated_at,
    }


def list_proposal_templates(db: Session) -> list[dict[str, Any]]:
    rows = db.query(ProposalTemplate).order_by(ProposalTemplate.name.asc(), ProposalTemplate.id.asc()).all()
    return [_template_summary(db, row) for row in rows]


def get_proposal_template_detail(db: Session, template: ProposalTemplate) -> dict[str, Any]:
    payload = _template_summary(db, template)
    versions = (
        db.query(ProposalTemplateVersion)
        .filter(ProposalTemplateVersion.template_id == template.id)
        .order_by(ProposalTemplateVersion.created_at.desc(), ProposalTemplateVersion.id.desc())
        .all()
    )
    payload["versions"] = [_version_summary(db, row) for row in versions]
    return payload


def _sheet_payload(sheet: ProposalTemplateSheet) -> dict[str, Any]:
    return {
        "id": sheet.id,
        "sheet_code": sheet.sheet_code,
        "sheet_name": sheet.sheet_name,
        "sheet_kind": sheet.sheet_kind,
        "display_order": sheet.display_order,
        "is_required": sheet.is_required,
        "notes": sheet.notes,
        "created_at": sheet.created_at,
        "updated_at": sheet.updated_at,
    }


def _version_registry_snapshot(db: Session, version: ProposalTemplateVersion) -> dict[str, list[dict[str, Any]]]:
    sheets = (
        db.query(ProposalTemplateSheet)
        .filter(ProposalTemplateSheet.template_version_id == version.id)
        .order_by(ProposalTemplateSheet.display_order.asc(), ProposalTemplateSheet.id.asc())
        .all()
    )
    sheet_by_id = {sheet.id: sheet for sheet in sheets}
    field_mappings = (
        db.query(ProposalTemplateFieldMapping)
        .filter(ProposalTemplateFieldMapping.template_version_id == version.id)
        .order_by(ProposalTemplateFieldMapping.display_order.asc(), ProposalTemplateFieldMapping.id.asc())
        .all()
    )
    table_mappings = (
        db.query(ProposalTemplateTableMapping)
        .filter(ProposalTemplateTableMapping.template_version_id == version.id)
        .order_by(ProposalTemplateTableMapping.id.asc())
        .all()
    )
    validation_rules = (
        db.query(ProposalTemplateValidationRule)
        .filter(ProposalTemplateValidationRule.template_version_id == version.id)
        .order_by(ProposalTemplateValidationRule.id.asc())
        .all()
    )
    return {
        "sheets": [_sheet_payload(sheet) for sheet in sheets],
        "field_mappings": [
            {
                "id": row.id,
                "sheet_code": sheet_by_id[row.sheet_id].sheet_code,
                "field_key": row.field_key,
                "target_cell": row.target_cell,
                "value_source": row.value_source,
                "transform_rule": row.transform_rule,
                "default_value": _json_load(row.default_value_json, None),
                "source_note_hint": row.source_note_hint,
                "is_required": row.is_required,
                "display_order": row.display_order,
                "created_at": row.created_at,
                "updated_at": row.updated_at,
            }
            for row in field_mappings
        ],
        "table_mappings": [
            {
                "id": row.id,
                "sheet_code": sheet_by_id[row.sheet_id].sheet_code,
                "table_key": row.table_key,
                "start_cell": row.start_cell,
                "row_source": row.row_source,
                "columns": _json_load(row.columns_json, []),
                "row_key_field": row.row_key_field,
                "append_mode": row.append_mode,
                "max_rows": row.max_rows,
                "notes": row.notes,
                "created_at": row.created_at,
                "updated_at": row.updated_at,
            }
            for row in table_mappings
        ],
        "validation_rules": [
            {
                "id": row.id,
                "sheet_code": sheet_by_id[row.sheet_id].sheet_code if row.sheet_id is not None else None,
                "rule_code": row.rule_code,
                "rule_type": row.rule_type,
                "severity": row.severity,
                "target_ref": row.target_ref,
                "rule_payload": _json_load(row.rule_payload_json, {}),
                "message": row.message,
                "created_at": row.created_at,
                "updated_at": row.updated_at,
            }
            for row in validation_rules
        ],
    }


def get_proposal_template_version_detail(db: Session, version: ProposalTemplateVersion) -> dict[str, Any]:
    template = db.get(ProposalTemplate, version.template_id)
    if template is None:
        raise ValueError("Template not found for version.")
    snapshot = _version_registry_snapshot(db, version)

    payload = _version_summary(db, version)
    payload["template_code"] = template.code
    payload["template_name"] = template.name
    payload["sheets"] = snapshot["sheets"]
    payload["field_mappings"] = snapshot["field_mappings"]
    payload["table_mappings"] = snapshot["table_mappings"]
    payload["validation_rules"] = snapshot["validation_rules"]
    return payload


def create_proposal_template(db: Session, data: ProposalTemplateCreate) -> ProposalTemplate:
    existing = db.query(ProposalTemplate).filter(ProposalTemplate.code == data.code).first()
    if existing is not None:
        raise ValueError(f"Template code already exists: {data.code}")
    template = ProposalTemplate(
        code=data.code,
        name=data.name,
        institution_type=data.institution_type,
        legacy_template_type=data.legacy_template_type,
        description=data.description,
        output_format=data.output_format,
        source_family=data.source_family,
        is_active=data.is_active,
    )
    db.add(template)
    db.flush()
    return template


def update_proposal_template(db: Session, template: ProposalTemplate, data: ProposalTemplateUpdate) -> ProposalTemplate:
    if data.name is not None:
        template.name = data.name
    if "institution_type" in data.model_fields_set:
        template.institution_type = data.institution_type
    if "legacy_template_type" in data.model_fields_set:
        template.legacy_template_type = data.legacy_template_type
    if "description" in data.model_fields_set:
        template.description = data.description
    if data.output_format is not None:
        template.output_format = data.output_format
    if data.source_family is not None:
        template.source_family = data.source_family
    if data.is_active is not None:
        template.is_active = data.is_active
    template.updated_at = _utcnow()
    db.flush()
    return template


def _resolve_sheet_inputs(data: ProposalTemplateVersionCreate) -> list[ProposalTemplateSheetInput]:
    if data.sheets:
        used_codes: set[str] = set()
        resolved_inputs: list[ProposalTemplateSheetInput] = []
        for index, item in enumerate(data.sheets, start=1):
            sheet_code = item.sheet_code or _sheet_code_from_name(item.sheet_name, index, used_codes)
            if sheet_code in used_codes and item.sheet_code:
                raise ValueError(f"Duplicate sheet_code: {sheet_code}")
            used_codes.add(sheet_code)
            resolved_inputs.append(
                ProposalTemplateSheetInput(
                    sheet_code=sheet_code,
                    sheet_name=item.sheet_name,
                    sheet_kind=item.sheet_kind,
                    display_order=item.display_order,
                    is_required=item.is_required,
                    notes=item.notes,
                )
            )
        return resolved_inputs
    if data.source_path and data.import_workbook_sheets:
        return _workbook_sheet_inputs(data.source_path)
    raise ValueError("At least one sheet definition or a readable workbook path is required.")


def _create_field_mappings(
    db: Session,
    *,
    version: ProposalTemplateVersion,
    sheet_id_by_code: dict[str, int],
    field_mappings: list[ProposalTemplateFieldMappingInput],
) -> None:
    for row in field_mappings:
        sheet_id = sheet_id_by_code.get(row.sheet_code)
        if sheet_id is None:
            raise ValueError(f"Unknown sheet_code in field mapping: {row.sheet_code}")
        db.add(
            ProposalTemplateFieldMapping(
                template_version_id=version.id,
                sheet_id=sheet_id,
                field_key=row.field_key,
                target_cell=row.target_cell,
                value_source=row.value_source,
                transform_rule=row.transform_rule,
                default_value_json=_serialize_json(row.default_value),
                source_note_hint=row.source_note_hint,
                is_required=row.is_required,
                display_order=row.display_order,
            )
        )


def _create_table_mappings(
    db: Session,
    *,
    version: ProposalTemplateVersion,
    sheet_id_by_code: dict[str, int],
    table_mappings: list[ProposalTemplateTableMappingInput],
) -> None:
    for row in table_mappings:
        sheet_id = sheet_id_by_code.get(row.sheet_code)
        if sheet_id is None:
            raise ValueError(f"Unknown sheet_code in table mapping: {row.sheet_code}")
        db.add(
            ProposalTemplateTableMapping(
                template_version_id=version.id,
                sheet_id=sheet_id,
                table_key=row.table_key,
                start_cell=row.start_cell,
                row_source=row.row_source,
                columns_json=_serialize_json([column.model_dump() for column in row.columns]) or "[]",
                row_key_field=row.row_key_field,
                append_mode=row.append_mode,
                max_rows=row.max_rows,
                notes=row.notes,
            )
        )


def _create_validation_rules(
    db: Session,
    *,
    version: ProposalTemplateVersion,
    sheet_id_by_code: dict[str, int],
    validation_rules: list[ProposalTemplateValidationRuleInput],
) -> None:
    for row in validation_rules:
        sheet_id = None
        if row.sheet_code is not None:
            sheet_id = sheet_id_by_code.get(row.sheet_code)
            if sheet_id is None:
                raise ValueError(f"Unknown sheet_code in validation rule: {row.sheet_code}")
        db.add(
            ProposalTemplateValidationRule(
                template_version_id=version.id,
                sheet_id=sheet_id,
                rule_code=row.rule_code,
                rule_type=row.rule_type,
                severity=row.severity,
                target_ref=row.target_ref,
                rule_payload_json=_serialize_json(row.rule_payload) or "{}",
                message=row.message,
            )
        )


def create_proposal_template_version(
    db: Session,
    *,
    template: ProposalTemplate,
    data: ProposalTemplateVersionCreate,
) -> ProposalTemplateVersion:
    duplicate = (
        db.query(ProposalTemplateVersion)
        .filter(
            ProposalTemplateVersion.template_id == template.id,
            ProposalTemplateVersion.version_label == data.version_label,
        )
        .first()
    )
    if duplicate is not None:
        raise ValueError(f"Version already exists: {data.version_label}")

    sheet_inputs = _resolve_sheet_inputs(data)
    source_filename = Path(data.source_path).name if data.source_path else None
    version = ProposalTemplateVersion(
        template_id=template.id,
        version_label=data.version_label,
        status=data.status,
        source_path=data.source_path,
        source_filename=source_filename,
        effective_from=data.effective_from,
        effective_to=data.effective_to,
        notes=data.notes,
    )
    db.add(version)
    db.flush()

    sheet_id_by_code: dict[str, int] = {}
    for item in sheet_inputs:
        sheet = ProposalTemplateSheet(
            template_version_id=version.id,
            sheet_code=item.sheet_code or "",
            sheet_name=item.sheet_name,
            sheet_kind=item.sheet_kind,
            display_order=item.display_order,
            is_required=item.is_required,
            notes=item.notes,
        )
        db.add(sheet)
        db.flush()
        sheet_id_by_code[sheet.sheet_code] = sheet.id

    _create_field_mappings(db, version=version, sheet_id_by_code=sheet_id_by_code, field_mappings=data.field_mappings)
    _create_table_mappings(db, version=version, sheet_id_by_code=sheet_id_by_code, table_mappings=data.table_mappings)
    _create_validation_rules(
        db,
        version=version,
        sheet_id_by_code=sheet_id_by_code,
        validation_rules=data.validation_rules,
    )

    if data.status == "active":
        activate_proposal_template_version(db, version=version)
    else:
        version.updated_at = _utcnow()
    db.flush()
    return version


def update_proposal_template_version_header(
    db: Session,
    *,
    version: ProposalTemplateVersion,
    data: ProposalTemplateVersionHeaderUpdate,
) -> ProposalTemplateVersion:
    if data.version_label is not None and data.version_label != version.version_label:
        duplicate = (
            db.query(ProposalTemplateVersion)
            .filter(
                ProposalTemplateVersion.template_id == version.template_id,
                ProposalTemplateVersion.version_label == data.version_label,
                ProposalTemplateVersion.id != version.id,
            )
            .first()
        )
        if duplicate is not None:
            raise ValueError(f"Version already exists: {data.version_label}")
        version.version_label = data.version_label
    if "source_path" in data.model_fields_set:
        version.source_path = data.source_path
        version.source_filename = Path(data.source_path).name if data.source_path else None
    if "effective_from" in data.model_fields_set:
        version.effective_from = data.effective_from
    if "effective_to" in data.model_fields_set:
        version.effective_to = data.effective_to
    if "notes" in data.model_fields_set:
        version.notes = data.notes
    if data.status is not None and data.status != version.status:
        if data.status == "active":
            activate_proposal_template_version(db, version=version)
        else:
            version.status = data.status
    version.updated_at = _utcnow()
    db.flush()
    return version


def clone_proposal_template_version(
    db: Session,
    *,
    source_version: ProposalTemplateVersion,
    data: ProposalTemplateVersionCloneRequest,
) -> ProposalTemplateVersion:
    template = db.get(ProposalTemplate, source_version.template_id)
    if template is None:
        raise ValueError("Template not found for version.")
    snapshot = _version_registry_snapshot(db, source_version)
    clone_data = ProposalTemplateVersionCreate(
        version_label=data.version_label,
        status=data.status,
        source_path=data.source_path if "source_path" in data.model_fields_set else source_version.source_path,
        effective_from=data.effective_from if "effective_from" in data.model_fields_set else source_version.effective_from,
        effective_to=data.effective_to if "effective_to" in data.model_fields_set else source_version.effective_to,
        notes=data.notes if "notes" in data.model_fields_set else source_version.notes,
        import_workbook_sheets=False,
        sheets=[
            ProposalTemplateSheetInput(
                sheet_code=row["sheet_code"],
                sheet_name=row["sheet_name"],
                sheet_kind=row["sheet_kind"],
                display_order=row["display_order"],
                is_required=row["is_required"],
                notes=row["notes"],
            )
            for row in snapshot["sheets"]
        ],
        field_mappings=[
            ProposalTemplateFieldMappingInput(
                sheet_code=row["sheet_code"],
                field_key=row["field_key"],
                target_cell=row["target_cell"],
                value_source=row["value_source"],
                transform_rule=row["transform_rule"],
                default_value=row["default_value"],
                source_note_hint=row["source_note_hint"],
                is_required=row["is_required"],
                display_order=row["display_order"],
            )
            for row in snapshot["field_mappings"]
        ],
        table_mappings=[
            ProposalTemplateTableMappingInput(
                sheet_code=row["sheet_code"],
                table_key=row["table_key"],
                start_cell=row["start_cell"],
                row_source=row["row_source"],
                columns=row["columns"],
                row_key_field=row["row_key_field"],
                append_mode=row["append_mode"],
                max_rows=row["max_rows"],
                notes=row["notes"],
            )
            for row in snapshot["table_mappings"]
        ],
        validation_rules=[
            ProposalTemplateValidationRuleInput(
                sheet_code=row["sheet_code"],
                rule_code=row["rule_code"],
                rule_type=row["rule_type"],
                severity=row["severity"],
                target_ref=row["target_ref"],
                rule_payload=row["rule_payload"],
                message=row["message"],
            )
            for row in snapshot["validation_rules"]
        ],
    )
    return create_proposal_template_version(db, template=template, data=clone_data)


def replace_proposal_template_version_registry(
    db: Session,
    *,
    version: ProposalTemplateVersion,
    data: ProposalTemplateVersionRegistryUpdate,
) -> ProposalTemplateVersion:
    if not data.sheets:
        raise ValueError("At least one sheet is required.")

    temp_create = ProposalTemplateVersionCreate(
        version_label=version.version_label,
        status=version.status,
        source_path=version.source_path,
        effective_from=version.effective_from,
        effective_to=version.effective_to,
        notes=version.notes,
        import_workbook_sheets=False,
        sheets=data.sheets,
        field_mappings=data.field_mappings,
        table_mappings=data.table_mappings,
        validation_rules=data.validation_rules,
    )
    sheet_inputs = _resolve_sheet_inputs(temp_create)

    (
        db.query(ProposalTemplateValidationRule)
        .filter(ProposalTemplateValidationRule.template_version_id == version.id)
        .delete(synchronize_session=False)
    )
    (
        db.query(ProposalTemplateTableMapping)
        .filter(ProposalTemplateTableMapping.template_version_id == version.id)
        .delete(synchronize_session=False)
    )
    (
        db.query(ProposalTemplateFieldMapping)
        .filter(ProposalTemplateFieldMapping.template_version_id == version.id)
        .delete(synchronize_session=False)
    )
    (
        db.query(ProposalTemplateSheet)
        .filter(ProposalTemplateSheet.template_version_id == version.id)
        .delete(synchronize_session=False)
    )

    sheet_id_by_code: dict[str, int] = {}
    for item in sheet_inputs:
        sheet = ProposalTemplateSheet(
            template_version_id=version.id,
            sheet_code=item.sheet_code or "",
            sheet_name=item.sheet_name,
            sheet_kind=item.sheet_kind,
            display_order=item.display_order,
            is_required=item.is_required,
            notes=item.notes,
        )
        db.add(sheet)
        db.flush()
        sheet_id_by_code[sheet.sheet_code] = sheet.id

    _create_field_mappings(db, version=version, sheet_id_by_code=sheet_id_by_code, field_mappings=data.field_mappings)
    _create_table_mappings(db, version=version, sheet_id_by_code=sheet_id_by_code, table_mappings=data.table_mappings)
    _create_validation_rules(
        db,
        version=version,
        sheet_id_by_code=sheet_id_by_code,
        validation_rules=data.validation_rules,
    )

    version.updated_at = _utcnow()
    db.flush()
    return version


def _diff_key(row: dict[str, Any], key_fields: tuple[str, ...]) -> str:
    values = []
    for field in key_fields:
        value = row.get(field)
        values.append("" if value is None else str(value))
    return ":".join(values)


def _normalize_diff_row(row: dict[str, Any], ignored_fields: set[str]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if key not in ignored_fields}


def _diff_rows(
    *,
    base_rows: list[dict[str, Any]],
    target_rows: list[dict[str, Any]],
    key_fields: tuple[str, ...],
) -> list[dict[str, Any]]:
    ignored_fields = {"id", "created_at", "updated_at"}
    base_by_key = {_diff_key(row, key_fields): row for row in base_rows}
    target_by_key = {_diff_key(row, key_fields): row for row in target_rows}
    all_keys = sorted(set(base_by_key) | set(target_by_key))

    changes: list[dict[str, Any]] = []
    for key in all_keys:
        before = base_by_key.get(key)
        after = target_by_key.get(key)
        if before is None:
            changes.append(
                {
                    "key": key,
                    "sheet_code": after.get("sheet_code"),
                    "change_type": "added",
                    "changed_fields": [],
                    "before": None,
                    "after": _normalize_diff_row(after, ignored_fields),
                }
            )
            continue
        if after is None:
            changes.append(
                {
                    "key": key,
                    "sheet_code": before.get("sheet_code"),
                    "change_type": "removed",
                    "changed_fields": [],
                    "before": _normalize_diff_row(before, ignored_fields),
                    "after": None,
                }
            )
            continue

        before_payload = _normalize_diff_row(before, ignored_fields)
        after_payload = _normalize_diff_row(after, ignored_fields)
        changed_fields = sorted(
            field
            for field in set(before_payload) | set(after_payload)
            if before_payload.get(field) != after_payload.get(field)
        )
        if changed_fields:
            changes.append(
                {
                    "key": key,
                    "sheet_code": after.get("sheet_code") or before.get("sheet_code"),
                    "change_type": "modified",
                    "changed_fields": changed_fields,
                    "before": before_payload,
                    "after": after_payload,
                }
            )
    return changes


def compare_proposal_template_versions(
    db: Session,
    *,
    base_version: ProposalTemplateVersion,
    target_version: ProposalTemplateVersion,
) -> dict[str, Any]:
    if base_version.template_id != target_version.template_id:
        raise ValueError("Only versions of the same template can be compared.")

    base_snapshot = _version_registry_snapshot(db, base_version)
    target_snapshot = _version_registry_snapshot(db, target_version)

    sheet_changes = _diff_rows(
        base_rows=base_snapshot["sheets"],
        target_rows=target_snapshot["sheets"],
        key_fields=("sheet_code",),
    )
    field_mapping_changes = _diff_rows(
        base_rows=base_snapshot["field_mappings"],
        target_rows=target_snapshot["field_mappings"],
        key_fields=("sheet_code", "field_key"),
    )
    table_mapping_changes = _diff_rows(
        base_rows=base_snapshot["table_mappings"],
        target_rows=target_snapshot["table_mappings"],
        key_fields=("sheet_code", "table_key"),
    )
    validation_rule_changes = _diff_rows(
        base_rows=base_snapshot["validation_rules"],
        target_rows=target_snapshot["validation_rules"],
        key_fields=("sheet_code", "rule_code"),
    )

    changed_sheet_codes = sorted(
        {
            sheet_code
            for change_group in (
                sheet_changes,
                field_mapping_changes,
                table_mapping_changes,
                validation_rule_changes,
            )
            for row in change_group
            for sheet_code in [row.get("sheet_code")]
            if sheet_code
        }
    )

    return {
        "base_version_id": base_version.id,
        "base_version_label": base_version.version_label,
        "target_version_id": target_version.id,
        "target_version_label": target_version.version_label,
        "sheet_changes": sheet_changes,
        "field_mapping_changes": field_mapping_changes,
        "table_mapping_changes": table_mapping_changes,
        "validation_rule_changes": validation_rule_changes,
        "changed_sheet_codes": changed_sheet_codes,
    }


def activate_proposal_template_version(db: Session, *, version: ProposalTemplateVersion) -> ProposalTemplateVersion:
    siblings = (
        db.query(ProposalTemplateVersion)
        .filter(
            ProposalTemplateVersion.template_id == version.template_id,
            ProposalTemplateVersion.id != version.id,
            ProposalTemplateVersion.status == "active",
        )
        .all()
    )
    for row in siblings:
        row.status = "archived"
        row.updated_at = _utcnow()
    version.status = "active"
    version.updated_at = _utcnow()
    db.flush()
    return version
